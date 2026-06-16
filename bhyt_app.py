import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from datetime import datetime
import os
import sys
import importlib
import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

# Import excel_compare
try:
    from excel_compare import render_excel_compare_tab
except ImportError as e:
    def render_excel_compare_tab():
        st.error(f"Không tải được excel_compare: {e}")

# Import bhyt_rules - từng hàm riêng để dễ phát hiện lỗi
try:
    import bhyt_rules as _br
    # Buộc reload để tránh cache module cũ trên Streamlit Cloud
    importlib.reload(_br)

    check_hanh_chinh_full            = _br.check_hanh_chinh_full
    check_nhom2_danh_muc_gia         = _br.check_nhom2_danh_muc_gia
    check_nhom3_hop_ly_chi_dinh      = _br.check_nhom3_hop_ly_chi_dinh
    _to_datetime                     = _br._to_datetime
    load_default_danh_muc            = _br.load_default_danh_muc
    merge_danh_muc                   = _br.merge_danh_muc
    rebuild_ma_thuoc_hoatchat_from_source = _br.rebuild_ma_thuoc_hoatchat_from_source
    rebuild_gia_dvkt_from_source     = _br.rebuild_gia_dvkt_from_source

    # Kiểm tra tất cả hàm tồn tại
    _MISSING = [name for name in [
        "check_hanh_chinh_full", "check_nhom2_danh_muc_gia",
        "check_nhom3_hop_ly_chi_dinh", "_to_datetime",
        "load_default_danh_muc", "merge_danh_muc",
        "rebuild_ma_thuoc_hoatchat_from_source", "rebuild_gia_dvkt_from_source"
    ] if not hasattr(_br, name)]
    if _MISSING:
        st.error(
            f"❌ bhyt_rules.py thiếu các hàm sau (cần upload lại file mới): "
            + ", ".join(_MISSING)
        )
        st.stop()

except Exception as _e:
    st.error(f"❌ Lỗi khi tải bhyt_rules.py: {_e}")
    st.info("Vui lòng đảm bảo đã upload đúng file bhyt_rules.py mới nhất lên GitHub.")
    st.stop()


st.set_page_config(page_title="Kiểm tra XML BHYT", layout="wide")

st.title("🔍 KIỂM TRA XML BHYT TRƯỚC KHI GỬI CỔNG GIÁM ĐỊNH")

# ==================== TẠO TAB ====================
tab1, tab2, tab3 = st.tabs(["📋 Kiểm tra XML", "📊 So sánh Excel", "🏥 Tra cứu TT06"])

# ==================== GLOBAL DEFINITIONS (FOR BOTH TABS) ====================

# Danh sách trường bắt buộc theo từng loại file (chuẩn 4210)
MANDATORY_FIELDS_PER_FILE = {
    "Hành chính": [
        'MA_LK', 'MA_BN', 'HO_TEN', 'NGAY_SINH', 'GIOI_TINH', 'MA_THE_BHYT',
        'MA_CSKCB', 'NGAY_VAO', 'NGAY_RA', 'MA_BENH_CHINH', 'MA_LOAI_KCB',
        'MUC_HUONG', 'TYLE_BHTT'
    ],
    "Thuốc": [
        'MA_LK', 'MA_THUOC', 'TEN_THUOC', 'SO_LUONG', 'DON_GIA',
        'THANH_TIEN_BH', 'NGAY_YL', 'NGAY_TH_YL'
    ],
    "Dịch vụ kỹ thuật": [
        'MA_LK', 'SO_LUONG', 'DON_GIA_BH',
        'THANH_TIEN_BH', 'NGAY_YL'
    ],
    "Mã máy xét nghiệm": ['MA_LK', 'MA_DICH_VU', 'TEN_CHI_SO', 'NGAY_KQ'],
    "Phẫu thuật thủ thuật": ['MA_LK', 'THOI_DIEM_DBLS']
}

FILE_TYPE_TO_KEY = {
    "Hành chính": "hanh_chinh",
    "Thuốc": "thuoc",
    "Dịch vụ kỹ thuật": "dvkt",
    "Mã máy xét nghiệm": "cdha",
    "Phẫu thuật thủ thuật": "pttt",
}

MUC_DO_COLOR = {
    "Lỗi": "FFC7CE",       # đỏ nhạt
    "Cảnh báo": "FFEB9C",  # vàng nhạt
}
MUC_DO_FONT_COLOR = {
    "Lỗi": "9C0006",
    "Cảnh báo": "9C6500",
}

# ==================== TAB 1: KIỂM TRA XML ====================
with tab1:
    st.markdown("### TTTYT Châu Thành | Công cụ kiểm tra dữ liệu BHYT trước khi gửi Cổng giám định")
    st.caption(
        "Áp dụng 4 nhóm quy tắc: (1) Hành chính & thẻ BHYT, "
        "(2) Danh mục thuốc/DVKT/VTYT - giá - định mức, "
        "(3) Hợp lý chỉ định y khoa (ICD-10, giới tính/tuổi, trùng lặp, liều dùng), "
        "(4) Thời gian KCB & trùng lặp đợt điều trị."
    )

    # ====================== ICD-10 ======================
    @st.cache_data
    def load_icd10_codes():
        """
        Nạp danh mục mã ICD-10 hợp lệ từ icd10.json (đặt cùng thư mục với
        bhyt_app.py). Cấu trúc file: {"A00": ["Bệnh tả", 4], "A00.0": [...], ...}
        Nhanh hơn đọc Excel và không cần upload mỗi lần.
        """
        try:
            if os.path.exists("icd10.json"):
                import json
                with open("icd10.json", encoding="utf-8") as f:
                    data = json.load(f)
                codes = set()
                for code in data.keys():
                    code = str(code).strip().upper()
                    codes.add(code)
                    if '.' in code:
                        codes.add(code.replace('.', ''))
                return codes
            return set()
        except Exception as e:
            st.error(f"Lỗi khi tải icd10.json: {e}")
            return set()

    # ====================== DANH MỤC BỔ SUNG (TÙY CHỌN) ======================
    @st.cache_data
    def load_gia_danh_muc(file_bytes, sheet_name=None, ma_col="MA", gia_col="GIA"):
        """Đọc 1 file Excel danh mục giá: 2 cột Mã - Giá (đơn vị đồng)."""
        result = {}
        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True)
            sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]
            try:
                idx_ma = headers.index(ma_col.upper())
                idx_gia = headers.index(gia_col.upper())
            except ValueError:
                return result
            for r in range(2, sheet.max_row + 1):
                ma = sheet.cell(r, idx_ma + 1).value
                gia = sheet.cell(r, idx_gia + 1).value
                if ma is None:
                    continue
                try:
                    result[str(ma).strip()] = float(gia)
                except (TypeError, ValueError):
                    continue
        except Exception:
            pass
        return result

    @st.cache_data
    def load_icd_chi_dinh(file_bytes):
        """
        Đọc danh mục đối chiếu ICD-10 <-> mã DVKT/Thuốc được phép chỉ định.
        Định dạng cột yêu cầu: MA_BENH | MA_DICH_VU (mỗi dòng 1 cặp).
        Trả về dict {MA_BENH: set(MA_DICH_VU,...)}
        """
        result = {}
        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]
            try:
                idx_benh = headers.index("MA_BENH")
                idx_dv = headers.index("MA_DICH_VU")
            except ValueError:
                return result
            for r in range(2, sheet.max_row + 1):
                ma_benh = sheet.cell(r, idx_benh + 1).value
                ma_dv = sheet.cell(r, idx_dv + 1).value
                if ma_benh is None or ma_dv is None:
                    continue
                ma_benh = str(ma_benh).strip().upper()
                ma_dv = str(ma_dv).strip()
                result.setdefault(ma_benh, set()).add(ma_dv)
        except Exception:
            pass
        return result

    # ====================== SIDEBAR ======================
    with st.sidebar:
        st.header("📋 Danh mục đối chiếu (tùy chọn)")
        st.caption("Tải các danh mục dưới đây để bật thêm các quy tắc kiểm tra chi tiết. "
                    "Nếu không tải, hệ thống vẫn kiểm tra các quy tắc logic cơ bản.")

        VALID_MA_BENH_CODES = load_icd10_codes()
        st.success(f"✅ Danh mục ICD-10 (icd10.json): {len(VALID_MA_BENH_CODES)} mã.")

        st.markdown("---")
        st.subheader("🔄 Cập nhật danh mục từ file gốc")
        st.caption(
            "Khi có danh mục thuốc/DVKT mới, chỉ cần tải lên file gốc (đầy đủ cột) "
            "ở đây - hệ thống tự động lọc và GHI ĐÈ danh mục đối chiếu đang dùng. "
            "Không cần tự tách file."
        )

        uploaded_file_thuoc_goc = st.file_uploader(
            "📄 File danh mục thuốc gốc (như FileDanhMucThuoc.xlsx - cần cột "
            "MA_THUOC, TEN_THUOC, TEN_HOAT_CHAT, TU_NGAY) - dùng để kiểm tra "
            "tương tác/chống chỉ định thuốc theo ICD-10",
            type=["xlsx"], key="file_thuoc_goc")
        if uploaded_file_thuoc_goc:
            try:
                from io import BytesIO as _BytesIO
                ket_qua = rebuild_ma_thuoc_hoatchat_from_source(_BytesIO(uploaded_file_thuoc_goc.getvalue()))
                st.success(f"✅ Đã cập nhật bảng tra hoạt chất ({ket_qua['so_ma_hoat_chat']} mã thuốc) từ file vừa tải lên.")
                st.cache_data.clear()
            except Exception as e:
                st.error(f"❌ Không thể cập nhật danh mục thuốc: {e}")

        uploaded_file_dvkt_goc = st.file_uploader(
            "📄 File danh mục DVKT/Giường gốc (như FileDichVuBV.xlsx - cần cột "
            "MA_TUONG_DUONG, DON_GIA, TUNGAY)",
            type=["xlsx"], key="file_dvkt_goc")
        if uploaded_file_dvkt_goc:
            try:
                from io import BytesIO as _BytesIO
                ket_qua = rebuild_gia_dvkt_from_source(_BytesIO(uploaded_file_dvkt_goc.getvalue()))
                st.success(f"✅ Đã cập nhật danh mục giá DVKT/Giường ({ket_qua['so_ma_gia_dvkt']} mã) từ file vừa tải lên.")
                st.cache_data.clear()
            except Exception as e:
                st.error(f"❌ Không thể cập nhật danh mục DVKT: {e}")

        st.markdown("---")
        st.caption(
            "✅ Đã tích hợp sẵn danh mục giá DVKT/Giường bệnh và bảng thuốc chống chỉ định "
            "theo ICD-10 của đơn vị (không cần tải lên mỗi lần). "
            "Các mục dưới đây chỉ dùng để BỔ SUNG/GHI ĐÈ thêm nếu cần."
        )
        uploaded_gia_dvkt = st.file_uploader(
            "2. (Tùy chọn) Bổ sung danh mục giá DVKT (cột MA, GIA)", type=["xlsx"], key="gia_dvkt")
        uploaded_icd_chidinh = st.file_uploader(
            "3. (Tùy chọn) Bảng đối chiếu ICD-10 ↔ chỉ định (cột MA_BENH, MA_DICH_VU)", type=["xlsx"], key="icd_chidinh")

        # Danh mục cố định kèm sẵn ứng dụng (giá DVKT/Giường, thuốc chống chỉ định theo ICD)
        DANH_MUC_MAC_DINH = load_default_danh_muc()
        st.success(
            f"📦 Danh mục mặc định: {len(DANH_MUC_MAC_DINH.get('gia_dvkt', {}))} mã giá DVKT/Giường, "
            f"{len(DANH_MUC_MAC_DINH.get('thuoc_chong_chi_dinh', []))} quy tắc thuốc chống chỉ định ICD-10, "
            f"{len(DANH_MUC_MAC_DINH.get('ma_thuoc_hoat_chat', {}))} mã thuốc có dữ liệu hoạt chất."
        )

        DANH_MUC_BO_SUNG = {}
        if uploaded_gia_dvkt:
            DANH_MUC_BO_SUNG["gia_dvkt"] = load_gia_danh_muc(uploaded_gia_dvkt.getvalue())
            st.success(f"✅ Bổ sung {len(DANH_MUC_BO_SUNG['gia_dvkt'])} mã giá DVKT")
        if uploaded_icd_chidinh:
            DANH_MUC_BO_SUNG["icd_chi_dinh"] = load_icd_chi_dinh(uploaded_icd_chidinh.getvalue())
            st.success(f"✅ Đã tải bảng đối chiếu ICD-10 ↔ chỉ định ({len(DANH_MUC_BO_SUNG['icd_chi_dinh'])} mã bệnh)")

        DANH_MUC = merge_danh_muc(DANH_MUC_MAC_DINH, DANH_MUC_BO_SUNG)

    # ====================== UPLOAD FILES ======================
    st.subheader("📤 Chọn 5 file cần kiểm tra (XML1 - XML5 theo chuẩn 4210)")
    cols = st.columns(5)
    files = [None] * 6
    file_types = ["", "Hành chính", "Thuốc", "Dịch vụ kỹ thuật", "Mã máy xét nghiệm", "Phẫu thuật thủ thuật"]

    for i in range(1, 6):
        with cols[i - 1]:
            st.markdown(f"**File {file_types[i]}**")
            files[i] = st.file_uploader(f"Chọn file {file_types[i]}", type=["xlsx", "xml"], key=f"f{i}_tab1")

    # ====================== HÀM ĐỌC FILE -> LIST HO_SO (XML ELEMENT) ======================
    def read_file_to_root(uploaded_file, file_type, all_errors):
        """Đọc file xlsx/xml và trả về root ET.Element <DATA><HO_SO_KCB>...</HO_SO_KCB></DATA>."""
        file_path = f"temp_{uploaded_file.name}"
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        root = None
        try:
            if uploaded_file.name.endswith('.xlsx'):
                wb = load_workbook(file_path)
                sheet = wb.active
                headers = [str(cell.value).strip().upper() if cell.value else "" for cell in sheet[1]]

                if not any(h for h in headers):
                    all_errors.append({
                        'ma_lk': "Không xác định",
                        'nhom': "Lỗi định dạng file",
                        'muc_do': "Lỗi",
                        'truong_loi': 'Headers',
                        'gia_tri': '',
                        'mo_ta_loi': f"File '{uploaded_file.name}' không có tiêu đề cột.",
                        'huong_xu_ly': "Kiểm tra lại dòng tiêu đề (dòng 1) của file Excel.",
                        'nguon_file': file_type
                    })
                    return None

                root = ET.Element("DATA")
                for r in range(2, sheet.max_row + 1):
                    ho_so = ET.SubElement(root, "HO_SO_KCB")
                    for idx, h in enumerate(headers):
                        if not h:
                            continue
                        val = sheet.cell(r, idx + 1).value
                        if isinstance(val, datetime):
                            val = val.strftime("%Y%m%d%H%M")
                        elif isinstance(val, (int, float)):
                            # Chỉ chuyển sang ngày-giờ nếu trông giống serial date của Excel
                            # và tên trường có liên quan đến NGAY/THOI_DIEM, tránh phá hỏng
                            # các trường số (đơn giá, số lượng...)
                            if ("NGAY" in h or "THOI_DIEM" in h) and 0 < val < 100000:
                                try:
                                    val = datetime.fromtimestamp((val - 25569) * 86400).strftime("%Y%m%d%H%M")
                                except Exception:
                                    pass
                            else:
                                val = val
                        elem = ET.SubElement(ho_so, h)
                        elem.text = str(val).strip() if val is not None else ""

            elif uploaded_file.name.endswith('.xml'):
                tree = ET.parse(file_path)
                root = tree.getroot()

            if root is None:
                all_errors.append({
                    'ma_lk': "Không xác định",
                    'nhom': "Lỗi đọc file",
                    'muc_do': "Lỗi",
                    'truong_loi': "File",
                    'gia_tri': '',
                    'mo_ta_loi': f"Không thể đọc hoặc phân tích file '{uploaded_file.name}'.",
                    'huong_xu_ly': "Kiểm tra lại định dạng/đuôi file đã chọn (.xlsx hoặc .xml).",
                    'nguon_file': file_type
                })
                return None

        except Exception as e:
            all_errors.append({
                'ma_lk': "Không xác định",
                'nhom': f"Lỗi xử lý file '{uploaded_file.name}'",
                'muc_do': "Lỗi",
                'truong_loi': "File",
                'gia_tri': '',
                'mo_ta_loi': f"Lỗi chung: {str(e)}",
                'huong_xu_ly': "Kiểm tra lại file đầu vào (định dạng, dữ liệu).",
                'nguon_file': file_type
            })
            return None
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)

        return root

    # ====================== HÀM KIỂM TRA TRƯỜNG BẮT BUỘC ======================
    def check_mandatory_fields(ho_so, ma_lk, file_type):
        errors = []
        for field in MANDATORY_FIELDS_PER_FILE.get(file_type, []):
            tag = ho_so.find(field)
            if tag is None or not (tag.text or "").strip():
                errors.append({
                    'ma_lk': ma_lk,
                    'nhom': "0. Trường bắt buộc",
                    'muc_do': "Lỗi",
                    'truong_loi': field,
                    'gia_tri': '(thiếu)',
                    'mo_ta_loi': f"Trường '{field}' bị thiếu hoặc rỗng - là trường bắt buộc của hồ sơ {file_type}.",
                    'huong_xu_ly': f"Bổ sung giá trị cho trường '{field}' trong file {file_type} trước khi gửi.",
                    'nguon_file': file_type
                })
        return errors

    def check_ngay_logic(ho_so, ma_lk, file_type):
        errors = []
        if file_type != "Hành chính":
            return errors
        ngay_vao_tag = ho_so.find('NGAY_VAO')
        ngay_ra_tag = ho_so.find('NGAY_RA')
        if ngay_vao_tag is not None and (ngay_vao_tag.text or "").strip() and \
           ngay_ra_tag is not None and (ngay_ra_tag.text or "").strip():
            v, r = ngay_vao_tag.text.strip(), ngay_ra_tag.text.strip()
            dt_v, dt_r = _to_datetime(v), _to_datetime(r)
            if dt_v is None or dt_r is None:
                errors.append({
                    'ma_lk': ma_lk,
                    'nhom': "0. Trường bắt buộc",
                    'muc_do': "Lỗi",
                    'truong_loi': 'NGAY_VAO/NGAY_RA',
                    'gia_tri': f"{v} / {r}",
                    'mo_ta_loi': "Ngày vào hoặc Ngày ra không đúng định dạng YYYYMMDDHHMM (12 số).",
                    'huong_xu_ly': "Sửa lại NGAY_VAO/NGAY_RA đúng định dạng năm-tháng-ngày-giờ-phút (12 số liên tiếp).",
                    'nguon_file': file_type
                })
            elif dt_r < dt_v:
                errors.append({
                    'ma_lk': ma_lk,
                    'nhom': "0. Trường bắt buộc",
                    'muc_do': "Lỗi",
                    'truong_loi': 'NGAY_RA',
                    'gia_tri': r,
                    'mo_ta_loi': f"Ngày ra ({r}) không thể nhỏ hơn Ngày vào ({v}).",
                    'huong_xu_ly': "Kiểm tra lại NGAY_VAO và NGAY_RA, đảm bảo Ngày ra >= Ngày vào.",
                    'nguon_file': file_type
                })
        return errors

    def check_ma_benh_icd10(ho_so, ma_lk, valid_codes):
        errors = []
        ma_benh_chinh_tag = ho_so.find('MA_BENH_CHINH')
        if ma_benh_chinh_tag is not None and (ma_benh_chinh_tag.text or "").strip():
            ma_benh_chinh = ma_benh_chinh_tag.text.strip().upper()
            if valid_codes and ma_benh_chinh not in valid_codes:
                errors.append({
                    'ma_lk': ma_lk,
                    'nhom': "0. Trường bắt buộc",
                    'muc_do': "Lỗi",
                    'truong_loi': 'MA_BENH_CHINH',
                    'gia_tri': ma_benh_chinh,
                    'mo_ta_loi': f"Mã bệnh chính '{ma_benh_chinh}' không hợp lệ theo danh mục ICD-10.",
                    'huong_xu_ly': "Tra lại mã ICD-10 đúng (kiểm tra chính tả, dấu chấm phân cấp) và sửa MA_BENH_CHINH.",
                    'nguon_file': "Hành chính"
                })
        return errors

    def check_thoi_diem_dbls(ho_so, ma_lk):
        errors = []
        tag = ho_so.find('THOI_DIEM_DBLS')
        if tag is not None and (tag.text or "").strip():
            if _to_datetime(tag.text.strip()) is None:
                errors.append({
                    'ma_lk': ma_lk,
                    'nhom': "0. Trường bắt buộc",
                    'muc_do': "Lỗi",
                    'truong_loi': 'THOI_DIEM_DBLS',
                    'gia_tri': tag.text.strip(),
                    'mo_ta_loi': "Thời điểm bắt đầu phẫu thuật/thủ thuật không đúng định dạng YYYYMMDDHHMM.",
                    'huong_xu_ly': "Sửa lại THOI_DIEM_DBLS đúng định dạng 12 số (năm-tháng-ngày-giờ-phút).",
                    'nguon_file': "Phẫu thuật thủ thuật"
                })
        return errors

    # ====================== NÚT KIỂM TRA ======================
    if st.button("🚀 BẮT ĐẦU KIỂM TRA TẤT CẢ", type="primary", use_container_width=True):
        VALID_MA_BENH_CODES = load_icd10_codes()
        all_errors = []  # list lỗi tổng, mỗi lỗi có key chuẩn ở trên

        # -------- Bước 1: đọc tất cả các file thành root XML --------
        roots = {}
        for i in range(1, 6):
            if files[i]:
                with st.spinner(f"Đang đọc file {file_types[i]}..."):
                    root = read_file_to_root(files[i], file_types[i], all_errors)
                    if root is not None:
                        roots[file_types[i]] = root

        # -------- Bước 2: build dữ liệu tra cứu chung --------
        HANH_CHINH_DATA = {}     # ma_lk -> {MA_HSBA, HO_TEN}
        HANH_CHINH_ELEMS = {}    # ma_lk -> element (để check nhóm 3)
        ALL_HANH_CHINH = []      # list dict phục vụ check nhóm 4 (trùng lịch)
        ITEMS_BY_MALK = {}       # ma_lk -> {"thuoc":[...], "dvkt":[...], "cdha":[...]}

        if "Hành chính" in roots:
            for ho_so in roots["Hành chính"].findall('.//HO_SO_KCB'):
                ma_lk_tag = ho_so.find('MA_LK')
                ma_lk = ma_lk_tag.text.strip().upper() if ma_lk_tag is not None and ma_lk_tag.text else None
                if not ma_lk:
                    continue
                HANH_CHINH_DATA[ma_lk] = {
                    'MA_HSBA': (ho_so.findtext('MA_HSBA') or "").strip(),
                    'HO_TEN': (ho_so.findtext('HO_TEN') or "").strip()
                }
                HANH_CHINH_ELEMS[ma_lk] = ho_so

                ngay_vao = (ho_so.findtext('NGAY_VAO') or "").strip()
                ngay_ra = (ho_so.findtext('NGAY_RA') or "").strip()
                ALL_HANH_CHINH.append({
                    "ma_lk": ma_lk,
                    "ma_the": (ho_so.findtext('MA_THE_BHYT') or "").strip(),
                    "ma_cskcb": (ho_so.findtext('MA_CSKCB') or "").strip(),
                    "ngay_vao": ngay_vao,
                    "ngay_ra": ngay_ra,
                    "dt_vao": _to_datetime(ngay_vao),
                    "dt_ra": _to_datetime(ngay_ra),
                })

        def collect_items(file_type, key, field_map):
            if file_type not in roots:
                return
            for ho_so in roots[file_type].findall('.//HO_SO_KCB'):
                ma_lk_tag = ho_so.find('MA_LK')
                ma_lk = ma_lk_tag.text.strip().upper() if ma_lk_tag is not None and ma_lk_tag.text else None
                if not ma_lk:
                    continue
                item = {}
                for tag in field_map:
                    item[tag] = (ho_so.findtext(tag) or "").strip()
                ITEMS_BY_MALK.setdefault(ma_lk, {}).setdefault(key, []).append(item)

        collect_items("Thuốc", "thuoc", [
            "MA_THUOC", "TEN_THUOC", "SO_LUONG", "DON_GIA", "THANH_TIEN_BH",
            "LIEU_DUNG", "DUONG_DUNG", "NGAY_YL"
        ])
        collect_items("Dịch vụ kỹ thuật", "dvkt", [
            "MA_DICH_VU", "TEN_DICH_VU", "SO_LUONG", "DON_GIA_BH",
            "THANH_TIEN_BH", "NGAY_YL", "MA_MAY"
        ])
        collect_items("Mã máy xét nghiệm", "cdha", [
            "MA_DICH_VU", "TEN_DICH_VU", "TEN_CHI_SO", "NGAY_KQ", "NGAY_YL"
        ])

        # -------- Bước 3: chạy validation cho từng file/hồ sơ --------
        for file_type, root in roots.items():
            for ho_so in root.findall('.//HO_SO_KCB'):
                ma_lk_tag = ho_so.find('MA_LK')
                ma_lk = ma_lk_tag.text.strip().upper() if ma_lk_tag is not None and ma_lk_tag.text else "Không xác định"

                # 0. Trường bắt buộc + logic ngày + ICD + thời điểm PTTT
                all_errors.extend(check_mandatory_fields(ho_so, ma_lk, file_type))
                all_errors.extend(check_ngay_logic(ho_so, ma_lk, file_type))
                if file_type == "Hành chính":
                    all_errors.extend(check_ma_benh_icd10(ho_so, ma_lk, VALID_MA_BENH_CODES))
                if file_type == "Phẫu thuật thủ thuật":
                    all_errors.extend(check_thoi_diem_dbls(ho_so, ma_lk))

                # Nhóm 1 & 4: Hành chính, thẻ BHYT, thời gian & trùng lặp KCB
                if file_type == "Hành chính":
                    all_errors.extend(check_hanh_chinh_full(
                        ho_so, all_hanh_chinh=ALL_HANH_CHINH, danh_muc=DANH_MUC))

                # Nhóm 2: Danh mục - giá - định mức (Thuốc, DVKT)
                if file_type in ("Thuốc", "Dịch vụ kỹ thuật"):
                    for e in check_nhom2_danh_muc_gia(ho_so, ma_lk, file_type, danh_muc=DANH_MUC):
                        all_errors.append(e)

        # Nhóm 3: Hợp lý chỉ định y khoa - chạy 1 lần / hồ sơ hành chính,
        # vì cần dữ liệu tổng hợp thuốc/dvkt/cdha theo MA_LK
        for ma_lk, ho_so_hc in HANH_CHINH_ELEMS.items():
            for e in check_nhom3_hop_ly_chi_dinh(
                    ho_so_hc, "Hành chính", ITEMS_BY_MALK, ma_lk, danh_muc=DANH_MUC):
                all_errors.append(e)

        # -------- Bước 4: hiển thị kết quả --------
        st.subheader("📊 KẾT QUẢ KIỂM TRA")

        # Thêm Mã HSBA / Họ tên cho từng lỗi
        for e in all_errors:
            lookup = HANH_CHINH_DATA.get(str(e.get('ma_lk', '')).strip().upper(), {})
            e['ma_hsba'] = lookup.get('MA_HSBA', '')
            e['ho_ten'] = lookup.get('HO_TEN', '')

        n_loi = sum(1 for e in all_errors if e.get('muc_do') == "Lỗi")
        n_canh_bao = sum(1 for e in all_errors if e.get('muc_do') == "Cảnh báo")

        c1, c2, c3 = st.columns(3)
        c1.metric("Tổng số dòng lỗi/cảnh báo", len(all_errors))
        c2.metric("🔴 Lỗi (cần sửa trước khi gửi)", n_loi)
        c3.metric("🟡 Cảnh báo (nên rà soát lại)", n_canh_bao)

        if not all_errors:
            st.success("🎉 Không phát hiện lỗi/cảnh báo nào trên các file đã chọn. Hồ sơ đã sẵn sàng để gửi Cổng giám định.")
        else:
            # Gom theo MA_LK để bác sĩ/điều dưỡng dễ tra theo từng hồ sơ bệnh án
            df_all = pd.DataFrame(all_errors)
            rename_map = {
                'ma_lk': 'Mã LK',
                'ma_hsba': 'Mã HSBA',
                'ho_ten': 'Họ và Tên',
                'nguon_file': 'File',
                'nhom': 'Nhóm quy tắc',
                'muc_do': 'Mức độ',
                'truong_loi': 'Trường dữ liệu',
                'gia_tri': 'Giá trị hiện tại',
                'mo_ta_loi': 'Mô tả lỗi',
                'huong_xu_ly': 'Hướng xử lý / Cách sửa',
            }
            df_all = df_all.rename(columns=rename_map)
            col_order = ['Mã LK', 'Mã HSBA', 'Họ và Tên', 'File', 'Nhóm quy tắc', 'Mức độ',
                          'Trường dữ liệu', 'Giá trị hiện tại', 'Mô tả lỗi', 'Hướng xử lý / Cách sửa']
            col_order = [c for c in col_order if c in df_all.columns]
            df_all = df_all[col_order]

            st.markdown("#### 📋 Danh sách lỗi/cảnh báo theo từng hồ sơ (Mã LK)")
            for ma_lk_val, df_grp in df_all.groupby('Mã LK', sort=False):
                ho_ten_val = df_grp['Họ và Tên'].iloc[0] if 'Họ và Tên' in df_grp.columns else ''
                ma_hsba_val = df_grp['Mã HSBA'].iloc[0] if 'Mã HSBA' in df_grp.columns else ''
                n_loi_grp = (df_grp['Mức độ'] == "Lỗi").sum()
                n_cb_grp = (df_grp['Mức độ'] == "Cảnh báo").sum()
                label = f"🆔 {ma_lk_val} | HSBA: {ma_hsba_val or '—'} | {ho_ten_val or '(chưa rõ họ tên)'} " \
                        f"— 🔴 {n_loi_grp} lỗi, 🟡 {n_cb_grp} cảnh báo"
                with st.expander(label, expanded=(n_loi_grp > 0)):
                    show_cols = [c for c in col_order if c not in ('Mã LK', 'Mã HSBA', 'Họ và Tên')]
                    st.dataframe(df_grp[show_cols], use_container_width=True, hide_index=True)

            # ====================== TẢI BÁO CÁO EXCEL ======================
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Bao cao loi BHYT"

            headers = col_order
            ws.append(headers)
            header_fill = PatternFill("solid", start_color="305496")
            header_font = Font(bold=True, color="FFFFFF")
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for _, row in df_all.iterrows():
                ws.append([row[h] for h in headers])
                r_idx = ws.max_row
                muc_do = row.get('Mức độ', '')
                fill_color = MUC_DO_COLOR.get(muc_do)
                font_color = MUC_DO_FONT_COLOR.get(muc_do)
                if fill_color:
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=r_idx, column=col_idx)
                        cell.fill = PatternFill("solid", start_color=fill_color)
                        if font_color:
                            cell.font = Font(color=font_color)
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Độ rộng cột
            col_widths = {
                'Mã LK': 16, 'Mã HSBA': 12, 'Họ và Tên': 18, 'File': 16,
                'Nhóm quy tắc': 26, 'Mức độ': 10, 'Trường dữ liệu': 16,
                'Giá trị hiện tại': 18, 'Mô tả lỗi': 50, 'Hướng xử lý / Cách sửa': 55,
            }
            for col_idx, h in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 20)

            ws.freeze_panes = "A2"
            wb.save(output)
            output.seek(0)

            st.download_button(
                "📥 Tải báo cáo Excel đầy đủ (có màu theo mức độ Lỗi/Cảnh báo)",
                data=output,
                file_name=f"BAO_CAO_LOI_BHYT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.caption(
            "💡 Màu 🔴 đỏ = Lỗi cần sửa trước khi gửi (nguy cơ bị từ chối/xuất toán cao); "
            "màu 🟡 vàng = Cảnh báo cần bác sĩ/điều dưỡng rà soát lại hồ sơ bệnh án. "
            "Các quy tắc đối chiếu danh mục giá, ICD-10↔chỉ định chỉ chạy đầy đủ khi đã tải "
            "danh mục tương ứng ở thanh bên."
        )

# ==================== TAB 2: SO SÁNH EXCEL ====================
with tab2:
    render_excel_compare_tab()

# ==================== TAB 3: KIỂM TRA THÔNG TƯ 06 ====================
with tab3:
    try:
        with open("tt06.html", "r", encoding="utf-8") as f:
            html_source = f.read()
        components.html(html_source, height=850, scrolling=True)
    except Exception as e:
        st.error(f"Không thể tải công cụ kiểm tra TT06. Vui lòng đảm bảo file tt06.html đã được up lên cùng thư mục. Chi tiết lỗi: {e}")
