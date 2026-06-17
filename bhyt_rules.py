# -*- coding: utf-8 -*-
"""
bhyt_rules.py
Bộ quy tắc kiểm tra dữ liệu XML1-XML5 (Quyết định 4210/QĐ-BYT) trước khi đẩy
lên Cổng tiếp nhận / Giám định BHYT.

4 nhóm quy tắc:
  1. Hành chính & thẻ BHYT
  2. Danh mục thuốc / DVKT / VTYT, giá, định mức kinh tế kỹ thuật
  3. Hợp lý chỉ định y khoa (bắt chéo ICD-10, giới tính/tuổi, trùng lặp, liều dùng)
  4. Thời gian KCB & trùng lặp đợt điều trị

Mỗi lỗi trả về là 1 dict với khóa thống nhất:
    ma_lk        : Mã liên kết hồ sơ
    nhom         : Tên nhóm quy tắc (1..4)
    muc_do       : "Lỗi" (chặn, chắc chắn bị xuất toán/từ chối) hoặc
                   "Cảnh báo" (cần bác sĩ/điều dưỡng xem lại)
    truong_loi   : Tên trường dữ liệu liên quan
    gia_tri      : Giá trị hiện tại trong hồ sơ (nếu có)
    mo_ta_loi    : Diễn giải lỗi
    huong_xu_ly  : Gợi ý cách sửa
    nguon_file   : Loại file (XML1..XML5)
"""

from datetime import datetime
import re
import os

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

# =====================================================================
#  DANH MỤC CỐ ĐỊNH ĐI KÈM ỨNG DỤNG (không cần tải lên mỗi lần kiểm tra)
#  Đặt các file Excel này CÙNG THƯ MỤC với bhyt_app.py / bhyt_rules.py:
#    - danh_muc_gia_dvkt.xlsx              (cột: MA, GIA)
#    - danh_muc_ma_thuoc_hoatchat.xlsx     (cột: MA_THUOC, TEN_THUOC, TEN_HOAT_CHAT)
#    - danh_muc_thuoc_chongchidinh.xlsx    (cột: HOAT_CHAT_KEYWORD, ICD_CHONGCHIDINH, GHI_CHU)
# =====================================================================
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_GIA_DVKT_FILE = os.path.join(_BASE_DIR, "danh_muc_gia_dvkt.xlsx")
DEFAULT_THUOC_CHONGCHIDINH_FILE = os.path.join(_BASE_DIR, "danh_muc_thuoc_chongchidinh.xlsx")
DEFAULT_MA_THUOC_HOATCHAT_FILE = os.path.join(_BASE_DIR, "danh_muc_ma_thuoc_hoatchat.xlsx")
DEFAULT_NHAN_VIEN_YTE_FILE = os.path.join(_BASE_DIR, "danh_muc_nhan_vien_yte.xlsx")

# Mã chức danh nghề nghiệp (CHUCDANH_NN) liên quan tới các quy tắc đối chiếu
# nhân sự - theo xác nhận của đơn vị:
#   1 = Bác sĩ, 2 = Y sĩ, 3 = Điều dưỡng, 4 = Dược sĩ
# (5,6,7,8,10: chưa xác nhận đầy đủ ý nghĩa, riêng 6 và 10 được yêu cầu dùng
#  cho điều kiện "người thực hiện DVKT/đọc kết quả CĐHA không phải BS")
CHUCDANH_BAC_SI = "1"
CHUCDANH_Y_SI = "2"
CHUCDANH_DIEU_DUONG = "3"
CHUCDANH_DUOC_SI = "4"
CHUCDANH_DUOC_KE_THUOC = {"1"}          # được kê đơn thuốc: chỉ Bác sĩ
CHUCDANH_THUC_HIEN_DVKT = {"6", "10"}    # được thực hiện DVKT/kỹ thuật viên
CHUCDANH_CANH_BAO_SAI_PHAM_VI = {"2", "3", "4"}  # Y sĩ/Điều dưỡng/Dược sĩ - cảnh báo "sai phạm vi hành nghề" khi kê thuốc/đọc KQ

# Từ khóa nhận diện dịch vụ CĐHA/TDCN cần bác sĩ đọc kết quả (siêu âm, điện tim, X-quang)
TEN_CHI_SO_CAN_BS_DOC_KQ = [
    "SIÊU ÂM", "SIEU AM", "ĐIỆN TIM", "DIEN TIM", "X-QUANG", "XQUANG", "X QUANG",
]

# =====================================================================
#  HẰNG SỐ DÙNG CHUNG
# =====================================================================

# Mã quyền lợi hưởng -> tỷ lệ hưởng BHYT (%) - theo Luật BHYT hiện hành
# (mã 5 - một số đối tượng đặc thù 100% có điều kiện - cần đối chiếu thêm
#  đối tượng tham gia, ở đây chỉ dùng cảnh báo nếu mã ngoài danh sách)
MA_MUC_HUONG_TY_LE = {
    "1": 100,
    "2": 95,
    "3": 80,
    "4": 100,   # nhóm ưu tiên đặc biệt (theo đối tượng)
    "5": 100,   # nhóm ưu tiên đặc biệt (theo đối tượng)
}

# Định dạng ngày giờ chuẩn 4210: YYYYMMDDHHMM (12 ký tự)
DATE_FMT = "%Y%m%d%H%M"
DATE_RE = re.compile(r"^\d{12}$")

# Các DVKT/CĐHA "nặng" thường bị soi khi không có ICD phù hợp
# -> dùng để cảnh báo khi KHÔNG có bảng đối chiếu ICD-DVKT do người dùng tải lên
HIGH_COST_DVKT_KEYWORDS = [
    "CT", "MRI", "CỘNG HƯỞNG TỪ", "CẮT LỚP VI TÍNH", "NỘI SOI",
    "SIÊU ÂM", "ĐIỆN TÂM ĐỒ", "XQUANG", "X-QUANG", "X QUANG",
]

# Từ khóa thuốc/DVKT chỉ dành riêng cho một giới tính (mẫu cơ bản, mở rộng
# thêm qua danh mục do người dùng tải lên nếu cần)
FEMALE_ONLY_KEYWORDS = [
    "THAI", "SẢN", "ÂM ĐẠO", "TỬ CUNG", "BUỒNG TRỨNG", "VÒNG TRÁNH THAI",
    "PHỤ KHOA", "CỔ TỬ CUNG", "VÚ", "TUYẾN VÚ", "MANG THAI",
]
MALE_ONLY_KEYWORDS = [
    "TUYẾN TIỀN LIỆT", "TIỀN LIỆT TUYẾN", "DƯƠNG VẬT", "TINH HOÀN",
    "BAO QUY ĐẦU", "NAM KHOA",
]

# Từ khóa DVKT/thuốc chỉ dành cho trẻ em (theo tuổi < 16, mở rộng tuỳ danh mục)
PEDIATRIC_KEYWORDS = ["NHI", "SƠ SINH", "TRẺ EM"]


# =====================================================================
#  HÀM TIỆN ÍCH
# =====================================================================

def _text(ho_so, tag):
    """Lấy text đã strip của 1 tag con, trả '' nếu không có."""
    el = ho_so.find(tag)
    return (el.text or "").strip() if el is not None and el.text else ""


def _to_float(value):
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _to_datetime(value):
    value = (value or "").strip()
    if not DATE_RE.match(value):
        return None
    try:
        return datetime.strptime(value, DATE_FMT)
    except ValueError:
        return None


def _err(ma_lk, nhom, muc_do, truong_loi, gia_tri, mo_ta_loi, huong_xu_ly, nguon_file):
    return {
        "ma_lk": ma_lk or "Không xác định",
        "nhom": nhom,
        "muc_do": muc_do,
        "truong_loi": truong_loi,
        "gia_tri": gia_tri,
        "mo_ta_loi": mo_ta_loi,
        "huong_xu_ly": huong_xu_ly,
        "nguon_file": nguon_file,
    }


def _tra_nhan_vien(macchn, nhan_vien_yte):
    """Tra cứu thông tin nhân viên theo MACCHN, trả về dict hoặc None nếu không có/không có danh mục."""
    if not nhan_vien_yte or not macchn:
        return None
    return nhan_vien_yte.get(str(macchn).strip())


def _kiem_tra_nguoi_ke_don(macchn, nhan_vien_yte):
    """
    Kiểm tra người kê đơn thuốc (MA_BAC_SI ở XML2) có đủ điều kiện (CHUCDANH_NN=1, Bác sĩ).
    Trả về tuple (trang_thai, ghi_chu):
      trang_thai: "khong_co_danh_muc" | "khong_tim_thay" | "dat" | "sai_pham_vi" | "khong_xac_dinh_cd"
      ghi_chu: thông tin nhân viên (họ tên, chức danh) nếu tìm thấy
    """
    nv = _tra_nhan_vien(macchn, nhan_vien_yte)
    if nv is None:
        if not nhan_vien_yte:
            return "khong_co_danh_muc", ""
        return "khong_tim_thay", ""
    cd = nv.get("chucdanh", "")
    if cd == CHUCDANH_BAC_SI:
        return "dat", nv.get("ho_ten", "")
    if cd in CHUCDANH_CANH_BAO_SAI_PHAM_VI:
        ten_cd = {"2": "Y sĩ", "3": "Điều dưỡng", "4": "Dược sĩ"}.get(cd, f"Chức danh {cd}")
        return "sai_pham_vi", f"{nv.get('ho_ten', '')} ({ten_cd})"
    return "khong_xac_dinh_cd", f"{nv.get('ho_ten', '')} (CHUCDANH_NN={cd})"


def _kiem_tra_nguoi_thuc_hien_dvkt(macchn, nhan_vien_yte, ma_dv=None):
    """
    Kiểm tra người thực hiện DVKT/PTTT (NGUOI_THUC_HIEN) có đủ điều kiện:
    CHUCDANH_NN thuộc {6,10}, HOẶC được phân công riêng qua DVKT_KHAC/VB_PHANCONG.
    Trả về tuple (trang_thai, ghi_chu) tương tự _kiem_tra_nguoi_ke_don.
    """
    nv = _tra_nhan_vien(macchn, nhan_vien_yte)
    if nv is None:
        if not nhan_vien_yte:
            return "khong_co_danh_muc", ""
        return "khong_tim_thay", ""
    cd = nv.get("chucdanh", "")
    if cd in CHUCDANH_THUC_HIEN_DVKT:
        return "dat", nv.get("ho_ten", "")
    # Được phân công riêng (DVKT_KHAC có giá trị và có VB_PHANCONG kèm theo)
    if nv.get("dvkt_khac") and nv.get("vb_phancong"):
        if not ma_dv or ma_dv.strip() in (nv.get("dvkt_khac") or ""):
            return "dat_phan_cong", f"{nv.get('ho_ten', '')} (phân công theo {nv.get('vb_phancong')})"
    if cd in CHUCDANH_CANH_BAO_SAI_PHAM_VI:
        ten_cd = {"2": "Y sĩ", "3": "Điều dưỡng", "4": "Dược sĩ"}.get(cd, f"Chức danh {cd}")
        return "sai_pham_vi", f"{nv.get('ho_ten', '')} ({ten_cd})"
    return "khong_dat", f"{nv.get('ho_ten', '')} (CHUCDANH_NN={cd})"


# =====================================================================
#  NẠP DANH MỤC CỐ ĐỊNH KÈM SẴN ỨNG DỤNG
# =====================================================================

def _load_gia_excel(path):
    """Đọc file Excel 2 cột MA / GIA -> dict {MA: GIA (float)}."""
    result = {}
    if not load_workbook or not os.path.exists(path):
        return result
    try:
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]
        try:
            idx_ma = headers.index("MA")
            idx_gia = headers.index("GIA")
        except ValueError:
            return result
        for r in range(2, sheet.max_row + 1):
            ma = sheet.cell(r, idx_ma + 1).value
            gia = sheet.cell(r, idx_gia + 1).value
            if ma is None:
                continue
            try:
                result[str(ma).strip()] = float(gia)
            except (TypeError, ValueError):
                continue
    except Exception:
        pass
    return result


def _load_thuoc_chongchidinh_excel(path):
    """
    Đọc file Excel danh mục thuốc chống chỉ định theo ICD-10.
    Cột: HOAT_CHAT_KEYWORD | ICD_CHONGCHIDINH (nhiều mã/prefix phân tách ';') | GHI_CHU
    Trả về list các dict {keyword, icd_prefixes: [...], ghi_chu}
    """
    result = []
    if not load_workbook or not os.path.exists(path):
        return result
    try:
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]
        try:
            idx_kw = headers.index("HOAT_CHAT_KEYWORD")
            idx_icd = headers.index("ICD_CHONGCHIDINH")
            idx_note = headers.index("GHI_CHU")
        except ValueError:
            return result
        for r in range(2, sheet.max_row + 1):
            kw = sheet.cell(r, idx_kw + 1).value
            icd = sheet.cell(r, idx_icd + 1).value
            note = sheet.cell(r, idx_note + 1).value
            if not kw or not icd:
                continue
            prefixes = [p.strip().upper().replace(".", "") for p in str(icd).split(";") if p.strip()]
            result.append({
                "keyword": str(kw).strip().lower(),
                "icd_prefixes": prefixes,
                "ghi_chu": (note or "").strip() if note else "",
            })
    except Exception:
        pass
    return result


def _load_ma_thuoc_hoatchat_excel(path):
    """Đọc bảng MA_THUOC | TEN_THUOC | TEN_HOAT_CHAT -> dict {MA_THUOC: TEN_HOAT_CHAT (lower)}."""
    result = {}
    if not load_workbook or not os.path.exists(path):
        return result
    try:
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]
        try:
            idx_ma = headers.index("MA_THUOC")
            idx_hc = headers.index("TEN_HOAT_CHAT")
        except ValueError:
            return result
        for r in range(2, sheet.max_row + 1):
            ma = sheet.cell(r, idx_ma + 1).value
            hc = sheet.cell(r, idx_hc + 1).value
            if ma is None:
                continue
            result[str(ma).strip()] = (str(hc).strip().lower() if hc else "")
    except Exception:
        pass
    return result


def _load_nhan_vien_yte_excel(path):
    """
    Đọc danh mục nhân viên y tế: cột MACCHN | HO_TEN | CHUCDANH_NN | DVKT_KHAC | VB_PHANCONG
    -> dict {MACCHN: {"ho_ten":..., "chucdanh": "1"/"2".../ "", "dvkt_khac": "", "vb_phancong": ""}}
    MACCHN được chuẩn hoá (strip) làm khoá tra cứu.
    """
    result = {}
    if not load_workbook or not os.path.exists(path):
        return result
    try:
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]
        try:
            idx_ma = headers.index("MACCHN")
        except ValueError:
            return result
        idx_ten = headers.index("HO_TEN") if "HO_TEN" in headers else None
        idx_cd = headers.index("CHUCDANH_NN") if "CHUCDANH_NN" in headers else None
        idx_dvkt = headers.index("DVKT_KHAC") if "DVKT_KHAC" in headers else None
        idx_vb = headers.index("VB_PHANCONG") if "VB_PHANCONG" in headers else None

        for r in range(2, sheet.max_row + 1):
            macchn = sheet.cell(r, idx_ma + 1).value
            if macchn is None or str(macchn).strip() == "":
                continue
            macchn = str(macchn).strip()

            chucdanh = ""
            if idx_cd is not None:
                cd_val = sheet.cell(r, idx_cd + 1).value
                if cd_val is not None:
                    try:
                        chucdanh = str(int(float(cd_val)))
                    except (TypeError, ValueError):
                        chucdanh = str(cd_val).strip()

            result[macchn] = {
                "ho_ten": str(sheet.cell(r, idx_ten + 1).value).strip() if idx_ten is not None and sheet.cell(r, idx_ten + 1).value else "",
                "chucdanh": chucdanh,
                "dvkt_khac": str(sheet.cell(r, idx_dvkt + 1).value).strip() if idx_dvkt is not None and sheet.cell(r, idx_dvkt + 1).value else "",
                "vb_phancong": str(sheet.cell(r, idx_vb + 1).value).strip() if idx_vb is not None and sheet.cell(r, idx_vb + 1).value else "",
            }
    except Exception:
        pass
    return result


def load_default_danh_muc():
    """
    Nạp các danh mục cố định kèm sẵn ứng dụng (đặt cùng thư mục với bhyt_rules.py):
      - gia_dvkt: dict {MA: GIA}
      - thuoc_chong_chi_dinh: list[{keyword, icd_prefixes, ghi_chu}]
      - ma_thuoc_hoat_chat: dict {MA_THUOC: ten_hoat_chat (lowercase)}
      - nhan_vien_yte: dict {MACCHN: {ho_ten, chucdanh, dvkt_khac, vb_phancong}}
    Trả về dict có thể truyền trực tiếp vào tham số `danh_muc` của các hàm check_*.
    Nếu file không tồn tại, trả về dict rỗng/{} cho mục đó (không lỗi).

    LƯU Ý: không còn nạp danh mục giá thuốc (gia_thuoc) - theo yêu cầu, việc
    kiểm tra thuốc chỉ tập trung vào tương tác/chống chỉ định theo ICD-10.
    """
    return {
        "gia_dvkt": _load_gia_excel(DEFAULT_GIA_DVKT_FILE),
        "thuoc_chong_chi_dinh": _load_thuoc_chongchidinh_excel(DEFAULT_THUOC_CHONGCHIDINH_FILE),
        "ma_thuoc_hoat_chat": _load_ma_thuoc_hoatchat_excel(DEFAULT_MA_THUOC_HOATCHAT_FILE),
        "nhan_vien_yte": _load_nhan_vien_yte_excel(DEFAULT_NHAN_VIEN_YTE_FILE),
    }



def merge_danh_muc(*danh_muc_dicts):
    """
    Gộp nhiều dict danh mục lại, dict sau ưu tiên ghi đè dict trước cho các
    danh mục dạng key->value (gia_dvkt, gia_thuoc, icd_chi_dinh), còn danh mục
    dạng list (thuoc_chong_chi_dinh) sẽ được nối (extend).
    Dùng để gộp danh mục mặc định (load_default_danh_muc) với danh mục người
    dùng tải lên thêm trong phiên làm việc.
    """
    out = {}
    for dm in danh_muc_dicts:
        if not dm:
            continue
        for k, v in dm.items():
            if k not in out:
                out[k] = v
                continue
            if isinstance(v, dict) and isinstance(out[k], dict):
                merged = dict(out[k])
                merged.update(v)
                out[k] = merged
            elif isinstance(v, list) and isinstance(out[k], list):
                out[k] = out[k] + v
            else:
                out[k] = v
    return out


# =====================================================================
#  NHÓM 1: HÀNH CHÍNH & THẺ BHYT  (chạy trên dữ liệu XML1 - Hành chính)
# =====================================================================

def check_nhom1_hanh_chinh(ho_so, ma_lk, danh_muc=None):
    """
    Kiểm tra:
      1.1 Hiệu lực thẻ BHYT (GT_THE_TU / GT_THE_DEN so với NGAY_VAO / NGAY_RA)
      1.2 Khớp định danh Họ tên / Ngày sinh / Giới tính (đã chuẩn hoá cơ bản)
      1.3 Mã quyền lợi (MA_LOAI_KCB / MA_DOITUONG_KCB -> MUC_HUONG) hợp lệ và
          khớp với tỷ lệ hưởng đã khai (TYLE_BHTT / MUC_HUONG)
      1.4 Thông tin chuyển tuyến (MA_LOAI_RV / NOICHUYEN / SO_GIAY_CHUYEN_TUYEN)
    """
    errors = []
    nguon = "XML1 - Hành chính"

    # ---- 1.1 Hiệu lực thẻ BHYT ----
    gt_the_tu = _text(ho_so, "GT_THE_TU")
    gt_the_den = _text(ho_so, "GT_THE_DEN")
    ngay_vao = _text(ho_so, "NGAY_VAO")
    ngay_ra = _text(ho_so, "NGAY_RA")

    dt_tu = _to_datetime(gt_the_tu) if gt_the_tu else None
    dt_den = _to_datetime(gt_the_den) if gt_the_den else None
    dt_vao = _to_datetime(ngay_vao)
    dt_ra = _to_datetime(ngay_ra)

    if gt_the_tu and gt_the_den:
        if dt_tu is None or dt_den is None:
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "GT_THE_TU/GT_THE_DEN", f"{gt_the_tu} - {gt_the_den}",
                "Định dạng giá trị thẻ (GT_THE_TU/GT_THE_DEN) không đúng YYYYMMDDHHMM.",
                "Kiểm tra lại định dạng ngày trên thẻ BHYT (12 số: năm/tháng/ngày/giờ/phút).", nguon
            ))
        else:
            if dt_vao and dt_vao < dt_tu:
                errors.append(_err(
                    ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                    "NGAY_VAO/GT_THE_TU", f"Vào viện: {ngay_vao} | Thẻ có giá trị từ: {gt_the_tu}",
                    "Ngày vào viện trước thời điểm thẻ BHYT có giá trị sử dụng -> thẻ chưa có hiệu lực.",
                    "Kiểm tra lại ngày vào viện hoặc kiểm tra thông tin gia hạn thẻ BHYT của người bệnh.", nguon
                ))
            if dt_ra and dt_den and dt_ra > dt_den:
                errors.append(_err(
                    ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                    "NGAY_RA/GT_THE_DEN", f"Ra viện: {ngay_ra} | Thẻ có giá trị đến: {gt_the_den}",
                    "Ngày ra viện sau thời điểm thẻ BHYT hết hạn -> phần chi phí phát sinh sau "
                    "ngày hết hạn thẻ có thể không được thanh toán.",
                    "Đề nghị người bệnh gia hạn/đóng tiếp BHYT hoặc kiểm tra lại GT_THE_DEN; "
                    "nếu thẻ đã được gia hạn cần cập nhật lại GT_THE_DEN cho đúng.", nguon
                ))
    else:
        errors.append(_err(
            ma_lk, "1. Hành chính & thẻ BHYT", "Cảnh báo",
            "GT_THE_TU/GT_THE_DEN", "(thiếu)",
            "Thiếu thông tin giá trị sử dụng thẻ BHYT (GT_THE_TU/GT_THE_DEN) để kiểm tra hiệu lực thẻ.",
            "Bổ sung GT_THE_TU, GT_THE_DEN lấy từ kết quả tra cứu thẻ BHYT (Cổng tiếp nhận).", nguon
        ))

    # ---- 1.2 Khớp định danh: Họ tên / Ngày sinh / Giới tính ----
    ho_ten = _text(ho_so, "HO_TEN")
    ngay_sinh = _text(ho_so, "NGAY_SINH")
    gioi_tinh = _text(ho_so, "GIOI_TINH")

    if ho_ten:
        # Cảnh báo các ký tự bất thường (số, ký tự đặc biệt) trong họ tên
        if re.search(r"[0-9@#$%^&*()_+=\[\]{};:\"\\|<>/~`]", ho_ten):
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "HO_TEN", ho_ten,
                "Họ tên chứa ký tự số hoặc ký tự đặc biệt, không hợp lệ.",
                "Sửa lại Họ tên đúng theo CCCD/Thẻ BHYT, chỉ gồm chữ cái và dấu cách.", nguon
            ))
        if ho_ten != ho_ten.strip() or "  " in ho_ten:
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Cảnh báo",
                "HO_TEN", ho_ten,
                "Họ tên có khoảng trắng thừa ở đầu/cuối hoặc giữa các từ.",
                "Chuẩn hoá lại Họ tên: xoá khoảng trắng thừa, viết hoa chữ đầu mỗi từ.", nguon
            ))

    if ngay_sinh:
        # Cho phép YYYYMMDDHHMM hoặc YYYY (chỉ năm sinh)
        if not (re.match(r"^\d{12}$", ngay_sinh) or re.match(r"^\d{4}$", ngay_sinh)):
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "NGAY_SINH", ngay_sinh,
                "Ngày sinh không đúng định dạng (YYYYMMDDHHMM hoặc YYYY nếu chỉ có năm sinh).",
                "Sửa lại Ngày sinh đúng định dạng theo CSDL quốc gia về bảo hiểm.", nguon
            ))
        else:
            if re.match(r"^\d{12}$", ngay_sinh):
                dt_sinh = _to_datetime(ngay_sinh)
                if dt_sinh and dt_vao and dt_sinh > dt_vao:
                    errors.append(_err(
                        ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                        "NGAY_SINH", ngay_sinh,
                        "Ngày sinh lớn hơn (sau) ngày vào viện - không hợp lý.",
                        "Kiểm tra lại Ngày sinh của người bệnh theo CCCD/giấy khai sinh.", nguon
                    ))

    if gioi_tinh:
        if gioi_tinh not in ("1", "2", "3"):
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "GIOI_TINH", gioi_tinh,
                "Mã giới tính không hợp lệ (chỉ nhận 1=Nam, 2=Nữ, 3=Khác/không xác định).",
                "Sửa lại GIOI_TINH theo đúng mã quy định (1: Nam, 2: Nữ).", nguon
            ))

    # ---- 1.3 Mã đối tượng và mức hưởng ----
    ma_doi_tuong = _text(ho_so, "MA_DOITUONG_KCB")
    muc_huong = _text(ho_so, "MUC_HUONG")
    ty_le_bhtt = _text(ho_so, "TYLE_BHTT")

    if muc_huong:
        if muc_huong not in MA_MUC_HUONG_TY_LE:
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "MUC_HUONG", muc_huong,
                "Mã mức hưởng (MUC_HUONG) không thuộc danh mục hợp lệ (1-5).",
                "Tra lại mã mức hưởng trên thẻ BHYT/kết quả kiểm tra thẻ, sửa MUC_HUONG cho đúng.", nguon
            ))
        else:
            ty_le_so = _to_float(ty_le_bhtt)
            ty_le_chuan = MA_MUC_HUONG_TY_LE[muc_huong]
            if ty_le_so is not None and abs(ty_le_so - ty_le_chuan) > 0.01:
                if muc_huong in ("1", "2", "3"):
                    errors.append(_err(
                        ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                        "TYLE_BHTT", ty_le_bhtt,
                        f"Tỷ lệ BHTT khai ({ty_le_bhtt}%) không khớp với tỷ lệ chuẩn của mức hưởng "
                        f"MUC_HUONG={muc_huong} (chuẩn = {ty_le_chuan}%).",
                        "Kiểm tra lại MUC_HUONG và TYLE_BHTT, đảm bảo khớp đúng theo mức hưởng trên thẻ.", nguon
                    ))
                else:
                    errors.append(_err(
                        ma_lk, "1. Hành chính & thẻ BHYT", "Cảnh báo",
                        "TYLE_BHTT", ty_le_bhtt,
                        f"Tỷ lệ BHTT khai ({ty_le_bhtt}%) khác mặc định ({ty_le_chuan}%) đối với "
                        f"MUC_HUONG={muc_huong}.",
                        "Mã mức hưởng 4/5 có thể có tỷ lệ đặc biệt theo từng đối tượng ưu tiên "
                        "(người có công, trẻ em dưới 6 tuổi, hộ nghèo...). Đối chiếu với quyết định/"
                        "giấy chứng nhận đối tượng để xác nhận TYLE_BHTT đã đúng trước khi gửi.", nguon
                    ))
    else:
        errors.append(_err(
            ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
            "MUC_HUONG", "(thiếu)",
            "Thiếu mã mức hưởng (MUC_HUONG) - không thể tính tỷ lệ thanh toán BHYT.",
            "Bổ sung MUC_HUONG (1-5) lấy từ kết quả tra cứu thẻ BHYT.", nguon
        ))

    # ---- 1.4 Thông tin chuyển tuyến ----
    # Dấu hiệu chuyển tuyến: chỉ dựa vào MA_NOI_DI / SO_GIAYCHUYEN đã có giá trị
    # (không dựa vào MA_LOAI_KCB vì cách phân loại mã này có thể khác nhau
    # giữa các phiên bản danh mục dùng chung).
    ma_noi_di = _text(ho_so, "MA_NOI_DI")
    so_giay_chuyen_tuyen = _text(ho_so, "SO_GIAYCHUYEN")

    if ma_noi_di or so_giay_chuyen_tuyen:
        # Có dấu hiệu chuyển tuyến -> bắt buộc đủ cả mã nơi chuyển đi và số giấy chuyển tuyến
        if not ma_noi_di:
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "MA_NOI_DI", "(thiếu)",
                "Hồ sơ có Số giấy chuyển tuyến (SO_GIAYCHUYEN) nhưng thiếu mã cơ sở chuyển đi (MA_NOI_DI).",
                "Bổ sung mã cơ sở y tế chuyển đi (MA_NOI_DI) theo Giấy chuyển tuyến.", nguon
            ))
        if not so_giay_chuyen_tuyen:
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "SO_GIAYCHUYEN", "(thiếu)",
                "Hồ sơ có mã cơ sở chuyển đi (MA_NOI_DI) nhưng thiếu Số giấy chuyển tuyến (SO_GIAYCHUYEN).",
                "Bổ sung số Giấy chuyển tuyến/Phiếu chuyển tuyến đầy đủ, đúng số đã cấp tại nơi chuyển đi.", nguon
            ))

    # ---- 1.5 Mã bệnh chính chỉ được chứa 1 mã duy nhất ----
    ma_benh_chinh = _text(ho_so, "MA_BENH_CHINH")
    if ma_benh_chinh and re.search(r"[;,]", ma_benh_chinh):
        errors.append(_err(
            ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
            "MA_BENH_CHINH", ma_benh_chinh,
            "Mã bệnh chính (MA_BENH_CHINH) chỉ được khai 1 mã ICD-10 duy nhất, "
            "không được ghi nhiều mã phân tách bằng ';' hoặc ','.",
            "Chỉ giữ lại 1 mã bệnh chính chính xác nhất; các mã bệnh khác (nếu có) "
            "chuyển sang MA_BENH_KT (bệnh kèm theo).", nguon
        ))

    # ---- 1.6 Công thức tổng chi phí: T_TONGCHI_BV = T_TONGCHI_BH + T_BNTT + T_BNCCT ----
    t_tongchi_bv = _to_float(_text(ho_so, "T_TONGCHI_BV"))
    t_tongchi_bh = _to_float(_text(ho_so, "T_TONGCHI_BH"))
    t_bntt = _to_float(_text(ho_so, "T_BNTT"))
    t_bncct = _to_float(_text(ho_so, "T_BNCCT"))

    if t_tongchi_bv is not None and t_tongchi_bh is not None and t_bntt is not None and t_bncct is not None:
        tong_tinh_lai = round(t_tongchi_bh + t_bntt + t_bncct, 0)
        if abs(round(t_tongchi_bv, 0) - tong_tinh_lai) > max(1, 0.01 * tong_tinh_lai):
            errors.append(_err(
                ma_lk, "1. Hành chính & thẻ BHYT", "Lỗi",
                "T_TONGCHI_BV", t_tongchi_bv,
                f"Tổng chi phí KCB (T_TONGCHI_BV={t_tongchi_bv:,.0f}) không khớp với tổng "
                f"T_TONGCHI_BH ({t_tongchi_bh:,.0f}) + T_BNTT ({t_bntt:,.0f}) + T_BNCCT "
                f"({t_bncct:,.0f}) = {tong_tinh_lai:,.0f}.",
                "Kiểm tra lại công thức tính: T_TONGCHI_BV = T_TONGCHI_BH + T_BNTT + T_BNCCT. "
                "Rà soát các khoản chi phí BHYT chi trả, bệnh nhân tự trả, bệnh nhân cùng chi trả.", nguon
            ))

    return errors


def check_nhom2_danh_muc_gia(ho_so, ma_lk, file_type, danh_muc=None):
    """
    danh_muc (optional, do người dùng tải lên) có thể chứa:
      danh_muc['gia_dvkt']    : dict {MA_DICH_VU: gia_duyet}
      danh_muc['gia_vtyt']    : dict {MA_VTYT: gia_duyet}
      danh_muc['dinh_muc']    : dict {MA_DICH_VU: {MA_VTYT: so_luong_toi_da}}
      danh_muc['nhan_vien_yte']: dict {MACCHN: {ho_ten, chucdanh, dvkt_khac, vb_phancong}}

    LƯU Ý: Theo yêu cầu, KHÔNG kiểm tra giá/đơn giá thuốc (file "Thuốc") trong
    nhóm này nữa - việc kiểm tra thuốc chỉ còn tập trung vào tương tác/chống
    chỉ định theo ICD-10 (xem nhóm 3 - mục 3.5) và chứng chỉ hành nghề người
    kê đơn (mục 2.0 dưới đây).
    Nếu không có danh mục giá DVKT, chỉ kiểm tra các quy tắc logic nội bộ
    (đơn giá x số lượng = thành tiền, thành tiền BH <= thành tiền, ...) cho DVKT.
    """
    errors = []
    danh_muc = danh_muc or {}
    nhan_vien_yte = danh_muc.get("nhan_vien_yte", {})

    if file_type == "Thuốc":
        # ---- 2.0 Chứng chỉ hành nghề người kê đơn (MA_BAC_SI phải là Bác sĩ) ----
        ma_bac_si = _text(ho_so, "MA_BAC_SI")
        ma_thuoc = _text(ho_so, "MA_THUOC")
        ten_thuoc = _text(ho_so, "TEN_THUOC")

        if ma_bac_si:
            trang_thai, ghi_chu = _kiem_tra_nguoi_ke_don(ma_bac_si, nhan_vien_yte)
            if trang_thai == "khong_tim_thay":
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                    "MA_BAC_SI", ma_bac_si,
                    f"Không tìm thấy mã chứng chỉ hành nghề '{ma_bac_si}' (người kê thuốc "
                    f"'{ma_thuoc}' - {ten_thuoc}) trong danh mục nhân viên y tế của đơn vị.",
                    "Kiểm tra lại MA_BAC_SI hoặc bổ sung nhân viên này vào danh mục nhân viên y tế "
                    "(danh_muc_nhan_vien_yte.xlsx).", "XML2 - Thuốc"
                ))
            elif trang_thai == "sai_pham_vi":
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                    "MA_BAC_SI", ma_bac_si,
                    f"Người kê đơn thuốc '{ma_thuoc}' ({ten_thuoc}) là {ghi_chu}, không phải Bác sĩ "
                    "(CHUCDANH_NN=1) - không đúng phạm vi hành nghề được kê đơn thuốc.",
                    "Kiểm tra lại người kê đơn (MA_BAC_SI); chỉ Bác sĩ mới được kê đơn thuốc theo "
                    "quy định về phạm vi hành nghề.", "XML2 - Thuốc"
                ))
            elif trang_thai == "khong_xac_dinh_cd":
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                    "MA_BAC_SI", ma_bac_si,
                    f"Người kê đơn thuốc '{ma_thuoc}' ({ten_thuoc}) - {ghi_chu} - chưa xác định rõ "
                    "có phải Bác sĩ hay không.",
                    "Kiểm tra lại chức danh nghề nghiệp (CHUCDANH_NN) của nhân viên này trong danh mục.",
                    "XML2 - Thuốc"
                ))
        # Không kiểm tra giá/đơn giá/thành tiền của thuốc nữa.
        # Việc kiểm tra thuốc tập trung ở Nhóm 3 - mục 3.5 (tương tác/chống chỉ định theo ICD-10).
        return errors

    elif file_type == "Dịch vụ kỹ thuật":
        nguon = "XML3 - DVKT"
        ma_dv = _text(ho_so, "MA_DICH_VU")
        ten_dv = _text(ho_so, "TEN_DICH_VU")
        so_luong = _to_float(_text(ho_so, "SO_LUONG"))
        don_gia_bh = _to_float(_text(ho_so, "DON_GIA_BH"))
        thanh_tien_bh = _to_float(_text(ho_so, "THANH_TIEN_BH"))

        gia_dvkt_dm = danh_muc.get("gia_dvkt", {})
        if gia_dvkt_dm:
            if ma_dv and ma_dv not in gia_dvkt_dm:
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                    "MA_DICH_VU", ma_dv,
                    f"Mã DVKT '{ma_dv}' ({ten_dv}) không có trong danh mục kỹ thuật được phê duyệt "
                    "thực hiện tại cơ sở.",
                    "Kiểm tra lại mã DVKT theo danh mục kỹ thuật đã được phê duyệt phân loại phẫu thuật, "
                    "thủ thuật/Danh mục dùng chung.", nguon
                ))
            elif ma_dv and don_gia_bh is not None:
                gia_duyet = gia_dvkt_dm.get(ma_dv)
                if gia_duyet is not None and abs(don_gia_bh - gia_duyet) > 0.5:
                    errors.append(_err(
                        ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                        "DON_GIA_BH", don_gia_bh,
                        f"Đơn giá DVKT '{ma_dv}' ({ten_dv}) = {don_gia_bh:,.0f} không khớp giá theo "
                        f"quy định hiện hành = {gia_duyet:,.0f}.",
                        "Sửa lại DON_GIA_BH theo đúng bảng giá dịch vụ kỹ thuật hiện hành (Thông tư giá).", nguon
                    ))

        if so_luong is not None and don_gia_bh is not None and thanh_tien_bh is not None:
            tinh_lai = round(so_luong * don_gia_bh, 0)
            if thanh_tien_bh - tinh_lai > max(1, 0.01 * tinh_lai):
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                    "THANH_TIEN_BH", thanh_tien_bh,
                    f"Thành tiền BHYT ({thanh_tien_bh:,.0f}) lớn hơn Số lượng x Đơn giá BH "
                    f"({so_luong:g} x {don_gia_bh:,.0f} = {tinh_lai:,.0f}).",
                    "Kiểm tra lại công thức tính thành tiền BHYT của dịch vụ kỹ thuật.", nguon
                ))

        if so_luong is not None and so_luong <= 0:
            errors.append(_err(
                ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                "SO_LUONG", so_luong,
                "Số lượng DVKT phải lớn hơn 0.",
                "Kiểm tra lại số lượng dịch vụ kỹ thuật đã thực hiện.", nguon
            ))

        # 2.3 Cảnh báo định mức VTYT hao phí kèm DVKT (nếu có danh mục)
        dinh_muc = danh_muc.get("dinh_muc", {})
        if dinh_muc and ma_dv in dinh_muc:
            # Việc đối chiếu thực tế VTYT đã xuất kèm DVKT cần dữ liệu chi tiết
            # VTYT theo từng DVKT (thường nằm trong XML3 mở rộng / XML7).
            # Ở đây chỉ nhắc nhở để bác sĩ/điều dưỡng tự rà soát theo định mức.
            errors.append(_err(
                ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                "MA_DICH_VU", ma_dv,
                f"DVKT '{ma_dv}' ({ten_dv}) có định mức VTYT hao phí kèm theo - cần rà soát "
                "số lượng VTYT xuất kèm không vượt định mức quy định.",
                "Đối chiếu số lượng VTYT đã xuất cho DVKT này với định mức kinh tế kỹ thuật đã tải lên; "
                "nếu vượt định mức cần ghi rõ lý do (biến chứng, tai biến) trong hồ sơ bệnh án.", nguon
            ))

        # ---- 2.4 MA_MAY phải TRỐNG khi là VTYT hoặc khi TEN_DICH_VU là "giường" ----
        ma_vat_tu = _text(ho_so, "MA_VAT_TU")
        ma_may = _text(ho_so, "MA_MAY")
        la_giuong = "giường" in (ten_dv or "").lower() or "giuong" in (ten_dv or "").lower()
        if (ma_vat_tu or la_giuong) and ma_may:
            ly_do = "là vật tư y tế (có MA_VAT_TU)" if ma_vat_tu else "là dịch vụ giường bệnh"
            errors.append(_err(
                ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                "MA_MAY", ma_may,
                f"Dòng {ly_do} không được khai mã máy thực hiện (MA_MAY), nhưng hiện đang có "
                f"giá trị '{ma_may}'.",
                "Xoá giá trị MA_MAY ở dòng này (VTYT/giường bệnh không gắn với máy thực hiện).", nguon
            ))

        # ---- 2.5 Người thực hiện DVKT (NGUOI_THUC_HIEN) phải đúng chứng chỉ hành nghề ----
        nguoi_th = _text(ho_so, "NGUOI_THUC_HIEN")
        if nguoi_th and not ma_vat_tu and not la_giuong:
            trang_thai, ghi_chu = _kiem_tra_nguoi_thuc_hien_dvkt(nguoi_th, nhan_vien_yte, ma_dv=ma_dv)
            if trang_thai == "khong_tim_thay":
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                    "NGUOI_THUC_HIEN", nguoi_th,
                    f"Không tìm thấy mã chứng chỉ hành nghề '{nguoi_th}' (người thực hiện DVKT "
                    f"'{ma_dv}' - {ten_dv}) trong danh mục nhân viên y tế.",
                    "Kiểm tra lại NGUOI_THUC_HIEN hoặc bổ sung nhân viên vào danh mục nhân viên y tế.", nguon
                ))
            elif trang_thai == "khong_dat":
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                    "NGUOI_THUC_HIEN", nguoi_th,
                    f"Người thực hiện DVKT '{ma_dv}' ({ten_dv}) - {ghi_chu} - không thuộc chức danh "
                    "kỹ thuật viên (CHUCDANH_NN=6,10) và không có văn bản phân công thực hiện kỹ thuật khác.",
                    "Kiểm tra lại người thực hiện hoặc bổ sung văn bản phân công (DVKT_KHAC/VB_PHANCONG) "
                    "trong danh mục nhân viên y tế nếu việc phân công là có thật.", nguon
                ))
            elif trang_thai == "sai_pham_vi":
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                    "NGUOI_THUC_HIEN", nguoi_th,
                    f"Người thực hiện DVKT '{ma_dv}' ({ten_dv}) là {ghi_chu}, không phải kỹ thuật viên "
                    "và không có văn bản phân công thực hiện kỹ thuật khác.",
                    "Kiểm tra lại người thực hiện hoặc bổ sung văn bản phân công nếu hợp lệ.", nguon
                ))

        # ---- 2.6 Logic thời gian: NGAY_YL <= NGAY_TH_YL < NGAY_KQ ----
        ngay_yl = _text(ho_so, "NGAY_YL")
        ngay_th_yl = _text(ho_so, "NGAY_TH_YL")
        ngay_kq = _text(ho_so, "NGAY_KQ")
        dt_yl = _to_datetime(ngay_yl) if ngay_yl else None
        dt_th_yl = _to_datetime(ngay_th_yl) if ngay_th_yl else None
        dt_kq = _to_datetime(ngay_kq) if ngay_kq else None

        if dt_yl and dt_th_yl and dt_th_yl < dt_yl:
            errors.append(_err(
                ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                "NGAY_TH_YL", ngay_th_yl,
                f"DVKT '{ma_dv}' ({ten_dv}): Ngày thực hiện y lệnh ({ngay_th_yl}) trước ngày "
                f"chỉ định (NGAY_YL={ngay_yl}) - không hợp lý.",
                "Kiểm tra lại NGAY_YL và NGAY_TH_YL, đảm bảo NGAY_TH_YL >= NGAY_YL.", nguon
            ))
        if dt_th_yl and dt_kq and dt_kq <= dt_th_yl:
            errors.append(_err(
                ma_lk, "2. Danh mục - Giá - Định mức", "Lỗi",
                "NGAY_KQ", ngay_kq,
                f"DVKT '{ma_dv}' ({ten_dv}): Ngày có kết quả ({ngay_kq}) phải lớn hơn ngày thực hiện "
                f"y lệnh (NGAY_TH_YL={ngay_th_yl}).",
                "Kiểm tra lại NGAY_TH_YL và NGAY_KQ, đảm bảo NGAY_KQ > NGAY_TH_YL.", nguon
            ))

        # ---- 2.7 Thời gian thực hiện tối thiểu 6 phút cho Điện tim/Siêu âm/X-quang ----
        la_dt_sa_xq = any(k in (ten_dv or "").upper() for k in
                           ["ĐIỆN TIM", "DIEN TIM", "SIÊU ÂM", "SIEU AM", "X-QUANG", "XQUANG", "X QUANG"])
        if la_dt_sa_xq and dt_th_yl and dt_kq:
            so_phut = (dt_kq - dt_th_yl).total_seconds() / 60
            if 0 <= so_phut < 6:
                errors.append(_err(
                    ma_lk, "2. Danh mục - Giá - Định mức", "Cảnh báo",
                    "NGAY_KQ", ngay_kq,
                    f"DVKT '{ma_dv}' ({ten_dv}): thời gian từ lúc thực hiện (NGAY_TH_YL={ngay_th_yl}) "
                    f"đến lúc có kết quả (NGAY_KQ={ngay_kq}) chỉ {so_phut:.0f} phút, ít hơn mức tối "
                    "thiểu 6 phút đối với Điện tim/Siêu âm/X-quang.",
                    "Kiểm tra lại thời gian thực hiện và trả kết quả; nếu thực hiện đúng quy trình "
                    "cần đảm bảo thời gian hợp lý (>=6 phút) để tránh bị nghi vấn khi giám định.", nguon
                ))

    return errors

def check_nhom3_hop_ly_chi_dinh(ho_so_hc, file_type, items_by_malk, ma_lk, danh_muc=None):
    """
    items_by_malk: dict {ma_lk: {"thuoc": [...], "dvkt": [...], "cdha": [...]}}
                    - mỗi item là dict các trường gốc đã đọc được từ XML tương ứng,
                      dùng để kiểm tra trùng lặp / liều dùng / giới tính-tuổi.
    danh_muc có thể chứa:
      danh_muc['icd_chi_dinh'] : dict {MA_BENH (ICD10): set(MA_DICH_VU hoặc MA_THUOC cho phép)}
    Trả về list lỗi áp dụng cho TỪNG hồ sơ (ma_lk), được gọi 1 lần / ma_lk
    sau khi đã đọc xong tất cả các file.
    """
    errors = []
    danh_muc = danh_muc or {}
    nhan_vien_yte = danh_muc.get("nhan_vien_yte", {})
    if ho_so_hc is None:
        return errors

    ma_benh_chinh = _text(ho_so_hc, "MA_BENH_CHINH")
    ma_benh_kt = _text(ho_so_hc, "MA_BENH_KT")  # mã bệnh kèm theo (có thể nhiều mã, phân tách bằng ;)
    cac_ma_benh = set()
    if ma_benh_chinh:
        cac_ma_benh.add(ma_benh_chinh.upper())
    if ma_benh_kt:
        for m in re.split(r"[;,]", ma_benh_kt):
            m = m.strip().upper()
            if m:
                cac_ma_benh.add(m)

    gioi_tinh = _text(ho_so_hc, "GIOI_TINH")  # 1=Nam, 2=Nữ
    ngay_sinh = _text(ho_so_hc, "NGAY_SINH")
    ngay_vao = _text(ho_so_hc, "NGAY_VAO")

    tuoi = None
    dt_vao = _to_datetime(ngay_vao)
    if re.match(r"^\d{12}$", ngay_sinh or ""):
        dt_sinh = _to_datetime(ngay_sinh)
        if dt_sinh and dt_vao:
            tuoi = (dt_vao - dt_sinh).days // 365
    elif re.match(r"^\d{4}$", ngay_sinh or "") and dt_vao:
        try:
            tuoi = dt_vao.year - int(ngay_sinh)
        except ValueError:
            tuoi = None

    data = items_by_malk.get(ma_lk, {})

    # ---- 3.1 Chỉ định theo ICD-10 (bắt chéo với danh mục nếu có) ----
    icd_chi_dinh = danh_muc.get("icd_chi_dinh", {})
    for nhom_item, ds in (("dvkt", data.get("dvkt", [])), ("thuoc", data.get("thuoc", []))):
        for item in ds:
            ma_code = item.get("MA_DICH_VU") or item.get("MA_THUOC")
            ten_code = item.get("TEN_DICH_VU") or item.get("TEN_THUOC") or ""
            nguon = "XML3 - DVKT" if nhom_item == "dvkt" else "XML2 - Thuốc"
            truong_loi = "MA_DICH_VU" if nhom_item == "dvkt" else "MA_THUOC"

            if icd_chi_dinh:
                cho_phep = False
                for ma_benh in cac_ma_benh:
                    allowed_set = icd_chi_dinh.get(ma_benh) or icd_chi_dinh.get(ma_benh.split(".")[0])
                    if allowed_set and ma_code in allowed_set:
                        cho_phep = True
                        break
                if not cho_phep:
                    errors.append(_err(
                        ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                        truong_loi, ma_code,
                        f"Chỉ định '{ma_code}' ({ten_code}) chưa thấy trong danh mục chỉ định phù hợp "
                        f"với mã bệnh chính/kèm theo ({', '.join(cac_ma_benh) or '(trống)'}).",
                        "Rà soát lại: nếu chỉ định này phù hợp với chẩn đoán, bổ sung mã bệnh kèm theo "
                        "(MA_BENH_KT) tương ứng vào hồ sơ hành chính; nếu không phù hợp, xem lại chỉ định "
                        "hoặc bổ sung lý do y khoa trong bệnh án để tránh bị xuất toán.", nguon
                    ))
            else:
                # Không có bảng đối chiếu ICD-DVKT: chỉ cảnh báo nhẹ cho các
                # DVKT chi phí cao thường bị soi, để bác sĩ tự rà soát ICD
                if nhom_item == "dvkt" and any(k in (ten_code or "").upper() for k in HIGH_COST_DVKT_KEYWORDS):
                    if not cac_ma_benh:
                        errors.append(_err(
                            ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                            truong_loi, ma_code,
                            f"[Gợi ý theo từ khóa] Chỉ định '{ma_code}' ({ten_code}) là DVKT chi phí cao "
                            "nhưng hồ sơ chưa có mã bệnh chính (MA_BENH_CHINH) để đối chiếu mức độ hợp lý.",
                            "Bổ sung MA_BENH_CHINH (và MA_BENH_KT nếu có) trong XML1 - Hành chính.", nguon
                        ))

    # ---- 3.2 Chỉ định theo giới tính / tuổi ----
    for nhom_item, ds in (("dvkt", data.get("dvkt", [])), ("thuoc", data.get("thuoc", []))):
        for item in ds:
            ma_code = item.get("MA_DICH_VU") or item.get("MA_THUOC")
            ten_code = (item.get("TEN_DICH_VU") or item.get("TEN_THUOC") or "").upper()
            nguon = "XML3 - DVKT" if nhom_item == "dvkt" else "XML2 - Thuốc"
            truong_loi = "MA_DICH_VU" if nhom_item == "dvkt" else "MA_THUOC"

            if gioi_tinh == "1" and any(k in ten_code for k in FEMALE_ONLY_KEYWORDS):
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    truong_loi, ma_code,
                    f"[Gợi ý theo từ khóa - cần xác nhận] Chỉ định '{ma_code}' ({ten_code.title()}) "
                    "có tên gợi ý dành cho bệnh nhân nữ, nhưng GIOI_TINH của hồ sơ là Nam (1).",
                    "Kiểm tra lại: nếu chỉ định này thực sự không phù hợp với giới tính người bệnh, "
                    "sửa lại chỉ định; nếu phù hợp (ví dụ tên dịch vụ trùng từ khóa nhưng bản chất khác), "
                    "có thể bỏ qua cảnh báo này.", nguon
                ))
            if gioi_tinh == "2" and any(k in ten_code for k in MALE_ONLY_KEYWORDS):
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    truong_loi, ma_code,
                    f"[Gợi ý theo từ khóa - cần xác nhận] Chỉ định '{ma_code}' ({ten_code.title()}) "
                    "có tên gợi ý dành cho bệnh nhân nam, nhưng GIOI_TINH của hồ sơ là Nữ (2).",
                    "Kiểm tra lại: nếu chỉ định này thực sự không phù hợp với giới tính người bệnh, "
                    "sửa lại chỉ định; nếu phù hợp (ví dụ tên dịch vụ trùng từ khóa nhưng bản chất khác), "
                    "có thể bỏ qua cảnh báo này.", nguon
                ))
            if tuoi is not None and tuoi >= 16 and any(k in ten_code for k in PEDIATRIC_KEYWORDS):
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    truong_loi, ma_code,
                    f"[Gợi ý theo từ khóa] Chỉ định '{ma_code}' ({ten_code.title()}) có dấu hiệu dành "
                    f"cho trẻ em, nhưng tuổi người bệnh ước tính ~{tuoi} tuổi (>=16).",
                    "Kiểm tra lại Ngày sinh và chỉ định/mã dịch vụ-thuốc cho phù hợp với độ tuổi người bệnh.", nguon
                ))

    # ---- 3.3 Trùng lặp chỉ định cùng loại trong ngày/đợt điều trị ----
    for nhom_item, label, key_field, name_field, nguon in (
        ("dvkt", "DVKT", "MA_DICH_VU", "TEN_DICH_VU", "XML3 - DVKT"),
        ("cdha", "CĐHA/TDCN", "MA_DICH_VU", "TEN_DICH_VU", "XML4 - CĐHA/TDCN"),
    ):
        ds = data.get(nhom_item, [])
        seen = {}
        for item in ds:
            ma_code = item.get(key_field)
            ten_code = item.get(name_field, "")
            ngay_yc = (item.get("NGAY_YL") or "")[:8]  # lấy phần ngày YYYYMMDD
            key = (ma_code, ngay_yc)
            if not ma_code:
                continue
            seen.setdefault(key, []).append(item)

        for (ma_code, ngay_yc), ds_dup in seen.items():
            if len(ds_dup) > 1:
                ten_code = ds_dup[0].get(name_field, "")
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    key_field, ma_code,
                    f"Chỉ định '{ma_code}' ({ten_code}) - nhóm {label} được lặp lại "
                    f"{len(ds_dup)} lần trong cùng ngày {ngay_yc or '(không rõ)'}.",
                    "Nếu việc thực hiện lại là cần thiết về chuyên môn (ví dụ theo dõi diễn biến, "
                    "có biến chứng), cần ghi rõ lý do y khoa trong bệnh án để tránh bị xuất toán do trùng chỉ định.", nguon
                ))

    # ---- 3.4 Liều dùng và số ngày điều trị (thuốc) ----
    nguon = "XML2 - Thuốc"
    ngay_ra = _text(ho_so_hc, "NGAY_RA")
    dt_ra = _to_datetime(ngay_ra)
    so_ngay_dieu_tri = None
    if dt_vao and dt_ra:
        so_ngay_dieu_tri = max((dt_ra - dt_vao).days, 1)

    for item in data.get("thuoc", []):
        ma_thuoc = item.get("MA_THUOC")
        ten_thuoc = item.get("TEN_THUOC", "")
        lieu_dung = item.get("LIEU_DUNG", "")
        duong_dung = item.get("DUONG_DUNG", "")
        so_luong = _to_float(item.get("SO_LUONG"))

        if not lieu_dung:
            errors.append(_err(
                ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                "LIEU_DUNG", "(thiếu)",
                f"Thuốc '{ma_thuoc}' ({ten_thuoc}) thiếu thông tin liều dùng (LIEU_DUNG).",
                "Bổ sung liều dùng/cách dùng (LIEU_DUNG, DUONG_DUNG) theo y lệnh để đối chiếu "
                "số lượng cấp phát với số ngày điều trị.", nguon
            ))
            continue

        # Tách số lần dùng/ngày từ chuỗi liều dùng dạng phổ biến, ví dụ:
        # "Uống ngày 2 lần, mỗi lần 1 viên" -> tổng = 2 viên/ngày
        so_lan = re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:lần|l)\b", lieu_dung, flags=re.IGNORECASE)
        so_moi_lan = re.findall(r"mỗi lần\s*(\d+(?:[.,]\d+)?)", lieu_dung, flags=re.IGNORECASE)

        if so_lan and so_moi_lan and so_luong is not None and so_ngay_dieu_tri:
            try:
                tong_ngay = float(so_lan[0].replace(",", ".")) * float(so_moi_lan[0].replace(",", "."))
                du_kien = tong_ngay * so_ngay_dieu_tri
                # Cho phép sai lệch +/- 20% (dự trù mang theo khi xuất viện, đợt điều trị ngắt quãng...)
                if so_luong > du_kien * 1.2:
                    errors.append(_err(
                        ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                        "SO_LUONG", so_luong,
                        f"Thuốc '{ma_thuoc}' ({ten_thuoc}): số lượng cấp = {so_luong:g}, theo liều dùng "
                        f"'{lieu_dung}' và {so_ngay_dieu_tri} ngày điều trị thì dự kiến chỉ cần "
                        f"~{du_kien:g}. Số lượng cấp đang vượt đáng kể so với dự kiến.",
                        "Kiểm tra lại số lượng thuốc kê/cấp so với liều dùng và số ngày điều trị "
                        "(NGAY_VAO - NGAY_RA); nếu có lý do hợp lý (mang về điều trị ngoại trú tiếp...) "
                        "cần ghi chú rõ trong bệnh án.", nguon
                    ))
            except (ValueError, ZeroDivisionError):
                pass

    # ---- 3.5 Thuốc chống chỉ định theo mã bệnh ICD-10 (hai chiều) ----
    thuoc_chong_chi_dinh = danh_muc.get("thuoc_chong_chi_dinh", [])
    ma_thuoc_hoat_chat = danh_muc.get("ma_thuoc_hoat_chat", {})
    if thuoc_chong_chi_dinh and cac_ma_benh:
        cac_ma_benh_nodot = {m.replace(".", "") for m in cac_ma_benh}

        for item in data.get("thuoc", []):
            ma_thuoc = item.get("MA_THUOC", "")
            ten_thuoc = item.get("TEN_THUOC", "")
            hoat_chat = ma_thuoc_hoat_chat.get(ma_thuoc, "") or (ten_thuoc or "").lower()

            for rule in thuoc_chong_chi_dinh:
                if rule["keyword"] not in hoat_chat:
                    continue
                canh_bao_added = False
                for ma_benh in cac_ma_benh_nodot:
                    for prefix in rule["icd_prefixes"]:
                        if ma_benh.startswith(prefix):
                            errors.append(_err(
                                ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                                "MA_THUOC", ma_thuoc,
                                f"[Gợi ý dược lý - cần xác nhận] Thuốc '{ma_thuoc}' ({ten_thuoc}) chứa "
                                f"hoạt chất liên quan '{rule['keyword']}', có thể chống chỉ định/thận trọng "
                                f"với mã bệnh '{ma_benh}' đã ghi nhận trong hồ sơ (MA_BENH_CHINH/MA_BENH_KT). "
                                f"{rule['ghi_chu']}",
                                "Rà soát lại chỉ định thuốc này theo hướng dẫn sử dụng/dược lý lâm sàng; "
                                "nếu việc sử dụng là có cơ sở (đã đánh giá lợi ích/nguy cơ), ghi rõ lý do "
                                "trong bệnh án. Nếu không phù hợp, xem xét đổi thuốc khác.", nguon
                            ))
                            canh_bao_added = True
                            break
                    if canh_bao_added:
                        break

    # ---- 3.6 Đối chiếu chéo DVKT (XML3) <-> Kết quả CĐHA-TDCN (XML4) ----
    ds_dvkt = data.get("dvkt", [])
    ds_cdha = data.get("cdha", [])
    ma_dvkt_set = {item.get("MA_DICH_VU") for item in ds_dvkt if item.get("MA_DICH_VU")}

    # Từ khóa nhận diện DVKT là xét nghiệm/CĐHA/TDCN cần có mã máy (MA_MAY ở XML3)
    DVKT_CAN_MA_MAY = [
        "XÉT NGHIỆM", "XET NGHIEM", "SIÊU ÂM", "SIEU AM",
        "X-QUANG", "XQUANG", "X QUANG", "CT ", "MRI",
        "CỘNG HƯỞNG TỪ", "CONG HUONG TU", "NỘI SOI", "NOI SOI",
        "ĐIỆN TIM", "DIEN TIM", "ĐIỆN NÃO", "DIEN NAO",
        "TDCN", "THĂM DÒ", "THAM DO", "ĐO CHỨC NĂNG", "DO CHUC NANG",
    ]

    # 3.6.a: MA_MAY ở XML3 - DVKT là xét nghiệm/CĐHA mà thiếu mã máy
    nguon = "XML3 - DVKT"
    for dv in ds_dvkt:
        ma_dv = dv.get("MA_DICH_VU", "")
        ten_dv = (dv.get("TEN_DICH_VU") or "").upper()
        ma_may = (dv.get("MA_MAY") or "").strip()

        # Chỉ kiểm tra khi DVKT có dấu hiệu là xét nghiệm/CĐHA
        la_xet_nghiem = any(k in ten_dv for k in DVKT_CAN_MA_MAY)
        if la_xet_nghiem and not ma_may:
            errors.append(_err(
                ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                "MA_MAY", "(thiếu)",
                f"DVKT '{ma_dv}' ({dv.get('TEN_DICH_VU', '')}) là xét nghiệm/CĐHA/TDCN "
                "nhưng chưa khai mã máy thực hiện (MA_MAY) trong XML3.",
                "Bổ sung MA_MAY (mã máy/thiết bị thực hiện) vào XML3 theo danh mục máy "
                "đã đăng ký với BHXH để tránh bị nghi vấn khi giám định.", nguon
            ))

    # 3.6.b: Mã DVKT trong kết quả XML4 phải có chỉ định tương ứng ở XML3
    nguon = "XML4 - CĐHA/TDCN"
    for item in ds_cdha:
        ma_dv_cdha = item.get("MA_DICH_VU", "")
        ten_dv_cdha = item.get("TEN_DICH_VU", "")
        ten_chi_so = item.get("TEN_CHI_SO", "")
        gia_tri = item.get("GIA_TRI", "")
        mo_ta = item.get("MO_TA", "")
        ket_luan = item.get("KET_LUAN", "")
        ma_bs_doc_kq = item.get("MA_BS_DOC_KQ", "")
        ngay_yl_item = item.get("NGAY_YL", "")
        ngay_th_yl_item = item.get("NGAY_TH_YL", "")
        ngay_kq = item.get("NGAY_KQ", "")
        ngay_yl_cdha = ngay_yl_item

        la_sa_dt_xq = any(k in (ten_chi_so or ten_dv_cdha or "").upper() for k in TEN_CHI_SO_CAN_BS_DOC_KQ)

        if ma_dv_cdha and ma_dvkt_set and ma_dv_cdha not in ma_dvkt_set:
            errors.append(_err(
                ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                "MA_DICH_VU", ma_dv_cdha,
                f"Kết quả CĐHA/TDCN (XML4) có mã DVKT '{ma_dv_cdha}' ({ten_dv_cdha}) "
                "nhưng không tìm thấy chỉ định tương ứng trong XML3 (Dịch vụ kỹ thuật) "
                "của cùng hồ sơ.",
                "Kiểm tra mã DVKT giữa XML3 và XML4 phải khớp nhau; nếu dịch vụ đã thực "
                "hiện thì phải có dòng tương ứng trong XML3, nếu không thì xoá dòng kết "
                "quả thừa trong XML4.", nguon
            ))

        # 3.6.d: Siêu âm/Điện tim/X-quang -> GIA_TRI phải TRỐNG, MO_TA + KET_LUAN bắt buộc có
        if la_sa_dt_xq:
            if gia_tri:
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                    "GIA_TRI", gia_tri,
                    f"'{ten_chi_so or ten_dv_cdha}' là Siêu âm/Điện tim/X-quang - không được khai "
                    f"GIA_TRI (chỉ áp dụng cho xét nghiệm định lượng), hiện đang có giá trị '{gia_tri}'.",
                    "Xoá giá trị GIA_TRI ở dòng này; mô tả/kết luận của SA-ĐT-XQ ghi vào MO_TA/KET_LUAN.", nguon
                ))
            if not mo_ta or not ket_luan:
                thieu = []
                if not mo_ta:
                    thieu.append("MO_TA")
                if not ket_luan:
                    thieu.append("KET_LUAN")
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                    "/".join(thieu), "(thiếu)",
                    f"'{ten_chi_so or ten_dv_cdha}' là Siêu âm/Điện tim/X-quang nhưng thiếu "
                    f"{' và '.join(thieu)} - đây là trường bắt buộc cho kết quả CĐHA/TDCN.",
                    "Bổ sung đầy đủ MO_TA (mô tả hình ảnh/kết quả) và KET_LUAN (kết luận chẩn đoán).", nguon
                ))

        # 3.6.e: Bác sĩ đọc kết quả (MA_BS_DOC_KQ) - SA/ĐT/XQ cần Bác sĩ (CHUCDANH_NN=1),
        # các loại khác cần kỹ thuật viên (CHUCDANH_NN=6,10)
        if ma_bs_doc_kq:
            if la_sa_dt_xq:
                trang_thai, ghi_chu = _kiem_tra_nguoi_ke_don(ma_bs_doc_kq, nhan_vien_yte)
                # _kiem_tra_nguoi_ke_don kiểm tra CHUCDANH_NN=1 (Bác sĩ) - đúng yêu cầu cho SA/ĐT/XQ
                yeu_cau = "Bác sĩ (CHUCDANH_NN=1)"
            else:
                trang_thai, ghi_chu = _kiem_tra_nguoi_thuc_hien_dvkt(ma_bs_doc_kq, nhan_vien_yte, ma_dv=ma_dv_cdha)
                yeu_cau = "Kỹ thuật viên (CHUCDANH_NN=6,10)"

            if trang_thai == "khong_tim_thay":
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    "MA_BS_DOC_KQ", ma_bs_doc_kq,
                    f"Không tìm thấy mã chứng chỉ hành nghề '{ma_bs_doc_kq}' (người đọc kết quả "
                    f"'{ten_chi_so or ten_dv_cdha}') trong danh mục nhân viên y tế.",
                    "Kiểm tra lại MA_BS_DOC_KQ hoặc bổ sung nhân viên vào danh mục nhân viên y tế.", nguon
                ))
            elif trang_thai in ("sai_pham_vi", "khong_dat", "khong_xac_dinh_cd"):
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    "MA_BS_DOC_KQ", ma_bs_doc_kq,
                    f"Người đọc kết quả '{ten_chi_so or ten_dv_cdha}' - {ghi_chu} - không đáp ứng yêu "
                    f"cầu phải là {yeu_cau}.",
                    "Kiểm tra lại người đọc/ký kết quả cho đúng phạm vi chuyên môn quy định.", nguon
                ))

        # 3.6.c: Ngày kết quả (NGAY_KQ) phải >= ngày chỉ định (NGAY_YL) tương ứng ở XML3,
        # và nếu XML4 có sẵn NGAY_YL/NGAY_TH_YL riêng thì kiểm tra NGAY_KQ > NGAY_TH_YL >= NGAY_YL
        dt_kq = _to_datetime(ngay_kq) if ngay_kq else None
        dt_yl_item = _to_datetime(ngay_yl_item) if ngay_yl_item else None
        dt_th_yl_item = _to_datetime(ngay_th_yl_item) if ngay_th_yl_item else None

        if dt_yl_item and dt_th_yl_item and dt_th_yl_item < dt_yl_item:
            errors.append(_err(
                ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                "NGAY_TH_YL", ngay_th_yl_item,
                f"'{ten_chi_so or ten_dv_cdha}': Ngày thực hiện y lệnh ({ngay_th_yl_item}) trước "
                f"ngày chỉ định (NGAY_YL={ngay_yl_item}) - không hợp lý.",
                "Kiểm tra lại NGAY_YL và NGAY_TH_YL trong XML4, đảm bảo NGAY_TH_YL >= NGAY_YL.", nguon
            ))
        if dt_th_yl_item and dt_kq and dt_kq <= dt_th_yl_item:
            errors.append(_err(
                ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                "NGAY_KQ", ngay_kq,
                f"'{ten_chi_so or ten_dv_cdha}': Ngày có kết quả ({ngay_kq}) phải lớn hơn ngày "
                f"thực hiện y lệnh (NGAY_TH_YL={ngay_th_yl_item}).",
                "Kiểm tra lại NGAY_TH_YL và NGAY_KQ trong XML4, đảm bảo NGAY_KQ > NGAY_TH_YL.", nguon
            ))

        if dt_kq and not dt_th_yl_item:
            # Không có NGAY_TH_YL riêng ở XML4 -> đối chiếu chéo với NGAY_YL của DVKT (XML3) như trước
            ngay_yl_dvkt = None
            for dv in ds_dvkt:
                if dv.get("MA_DICH_VU") == ma_dv_cdha and dv.get("NGAY_YL"):
                    ngay_yl_dvkt = dv.get("NGAY_YL")
                    break
            if not ngay_yl_dvkt:
                ngay_yl_dvkt = ngay_yl_cdha

            dt_yl = _to_datetime(ngay_yl_dvkt) if ngay_yl_dvkt else None
            if dt_yl and dt_kq < dt_yl:
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                    "NGAY_KQ", ngay_kq,
                    f"Ngày có kết quả ({ngay_kq}) của '{ten_dv_cdha or ma_dv_cdha}' "
                    f"({ten_chi_so}) trước ngày chỉ định thực hiện ({ngay_yl_dvkt}) - "
                    "không hợp lý.",
                    "Kiểm tra lại NGAY_YL (ngày chỉ định/thực hiện ở XML3) và NGAY_KQ "
                    "(ngày có kết quả ở XML4), đảm bảo NGAY_KQ >= NGAY_YL.", nguon
                ))

    # ---- 3.7 Phẫu thuật/thủ thuật (XML5): THOI_DIEM_DBLS > NGAY_VAO, NGUOI_THUC_HIEN phải là Bác sĩ ----
    nguon = "XML5 - Phẫu thuật thủ thuật"
    for item in data.get("pttt", []):
        ma_pttt = item.get("MA_PTTT", "")
        ten_pttt = item.get("TEN_PTTT", "")
        thoi_diem_dbls = item.get("THOI_DIEM_DBLS", "")
        nguoi_th_pttt = item.get("NGUOI_THUC_HIEN", "")

        dt_dbls = _to_datetime(thoi_diem_dbls) if thoi_diem_dbls else None
        if dt_dbls and dt_vao and dt_dbls <= dt_vao:
            errors.append(_err(
                ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                "THOI_DIEM_DBLS", thoi_diem_dbls,
                f"PTTT '{ma_pttt}' ({ten_pttt}): Thời điểm bắt đầu phẫu thuật/thủ thuật "
                f"({thoi_diem_dbls}) phải lớn hơn Ngày vào viện (NGAY_VAO={ngay_vao}).",
                "Kiểm tra lại THOI_DIEM_DBLS trong XML5 và NGAY_VAO trong XML1, đảm bảo "
                "THOI_DIEM_DBLS > NGAY_VAO.", nguon
            ))

        if nguoi_th_pttt:
            trang_thai, ghi_chu = _kiem_tra_nguoi_ke_don(nguoi_th_pttt, nhan_vien_yte)
            if trang_thai == "khong_tim_thay":
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Cảnh báo",
                    "NGUOI_THUC_HIEN", nguoi_th_pttt,
                    f"Không tìm thấy mã chứng chỉ hành nghề '{nguoi_th_pttt}' (người thực hiện "
                    f"PTTT '{ma_pttt}' - {ten_pttt}) trong danh mục nhân viên y tế.",
                    "Kiểm tra lại NGUOI_THUC_HIEN hoặc bổ sung nhân viên vào danh mục nhân viên y tế.", nguon
                ))
            elif trang_thai in ("sai_pham_vi", "khong_xac_dinh_cd"):
                errors.append(_err(
                    ma_lk, "3. Hợp lý chỉ định y khoa", "Lỗi",
                    "NGUOI_THUC_HIEN", nguoi_th_pttt,
                    f"Người thực hiện PTTT '{ma_pttt}' ({ten_pttt}) - {ghi_chu} - không phải Bác sĩ "
                    "(CHUCDANH_NN=1). Phẫu thuật/thủ thuật phải do Bác sĩ thực hiện.",
                    "Kiểm tra lại người thực hiện (NGUOI_THUC_HIEN) trong XML5.", nguon
                ))

    return errors

def check_nhom4_thoi_gian_trung_lap(ho_so, file_type, all_hanh_chinh=None):
    """
    all_hanh_chinh: list các dict thông tin hành chính đã đọc, mỗi dict gồm
                    ít nhất MA_LK, MA_THE_BHYT, NGAY_VAO, NGAY_RA, MA_CSKCB
                    -> dùng để phát hiện trùng lịch KCB cùng mã thẻ, các
                       khoảng thời gian giao nhau.
    Hàm này được gọi riêng cho XML1 - Hành chính, kiểm tra:
      4.1 Thời gian đẩy hồ sơ (NGAY_TT so với NGAY_RA)
      4.2 Trùng lịch KCB (chồng lấp thời gian, cùng mã thẻ, khác MA_LK)
      4.3 Tách đợt điều trị ngoại trú trong cùng 1 ngày tại cùng cơ sở
    """
    errors = []
    nguon = "XML1 - Hành chính"

    ma_lk_tag = ho_so.find('MA_LK')
    ma_lk = ma_lk_tag.text.strip().upper() if ma_lk_tag is not None and ma_lk_tag.text else "Không xác định"

    ngay_vao = _text(ho_so, "NGAY_VAO")
    ngay_ra = _text(ho_so, "NGAY_RA")
    ma_the = _text(ho_so, "MA_THE_BHYT")
    ma_cskcb = _text(ho_so, "MA_CSKCB")

    dt_vao = _to_datetime(ngay_vao)
    dt_ra = _to_datetime(ngay_ra)

    # ---- 4.1 Thời hạn đẩy hồ sơ ----
    ngay_tt = _text(ho_so, "NGAY_TT")  # ngày thanh toán/quyết toán (nếu có khai trong XML1)
    if ngay_tt and dt_ra:
        dt_tt = _to_datetime(ngay_tt)
        if dt_tt:
            so_ngay_cham = (dt_tt - dt_ra).days
            if so_ngay_cham > 10:
                errors.append(_err(
                    ma_lk, "4. Thời gian & trùng lặp KCB", "Cảnh báo",
                    "NGAY_TT", ngay_tt,
                    f"Khoảng cách từ ngày ra viện ({ngay_ra}) đến ngày quyết toán ({ngay_tt}) là "
                    f"{so_ngay_cham} ngày, có thể vượt thời hạn gửi dữ liệu lên Cổng giám định.",
                    "Đẩy hồ sơ lên Cổng tiếp nhận ngay sau khi người bệnh ra viện theo đúng thời hạn "
                    "quy định; nếu chậm cần kèm văn bản giải trình lý do.", nguon
                ))

    # ---- 4.2 & 4.3: cần dữ liệu tổng hợp tất cả hồ sơ ----
    if all_hanh_chinh and dt_vao and dt_ra and ma_the:
        for other in all_hanh_chinh:
            if other["ma_lk"] == ma_lk:
                continue
            if other["ma_the"] != ma_the:
                continue
            o_vao, o_ra = other["dt_vao"], other["dt_ra"]
            if not (o_vao and o_ra):
                continue

            # Kiểm tra giao nhau về thời gian
            giao_nhau = dt_vao <= o_ra and o_vao <= dt_ra
            if giao_nhau:
                if other["ma_cskcb"] != ma_cskcb:
                    errors.append(_err(
                        ma_lk, "4. Thời gian & trùng lặp KCB", "Lỗi",
                        "NGAY_VAO/NGAY_RA", f"{ngay_vao} - {ngay_ra}",
                        f"Khoảng thời gian KCB ({ngay_vao} - {ngay_ra}) trùng/giao với một hồ sơ khác "
                        f"(MA_LK={other['ma_lk']}) tại cơ sở khác (MA_CSKCB={other['ma_cskcb']}) "
                        f"cùng mã thẻ BHYT {ma_the}.",
                        "Kiểm tra lại thời gian vào/ra viện; nếu là khám chuyên khoa đặc thù được phép "
                        "song song (chạy thận, hoá trị...) cần lưu hồ sơ chứng minh; nếu không, "
                        "phối hợp với cơ sở liên quan để xác minh và điều chỉnh.", nguon
                    ))
                else:
                    # Cùng 1 cơ sở, cùng ngày, nhiều đợt khám -> kiểm tra tách đợt ngoại trú
                    if dt_vao.date() == o_vao.date() and ngay_vao != other["ngay_vao"]:
                        errors.append(_err(
                            ma_lk, "4. Thời gian & trùng lặp KCB", "Cảnh báo",
                            "NGAY_VAO", ngay_vao,
                            f"Cùng ngày {dt_vao.date()} tại cùng cơ sở (MA_CSKCB={ma_cskcb}), người bệnh "
                            f"có nhiều hồ sơ KCB (MA_LK={ma_lk} và {other['ma_lk']}).",
                            "Khám nhiều chuyên khoa trong cùng một ngày tại một cơ sở chỉ được tính là "
                            "một đợt khám (trừ trường hợp được phép tách riêng theo quy định). "
                            "Kiểm tra gộp thành 1 đợt khám nếu không thuộc trường hợp ngoại lệ.", nguon
                        ))

    return errors


# =====================================================================
#  HÀM TỔNG HỢP: CHẠY TẤT CẢ QUY TẮC CHO 1 HỒ SƠ HÀNH CHÍNH
#  (dùng cho nhóm 1 và 4, gọi từ vòng lặp chính trong app)
# =====================================================================

def check_hanh_chinh_full(ho_so, all_hanh_chinh=None, danh_muc=None):
    ma_lk_tag = ho_so.find('MA_LK')
    ma_lk = ma_lk_tag.text.strip().upper() if ma_lk_tag is not None and ma_lk_tag.text else "Không xác định"

    errors = []
    errors.extend(check_nhom1_hanh_chinh(ho_so, ma_lk, danh_muc=danh_muc))
    errors.extend(check_nhom4_thoi_gian_trung_lap(ho_so, "Hành chính", all_hanh_chinh=all_hanh_chinh))
    return errors


# =====================================================================
#  CẬP NHẬT DANH MỤC TỪ FILE GỐC (do người dùng tải lên 1 lần,
#  tự động lọc và GHI ĐÈ các file danh mục cố định kèm sẵn ứng dụng)
# =====================================================================

def _save_gia_excel(data_dict, path):
    """Lưu dict {MA: GIA} ra file Excel 2 cột MA, GIA."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["MA", "GIA"])
    for ma, gia in data_dict.items():
        ws.append([ma, gia])
    wb.save(path)


def rebuild_ma_thuoc_hoatchat_from_source(file_bytes_or_path):
    """
    Nhận file Excel danh mục thuốc GỐC (cấu trúc như FileDanhMucThuoc.xlsx, có
    các cột MA_THUOC, TEN_THUOC, TEN_HOAT_CHAT, và tùy chọn TU_NGAY để lấy bản
    ghi mới nhất cho mỗi mã thuốc).

    Tự động lọc ra bảng tra MA_THUOC -> TEN_HOAT_CHAT -> ghi đè
    danh_muc_ma_thuoc_hoatchat.xlsx (dùng cho quy tắc 3.5 - tương tác/chống chỉ
    định thuốc theo ICD-10).

    Không còn xử lý giá thuốc (DON_GIA_BH) - việc kiểm tra giá thuốc đã được bỏ.

    Trả về dict {"so_ma_hoat_chat": int}. Raise Exception nếu file không đúng
    cấu trúc tối thiểu.
    """
    wb = load_workbook(file_bytes_or_path, data_only=True)
    sheet = wb.active
    headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]

    required = ["MA_THUOC", "TEN_THUOC", "TEN_HOAT_CHAT"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(
            "File danh mục thuốc thiếu các cột bắt buộc: " + ", ".join(missing) +
            ". Cần có đủ các cột: MA_THUOC, TEN_THUOC, TEN_HOAT_CHAT "
            "(và TU_NGAY nếu muốn lấy bản ghi mới nhất khi 1 mã có nhiều dòng)."
        )

    idx = {h: headers.index(h) for h in headers if h}
    idx_ma = idx["MA_THUOC"]
    idx_ten = idx["TEN_THUOC"]
    idx_hc = idx["TEN_HOAT_CHAT"]
    idx_tungay = idx.get("TU_NGAY")

    # Gom theo MA_THUOC, nếu có TU_NGAY thì giữ bản ghi có TU_NGAY lớn nhất (mới nhất)
    records = {}  # ma_thuoc -> (tu_ngay_sortkey, ten_thuoc, hoat_chat)
    for r in range(2, sheet.max_row + 1):
        ma = sheet.cell(r, idx_ma + 1).value
        if ma is None or str(ma).strip() == "":
            continue
        ma = str(ma).strip()
        ten = sheet.cell(r, idx_ten + 1).value
        hc = sheet.cell(r, idx_hc + 1).value

        sort_key = ""
        if idx_tungay is not None:
            tu_ngay_val = sheet.cell(r, idx_tungay + 1).value
            sort_key = str(tu_ngay_val) if tu_ngay_val is not None else ""

        prev = records.get(ma)
        if prev is None or sort_key >= prev[0]:
            records[ma] = (sort_key, str(ten).strip() if ten else "",
                            str(hc).strip() if hc else "")

    from openpyxl import Workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append(["MA_THUOC", "TEN_THUOC", "TEN_HOAT_CHAT"])
    for ma, rec in records.items():
        ws_out.append([ma, rec[1], rec[2]])
    wb_out.save(DEFAULT_MA_THUOC_HOATCHAT_FILE)

    return {"so_ma_hoat_chat": len(records)}


def rebuild_gia_dvkt_from_source(file_bytes_or_path):
    """
    Nhận file Excel danh mục DVKT/Giường GỐC (cấu trúc như FileDichVuBV.xlsx,
    có các cột MA_TUONG_DUONG, DON_GIA, và tùy chọn TUNGAY để lấy bản ghi mới
    nhất cho mỗi mã).

    Tự động lọc ra danh mục giá DVKT (MA, GIA) -> ghi đè danh_muc_gia_dvkt.xlsx.

    Trả về dict {"so_ma_gia_dvkt": int}. Raise Exception nếu file không đúng
    cấu trúc tối thiểu.
    """
    wb = load_workbook(file_bytes_or_path, data_only=True)
    sheet = wb.active
    headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]

    required = ["MA_TUONG_DUONG", "DON_GIA"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(
            "File danh mục DVKT thiếu các cột bắt buộc: " + ", ".join(missing) +
            ". Cần có đủ các cột: MA_TUONG_DUONG, DON_GIA "
            "(và TUNGAY nếu muốn lấy bản ghi mới nhất khi 1 mã có nhiều dòng)."
        )

    idx = {h: headers.index(h) for h in headers if h}
    idx_ma = idx["MA_TUONG_DUONG"]
    idx_gia = idx["DON_GIA"]
    idx_tungay = idx.get("TUNGAY")

    records = {}  # ma -> (sort_key, gia)
    for r in range(2, sheet.max_row + 1):
        ma = sheet.cell(r, idx_ma + 1).value
        gia = sheet.cell(r, idx_gia + 1).value
        if ma is None or str(ma).strip() == "":
            continue
        try:
            gia_f = float(gia)
        except (TypeError, ValueError):
            continue
        ma = str(ma).strip()

        sort_key = ""
        if idx_tungay is not None:
            tu_ngay_val = sheet.cell(r, idx_tungay + 1).value
            sort_key = str(tu_ngay_val) if tu_ngay_val is not None else ""

        prev = records.get(ma)
        if prev is None or sort_key >= prev[0]:
            records[ma] = (sort_key, gia_f)

    gia_dvkt = {ma: rec[1] for ma, rec in records.items()}
    _save_gia_excel(gia_dvkt, DEFAULT_GIA_DVKT_FILE)

    return {"so_ma_gia_dvkt": len(gia_dvkt)}


def rebuild_nhan_vien_yte_from_source(file_bytes_or_path):
    """
    Nhận file Excel danh mục nhân viên y tế GỐC (cấu trúc như FileNhanVienYTe.xlsx,
    cần có cột MACCHN, HO_TEN, CHUCDANH_NN; tùy chọn DVKT_KHAC, VB_PHANCONG).

    Tự động lọc và ghi đè danh_muc_nhan_vien_yte.xlsx (5 cột rút gọn).

    Trả về dict {"so_nhan_vien": int}. Raise ValueError nếu thiếu cột bắt buộc.
    """
    wb = load_workbook(file_bytes_or_path, data_only=True)
    sheet = wb.active
    headers = [str(c.value).strip().upper() if c.value else "" for c in sheet[1]]

    required = ["MACCHN", "HO_TEN", "CHUCDANH_NN"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(
            "File danh mục nhân viên y tế thiếu các cột bắt buộc: " + ", ".join(missing) +
            ". Cần có đủ các cột: MACCHN, HO_TEN, CHUCDANH_NN (và tùy chọn DVKT_KHAC, VB_PHANCONG)."
        )

    idx = {h: headers.index(h) for h in headers if h}
    idx_ma = idx["MACCHN"]
    idx_ten = idx["HO_TEN"]
    idx_cd = idx["CHUCDANH_NN"]
    idx_dvkt = idx.get("DVKT_KHAC")
    idx_vb = idx.get("VB_PHANCONG")

    from openpyxl import Workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append(["MACCHN", "HO_TEN", "CHUCDANH_NN", "DVKT_KHAC", "VB_PHANCONG"])

    count = 0
    for r in range(2, sheet.max_row + 1):
        macchn = sheet.cell(r, idx_ma + 1).value
        if macchn is None or str(macchn).strip() == "":
            continue
        ten = sheet.cell(r, idx_ten + 1).value
        cd = sheet.cell(r, idx_cd + 1).value
        try:
            cd = str(int(float(cd))) if cd is not None else ""
        except (TypeError, ValueError):
            cd = str(cd).strip() if cd else ""
        dvkt = sheet.cell(r, idx_dvkt + 1).value if idx_dvkt is not None else None
        vb = sheet.cell(r, idx_vb + 1).value if idx_vb is not None else None

        ws_out.append([
            str(macchn).strip(),
            str(ten).strip() if ten else "",
            cd,
            str(dvkt).strip() if dvkt else "",
            str(vb).strip() if vb else "",
        ])
        count += 1

    wb_out.save(DEFAULT_NHAN_VIEN_YTE_FILE)
    return {"so_nhan_vien": count}
