import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re


# ====================== HELPER FUNCTIONS ======================

def read_excel_file(uploaded_file, sheet_name=0):
    try:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, dtype=str)
        df = df.fillna("")
        df.columns = [str(c).strip() for c in df.columns]
        return df, None
    except Exception as e:
        return None, str(e)


def get_sheet_names(uploaded_file):
    try:
        uploaded_file.seek(0)
        xl = pd.ExcelFile(uploaded_file)
        return xl.sheet_names
    except Exception:
        return ["Sheet1"]


def to_excel_bytes(df, sheet_name="Sheet1", highlight_rows=None, highlight_color="FFD7D7"):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        if highlight_rows:
            hl_fill = PatternFill("solid", fgColor=highlight_color)
            for row_idx in highlight_rows:
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx + 2, column=col).fill = hl_fill
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    output.seek(0)
    return output


def multi_sheet_excel(sheets_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sname, df in sheets_dict.items():
            sname_safe = sname[:31]
            df.to_excel(writer, index=False, sheet_name=sname_safe)
            ws = writer.sheets[sname_safe]
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    output.seek(0)
    return output


def dl_btn(label, data, filename, key):
    st.download_button(label, data, filename, key=key, use_container_width=True)


# ====================== FEATURE LOGIC FUNCTIONS ======================

def feature_compare_two_files(df1, df2, key_col1, key_col2, compare_cols1, compare_cols2):
    results = {}
    df1_keys = set(df1[key_col1].str.strip().str.upper())
    df2_keys = set(df2[key_col2].str.strip().str.upper())
    only_in_f1 = df1_keys - df2_keys
    only_in_f2 = df2_keys - df1_keys
    in_both = df1_keys & df2_keys
    results["only_in_file1"] = df1[df1[key_col1].str.strip().str.upper().isin(only_in_f1)].copy()
    results["only_in_file2"] = df2[df2[key_col2].str.strip().str.upper().isin(only_in_f2)].copy()
    diff_rows = []
    if compare_cols1 and compare_cols2 and len(compare_cols1) == len(compare_cols2):
        df1_idx = df1.set_index(df1[key_col1].str.strip().str.upper())
        df2_idx = df2.set_index(df2[key_col2].str.strip().str.upper())
        for key in in_both:
            if key not in df1_idx.index or key not in df2_idx.index:
                continue
            row1 = df1_idx.loc[key]
            row2 = df2_idx.loc[key]
            diffs = []
            for c1, c2 in zip(compare_cols1, compare_cols2):
                v1 = str(row1[c1]).strip() if c1 in row1.index else ""
                v2 = str(row2[c2]).strip() if c2 in row2.index else ""
                if v1 != v2:
                    diffs.append(f"{c1}: [{v1}] ≠ [{v2}]")
            if diffs:
                diff_rows.append({"Khóa": key, "Điểm khác biệt": " | ".join(diffs), "Số trường khác": len(diffs)})
    results["differences"] = pd.DataFrame(diff_rows) if diff_rows else pd.DataFrame()
    results["stats"] = {
        "Tổng file 1": len(df1), "Tổng file 2": len(df2),
        "Chỉ có trong file 1": len(only_in_f1), "Chỉ có trong file 2": len(only_in_f2),
        "Có trong cả 2": len(in_both), "Dòng có khác biệt": len(diff_rows)
    }
    return results


def feature_find_duplicates(df, subset_cols, keep="first"):
    dup_mask = df.duplicated(subset=subset_cols, keep=False)
    duplicates = df[dup_mask].copy()
    if not duplicates.empty:
        duplicates["__Nhóm_trùng__"] = duplicates.groupby(subset_cols).ngroup() + 1
        duplicates = duplicates.sort_values("__Nhóm_trùng__")
    cleaned = df.drop_duplicates(subset=subset_cols, keep=keep).copy()
    removed = df[df.duplicated(subset=subset_cols, keep=keep)].copy()
    return duplicates, cleaned, removed


def feature_vlookup(df_main, df_lookup, main_key, lookup_key, lookup_cols):
    df_lookup_sub = df_lookup[[lookup_key] + lookup_cols].drop_duplicates(subset=[lookup_key])
    merge_map = {c: f"{c}_lookup" for c in lookup_cols if c in df_main.columns}
    df_lookup_sub = df_lookup_sub.rename(columns=merge_map)
    result = df_main.merge(df_lookup_sub, left_on=main_key, right_on=lookup_key, how="left")
    not_found = result[result[lookup_key].isna()][main_key].tolist()
    return result, not_found


def feature_filter_advanced(df, conditions):
    mask = pd.Series([True] * len(df), index=df.index)
    for cond in conditions:
        col, op, val = cond.get("col"), cond.get("operator"), cond.get("value", "")
        if col not in df.columns:
            continue
        series = df[col].astype(str).str.strip()
        if op == "=":
            mask &= series.str.upper() == str(val).strip().upper()
        elif op == "≠":
            mask &= series.str.upper() != str(val).strip().upper()
        elif op == "chứa":
            mask &= series.str.contains(str(val), case=False, na=False)
        elif op == "không chứa":
            mask &= ~series.str.contains(str(val), case=False, na=False)
        elif op == "bắt đầu bằng":
            mask &= series.str.startswith(str(val), na=False)
        elif op == "trống":
            mask &= (series == "") | (series.str.upper() == "NAN")
        elif op == "không trống":
            mask &= (series != "") & (series.str.upper() != "NAN")
        elif op == ">":
            try:
                mask &= pd.to_numeric(df[col], errors="coerce") > float(val)
            except Exception:
                pass
        elif op == "<":
            try:
                mask &= pd.to_numeric(df[col], errors="coerce") < float(val)
            except Exception:
                pass
    return df[mask].copy()


def feature_summary_stats(df, group_cols, agg_cols):
    agg_dict = {}
    for col in agg_cols:
        df[col + "_num"] = pd.to_numeric(df[col], errors="coerce")
        agg_dict[col + "_num"] = ["count", "sum", "mean", "min", "max"]
    grouped = df.groupby(group_cols).agg(agg_dict).round(2)
    grouped.columns = ["_".join(c).strip() for c in grouped.columns]
    return grouped.reset_index()


def feature_check_missing_values(df):
    report = []
    for col in df.columns:
        empty = ((df[col].astype(str).str.strip() == "") | (df[col].astype(str).str.upper() == "NAN")).sum()
        if empty > 0:
            report.append({"Cột": col, "Số dòng trống": empty, "Tỷ lệ (%)": round(empty / len(df) * 100, 1)})
    return pd.DataFrame(report)


def feature_normalize_data(df, cols, upper=False, lower=False, remove_spaces=False, remove_special=False):
    result = df.copy()
    for col in cols:
        if col not in result.columns:
            continue
        s = result[col].astype(str).str.strip()
        if upper:
            s = s.str.upper()
        elif lower:
            s = s.str.lower()
        if remove_spaces:
            s = s.str.replace(r'\s+', ' ', regex=True)
        if remove_special:
            s = s.str.replace(r'[^a-zA-Z0-9\s\u00C0-\u024F\u1E00-\u1EFF]', '', regex=True)
        result[col] = s
    return result


def feature_cross_check(df1, df2, check_col1, check_col2):
    vals1 = set(df1[check_col1].astype(str).str.strip().str.upper())
    vals2 = set(df2[check_col2].astype(str).str.strip().str.upper())
    missing_in_f2 = vals1 - vals2
    missing_in_f1 = vals2 - vals1
    rows_missing = df1[df1[check_col1].astype(str).str.strip().str.upper().isin(missing_in_f2)].copy()
    return rows_missing, len(missing_in_f2), len(missing_in_f1)


def feature_concat_files(dfs, add_source=True, source_names=None):
    all_dfs = []
    for i, df in enumerate(dfs):
        d = df.copy()
        if add_source:
            name = source_names[i] if source_names and i < len(source_names) else f"File {i+1}"
            d.insert(0, "__Nguồn__", name)
        all_dfs.append(d)
    return pd.concat(all_dfs, ignore_index=True, sort=False)


# ===== NEW: NHÓM TÀI CHÍNH / BHYT =====

def feature_doi_chieu_bhyt(df, col_ma_the, col_tien_bhyt, col_tien_tu_tra, col_tong=None):
    """Đối chiếu BHYT vs tự trả, kiểm tra tổng"""
    result = df.copy()
    result["__BHYT_num__"] = pd.to_numeric(result[col_tien_bhyt], errors="coerce").fillna(0)
    result["__TuTra_num__"] = pd.to_numeric(result[col_tien_tu_tra], errors="coerce").fillna(0)
    result["__Tong_tinh__"] = result["__BHYT_num__"] + result["__TuTra_num__"]
    issues = []
    if col_tong:
        result["__Tong_goc__"] = pd.to_numeric(result[col_tong], errors="coerce").fillna(0)
        result["__Chenh_lech__"] = (result["__Tong_tinh__"] - result["__Tong_goc__"]).round(0)
        issues_df = result[result["__Chenh_lech__"].abs() > 1].copy()
        issues = issues_df
    summary = {
        "Tổng bệnh nhân": len(df),
        "Tổng tiền BHYT": result["__BHYT_num__"].sum(),
        "Tổng tiền tự trả": result["__TuTra_num__"].sum(),
        "Tổng cộng": result["__Tong_tinh__"].sum(),
    }
    return result, issues, summary


def feature_kiem_tra_vuot_tran(df, col_khoa, col_tien, nguong_dict=None):
    """Kiểm tra chi phí vượt trần theo khoa/loại"""
    result = df.copy()
    result["__Tien_num__"] = pd.to_numeric(result[col_tien], errors="coerce").fillna(0)
    tong_theo_khoa = result.groupby(col_khoa)["__Tien_num__"].agg(
        Tổng_chi_phí="sum", Số_hồ_sơ="count", Trung_bình="mean"
    ).round(0).reset_index()
    tong_theo_khoa.columns = [col_khoa, "Tổng chi phí (đ)", "Số hồ sơ", "Chi phí TB/hồ sơ (đ)"]
    if nguong_dict:
        tong_theo_khoa["Ngưỡng trần (đ)"] = tong_theo_khoa[col_khoa].map(nguong_dict).fillna(0)
        tong_theo_khoa["Vượt trần (đ)"] = (tong_theo_khoa["Tổng chi phí (đ)"] - tong_theo_khoa["Ngưỡng trần (đ)"]).clip(lower=0)
        tong_theo_khoa["Trạng thái"] = tong_theo_khoa["Vượt trần (đ)"].apply(lambda x: "⚠️ Vượt trần" if x > 0 else "✅ Trong trần")
    return tong_theo_khoa


def feature_bao_cao_tong_hop(df, col_ngay, col_tien_bhyt, col_tien_tu_tra, col_khoa=None):
    """Báo cáo tổng hợp theo tháng"""
    result = df.copy()
    # Parse date
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]:
        try:
            result["__Ngay__"] = pd.to_datetime(result[col_ngay], format=fmt, errors="coerce")
            if result["__Ngay__"].notna().sum() > 0:
                break
        except Exception:
            pass
    result["__Thang__"] = result["__Ngay__"].dt.to_period("M").astype(str)
    result["__BHYT__"] = pd.to_numeric(result[col_tien_bhyt], errors="coerce").fillna(0)
    result["__TuTra__"] = pd.to_numeric(result[col_tien_tu_tra], errors="coerce").fillna(0)
    result["__Tong__"] = result["__BHYT__"] + result["__TuTra__"]
    group_cols = ["__Thang__"]
    if col_khoa:
        group_cols.append(col_khoa)
    bao_cao = result.groupby(group_cols).agg(
        Số_hồ_sơ=("__Tong__", "count"),
        Tiền_BHYT=("__BHYT__", "sum"),
        Tiền_tự_trả=("__TuTra__", "sum"),
        Tổng_chi_phí=("__Tong__", "sum"),
    ).round(0).reset_index()
    bao_cao.rename(columns={"__Thang__": "Tháng", col_khoa if col_khoa else "__Thang__": col_khoa or "Tháng"}, inplace=True)
    return bao_cao


# ===== NEW: NHÓM THUỐC & VẬT TƯ =====

def feature_kiem_tra_han_su_dung(df, col_ten, col_han, ngay_canh_bao=90):
    """Kiểm tra thuốc/vật tư sắp hết hạn hoặc đã hết hạn"""
    result = df.copy()
    parsed = None
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]:
        try:
            parsed = pd.to_datetime(result[col_han], format=fmt, errors="coerce")
            if parsed.notna().sum() > len(result) * 0.3:
                break
        except Exception:
            pass
    if parsed is None:
        parsed = pd.to_datetime(result[col_han], errors="coerce")
    result["__HanDate__"] = parsed
    today = pd.Timestamp.today().normalize()
    result["__ConLai__"] = (result["__HanDate__"] - today).dt.days
    da_het = result[result["__ConLai__"] < 0].copy()
    sap_het = result[(result["__ConLai__"] >= 0) & (result["__ConLai__"] <= ngay_canh_bao)].copy()
    con_han = result[result["__ConLai__"] > ngay_canh_bao].copy()
    da_het["__Trạng_thái__"] = "❌ Đã hết hạn"
    sap_het["__Trạng_thái__"] = sap_het["__ConLai__"].apply(lambda x: f"⚠️ Còn {int(x)} ngày")
    con_han["__Trạng_thái__"] = "✅ Còn hạn"
    return da_het, sap_het, con_han


def feature_doi_chieu_danh_muc_bhyt(df_don, df_dm, col_ma_don, col_ma_dm, col_ten_don=None):
    """Đối chiếu thuốc/dịch vụ trong đơn với danh mục BHYT được chi trả"""
    dm_keys = set(df_dm[col_ma_dm].astype(str).str.strip().str.upper())
    result = df_don.copy()
    result["__Trong_DM_BHYT__"] = result[col_ma_don].astype(str).str.strip().str.upper().isin(dm_keys)
    in_dm = result[result["__Trong_DM_BHYT__"]].copy()
    not_in_dm = result[~result["__Trong_DM_BHYT__"]].copy()
    not_in_dm["__Ghi_chú__"] = "❌ Không có trong danh mục BHYT"
    return in_dm, not_in_dm


def feature_ton_kho(df, col_ten, col_nhap, col_xuat, col_ton_dau=None, nguong_canh_bao=10):
    """Thống kê xuất nhập tồn kho, cảnh báo dưới mức tối thiểu"""
    result = df.copy()
    result["__Nhap__"] = pd.to_numeric(result[col_nhap], errors="coerce").fillna(0)
    result["__Xuat__"] = pd.to_numeric(result[col_xuat], errors="coerce").fillna(0)
    if col_ton_dau and col_ton_dau in result.columns:
        result["__TonDau__"] = pd.to_numeric(result[col_ton_dau], errors="coerce").fillna(0)
    else:
        result["__TonDau__"] = 0
    result["__TonCuoi__"] = result["__TonDau__"] + result["__Nhap__"] - result["__Xuat__"]
    result["__CanhBao__"] = result["__TonCuoi__"].apply(
        lambda x: "⚠️ Dưới mức tối thiểu" if x <= nguong_canh_bao else "✅ Đủ tồn kho"
    )
    canh_bao = result[result["__TonCuoi__"] <= nguong_canh_bao].copy()
    return result, canh_bao


# ===== NEW: NHÓM NHÂN SỰ / LƯƠNG =====

def feature_tinh_luong(df, col_he_so, col_luong_cb, col_phu_cap=None, col_ngay_cong=None, col_ngay_chuan=None):
    """Tính lương theo hệ số và lương cơ bản"""
    result = df.copy()
    result["__HeSo__"] = pd.to_numeric(result[col_he_so], errors="coerce").fillna(0)
    result["__LuongCB__"] = pd.to_numeric(result[col_luong_cb], errors="coerce").fillna(0)
    result["__LuongCoBan__"] = (result["__HeSo__"] * result["__LuongCB__"]).round(0)
    result["__PhuCap__"] = pd.to_numeric(result[col_phu_cap], errors="coerce").fillna(0) if col_phu_cap else 0
    # Tính theo ngày công thực tế
    if col_ngay_cong and col_ngay_chuan:
        result["__NgayCong__"] = pd.to_numeric(result[col_ngay_cong], errors="coerce").fillna(0)
        result["__NgayChuan__"] = pd.to_numeric(result[col_ngay_chuan], errors="coerce").fillna(26)
        result["__Luong_Thuc_te__"] = (result["__LuongCoBan__"] / result["__NgayChuan__"] * result["__NgayCong__"]).round(0)
    else:
        result["__Luong_Thuc_te__"] = result["__LuongCoBan__"]
    result["__Luong_Truoc_KT__"] = result["__Luong_Thuc_te__"] + result["__PhuCap__"]
    # Khấu trừ BHXH 8%, BHYT 1.5%, BHTN 1%
    result["__BHXH_KT__"] = (result["__LuongCoBan__"] * 0.08).round(0)
    result["__BHYT_KT__"] = (result["__LuongCoBan__"] * 0.015).round(0)
    result["__BHTN_KT__"] = (result["__LuongCoBan__"] * 0.01).round(0)
    result["__Tong_KT__"] = result["__BHXH_KT__"] + result["__BHYT_KT__"] + result["__BHTN_KT__"]
    result["__Luong_Thuc_Lanh__"] = (result["__Luong_Truoc_KT__"] - result["__Tong_KT__"]).round(0)
    return result


def feature_tong_hop_ngay_cong(df, col_ten, col_khoa, col_thang, col_ngay_cong):
    """Tổng hợp ngày công theo khoa / tháng"""
    result = df.copy()
    result["__NgayCong__"] = pd.to_numeric(result[col_ngay_cong], errors="coerce").fillna(0)
    pivot = result.groupby([col_khoa, col_thang])["__NgayCong__"].agg(
        Tổng_ngày_công="sum", Số_nhân_viên="count", TB_ngày_công="mean"
    ).round(1).reset_index()
    return pivot


def feature_bao_hiem_xa_hoi(df, col_ten, col_luong_cb, col_he_so):
    """Bảng tính và khấu trừ BHXH/BHYT/BHTN"""
    result = df.copy()
    result["__LuongCB__"] = pd.to_numeric(result[col_luong_cb], errors="coerce").fillna(0)
    result["__HeSo__"] = pd.to_numeric(result[col_he_so], errors="coerce").fillna(0)
    result["__Luong_dong_BH__"] = (result["__LuongCB__"] * result["__HeSo__"]).round(0)
    # Người lao động đóng
    result["NLĐ_BHXH (8%)"] = (result["__Luong_dong_BH__"] * 0.08).round(0)
    result["NLĐ_BHYT (1.5%)"] = (result["__Luong_dong_BH__"] * 0.015).round(0)
    result["NLĐ_BHTN (1%)"] = (result["__Luong_dong_BH__"] * 0.01).round(0)
    result["NLĐ_Tổng khấu trừ"] = result["NLĐ_BHXH (8%)"] + result["NLĐ_BHYT (1.5%)"] + result["NLĐ_BHTN (1%)"]
    # Đơn vị sử dụng lao động đóng
    result["ĐVSDLĐ_BHXH (17%)"] = (result["__Luong_dong_BH__"] * 0.17).round(0)
    result["ĐVSDLĐ_BHYT (3%)"] = (result["__Luong_dong_BH__"] * 0.03).round(0)
    result["ĐVSDLĐ_BHTN (1%)"] = (result["__Luong_dong_BH__"] * 0.01).round(0)
    result["ĐVSDLĐ_Tổng đóng"] = result["ĐVSDLĐ_BHXH (17%)"] + result["ĐVSDLĐ_BHYT (3%)"] + result["ĐVSDLĐ_BHTN (1%)"]
    return result


# ===== NEW: NHÓM KIỂM SOÁT / PHÁT HIỆN LỖI =====

def feature_validate_ma(df, col_ma, loai="the_bhyt"):
    """Kiểm tra định dạng mã thẻ BHYT / mã bệnh nhân / CMND"""
    patterns = {
        "the_bhyt": r'^[A-Z]{2}\d{13}$',          # VD: DN4012345678901
        "ma_benh_nhan": r'^\d{6,12}$',
        "cmnd_cccd": r'^\d{9}$|^\d{12}$',
        "ma_icd10": r'^[A-Z]\d{2}(\.\d{1,2})?$',  # VD: J06.9
        "so_dien_thoai": r'^(0|\+84)\d{8,9}$',
    }
    pattern = patterns.get(loai, r'.+')
    result = df.copy()
    series = result[col_ma].astype(str).str.strip()
    result["__Hop_le__"] = series.str.match(pattern, na=False)
    hop_le = result[result["__Hop_le__"]].copy()
    khong_hop_le = result[~result["__Hop_le__"]].copy()
    khong_hop_le["__Lỗi__"] = f"Sai định dạng {loai}"
    return hop_le, khong_hop_le


def feature_detect_anomaly_numbers(df, col_tien, cho_phep_am=False):
    """Phát hiện số âm bất thường, giá trị 0, giá trị quá lớn bất thường"""
    result = df.copy()
    result["__So__"] = pd.to_numeric(result[col_tien], errors="coerce")
    issues = []
    # Số âm
    if not cho_phep_am:
        so_am = result[result["__So__"] < 0].copy()
        so_am["__Loại_lỗi__"] = "❌ Số âm"
        issues.append(so_am)
    # Giá trị 0
    bang_0 = result[result["__So__"] == 0].copy()
    bang_0["__Loại_lỗi__"] = "⚠️ Giá trị bằng 0"
    issues.append(bang_0)
    # Không parse được (text lẫn)
    loi_parse = result[result["__So__"].isna() & (result[col_tien].astype(str).str.strip() != "")].copy()
    loi_parse["__Loại_lỗi__"] = "❌ Không phải số"
    issues.append(loi_parse)
    # Outlier (> 3 std)
    mean = result["__So__"].mean()
    std = result["__So__"].std()
    if std > 0:
        outliers = result[result["__So__"] > mean + 3 * std].copy()
        outliers["__Loại_lỗi__"] = f"⚠️ Bất thường (> {(mean + 3*std):,.0f})"
        issues.append(outliers)
    all_issues = pd.concat([i for i in issues if not i.empty], ignore_index=True) if issues else pd.DataFrame()
    return all_issues


def feature_validate_ngay(df, col_ngay):
    """Kiểm tra ngày tháng sai định dạng, ngày không hợp lệ, ngày trong tương lai"""
    result = df.copy()
    parsed = pd.to_datetime(result[col_ngay], errors="coerce", dayfirst=True)
    result["__NgayParsed__"] = parsed
    today = pd.Timestamp.today().normalize()
    sai_dinh_dang = result[result["__NgayParsed__"].isna() & (result[col_ngay].astype(str).str.strip() != "")].copy()
    sai_dinh_dang["__Lỗi__"] = "❌ Không parse được ngày"
    tuong_lai = result[result["__NgayParsed__"] > today].copy()
    tuong_lai["__Lỗi__"] = "⚠️ Ngày trong tương lai"
    qua_cu = result[result["__NgayParsed__"] < pd.Timestamp("1900-01-01")].copy()
    qua_cu["__Lỗi__"] = "❌ Ngày quá cũ (< 1900)"
    hop_le = result[result["__NgayParsed__"].notna() & (result["__NgayParsed__"] <= today) & (result["__NgayParsed__"] >= pd.Timestamp("1900-01-01"))].copy()
    return sai_dinh_dang, tuong_lai, qua_cu, hop_le


# ===== NEW: NHÓM XỬ LÝ & CHUYỂN ĐỔI =====

def feature_tach_cot(df, col_nguon, loai="ho_ten", sep=None):
    """Tách cột: họ tên, ngày tháng năm, hoặc theo ký tự phân cách"""
    result = df.copy()
    series = result[col_nguon].astype(str).str.strip()
    if loai == "ho_ten":
        parts = series.str.split(r'\s+', n=1, expand=True)
        result[col_nguon + "_Họ"] = parts[0].fillna("")
        result[col_nguon + "_Tên"] = parts[1].fillna("") if 1 in parts.columns else ""
    elif loai == "ngay_thang":
        parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
        result[col_nguon + "_Ngày"] = parsed.dt.day.astype("Int64").astype(str).replace("<NA>", "")
        result[col_nguon + "_Tháng"] = parsed.dt.month.astype("Int64").astype(str).replace("<NA>", "")
        result[col_nguon + "_Năm"] = parsed.dt.year.astype("Int64").astype(str).replace("<NA>", "")
    elif loai == "tu_chon" and sep:
        parts = series.str.split(re.escape(sep), expand=True)
        for i in range(len(parts.columns)):
            result[f"{col_nguon}_phần{i+1}"] = parts[i].fillna("")
    return result


def feature_chuyen_doi_ngay(df, col_ngay, dinh_dang_dau, dinh_dang_cuoi):
    """Chuyển đổi định dạng ngày tháng"""
    result = df.copy()
    parsed = pd.to_datetime(result[col_ngay], errors="coerce", dayfirst=(dinh_dang_dau.startswith("%d")))
    result[col_ngay + "_mới"] = parsed.dt.strftime(dinh_dang_cuoi).where(parsed.notna(), "")
    loi = result[parsed.isna() & (result[col_ngay].astype(str).str.strip() != "")].copy()
    return result, loi


def feature_gop_cot(df, cols, ten_cot_moi, ngan_cach=" "):
    """Gộp nhiều cột thành 1"""
    result = df.copy()
    result[ten_cot_moi] = result[cols].astype(str).apply(
        lambda row: ngan_cach.join([v.strip() for v in row if v.strip() and v.strip().lower() != "nan"]),
        axis=1
    )
    return result


def feature_pivot_nhanh(df, index_col, col_col, value_col, aggfunc="sum"):
    """Pivot table nhanh"""
    try:
        numeric_values = pd.to_numeric(df[value_col], errors="coerce")
        df_tmp = df.copy()
        df_tmp[value_col] = numeric_values
        pivot = pd.pivot_table(
            df_tmp, values=value_col, index=index_col,
            columns=col_col, aggfunc=aggfunc, fill_value=0
        ).round(0)
        pivot = pivot.reset_index()
        pivot.columns = [str(c) for c in pivot.columns]
        return pivot, None
    except Exception as e:
        return None, str(e)


# ====================== NHÓM PHÂN TÍCH NÂNG CAO ======================

def feature_trend_analysis(df, col_ngay, col_gia_tri, col_nhom=None, ky="M"):
    """Phân tích xu hướng theo thời gian, tính tốc độ tăng trưởng MoM/QoQ"""
    result = df.copy()
    parsed = pd.to_datetime(result[col_ngay], errors="coerce", dayfirst=True)
    result["__Date__"] = parsed
    result["__Val__"] = pd.to_numeric(result[col_gia_tri], errors="coerce").fillna(0)
    label_map = {"M": "Tháng", "Q": "Quý", "Y": "Năm"}
    result["__Ky__"] = result["__Date__"].dt.to_period(ky).astype(str)
    group_cols = ["__Ky__"]
    if col_nhom and col_nhom in result.columns:
        group_cols.append(col_nhom)
    agg = result.groupby(group_cols).agg(
        Tổng=("__Val__", "sum"),
        Số_hồ_sơ=("__Val__", "count"),
        Trung_bình=("__Val__", "mean"),
    ).round(0).reset_index()
    agg.rename(columns={"__Ky__": label_map.get(ky, "Kỳ")}, inplace=True)
    ky_col = label_map.get(ky, "Kỳ")
    if col_nhom is None:
        agg["Tăng_trưởng (%)"] = agg["Tổng"].pct_change().mul(100).round(2)
        agg["Tăng_trưởng (%)"] = agg["Tăng_trưởng (%)"].fillna(0)
    return agg, ky_col


def feature_auto_report_bhyt(df, col_ngay, col_khoa, col_benh, col_tien_bhyt, col_tien_tt):
    """Tạo báo cáo BHYT đa chiều tự động: theo khoa, theo nhóm bệnh, theo tháng"""
    r = df.copy()
    r["__BHYT__"] = pd.to_numeric(r[col_tien_bhyt], errors="coerce").fillna(0)
    r["__TT__"] = pd.to_numeric(r[col_tien_tt], errors="coerce").fillna(0)
    r["__Tong__"] = r["__BHYT__"] + r["__TT__"]
    # Parse date
    parsed = pd.to_datetime(r[col_ngay], errors="coerce", dayfirst=True)
    r["__Thang__"] = parsed.dt.to_period("M").astype(str)
    # 1. Theo khoa
    theo_khoa = r.groupby(col_khoa).agg(
        Số_hồ_sơ=("__Tong__", "count"),
        Tiền_BHYT=("__BHYT__", "sum"),
        Tiền_tự_trả=("__TT__", "sum"),
        Tổng_chi_phí=("__Tong__", "sum"),
        TB_chi_phí=("__Tong__", "mean"),
    ).round(0).reset_index()
    theo_khoa["Tỷ_lệ_BHYT (%)"] = (theo_khoa["Tiền_BHYT"] / theo_khoa["Tổng_chi_phí"] * 100).round(1)
    theo_khoa = theo_khoa.sort_values("Tổng_chi_phí", ascending=False)
    # 2. Theo bệnh
    theo_benh = r.groupby(col_benh).agg(
        Số_lượt=("__Tong__", "count"),
        Tiền_BHYT=("__BHYT__", "sum"),
        Tổng_chi_phí=("__Tong__", "sum"),
    ).round(0).reset_index()
    theo_benh = theo_benh.sort_values("Tổng_chi_phí", ascending=False).head(50)
    # 3. Theo tháng
    theo_thang = r.groupby("__Thang__").agg(
        Số_hồ_sơ=("__Tong__", "count"),
        Tiền_BHYT=("__BHYT__", "sum"),
        Tiền_tự_trả=("__TT__", "sum"),
        Tổng_chi_phí=("__Tong__", "sum"),
    ).round(0).reset_index().rename(columns={"__Thang__": "Tháng"})
    theo_thang["Tăng_trưởng (%)"] = theo_thang["Tổng_chi_phí"].pct_change().mul(100).round(2).fillna(0)
    # 4. Tổng quan
    tong_quan = pd.DataFrame([{
        "Chỉ tiêu": "Tổng hồ sơ", "Giá trị": f"{len(r):,}",
        "Chỉ tiêu 2": "Tổng tiền BHYT", "Giá trị 2": f"{r['__BHYT__'].sum():,.0f} đ",
    }, {
        "Chỉ tiêu": "Số khoa", "Giá trị": f"{r[col_khoa].nunique()}",
        "Chỉ tiêu 2": "Tổng tiền tự trả", "Giá trị 2": f"{r['__TT__'].sum():,.0f} đ",
    }, {
        "Chỉ tiêu": "Số nhóm bệnh", "Giá trị": f"{r[col_benh].nunique()}",
        "Chỉ tiêu 2": "Tổng chi phí", "Giá trị 2": f"{r['__Tong__'].sum():,.0f} đ",
    }, {
        "Chỉ tiêu": "Chi phí TB/hồ sơ", "Giá trị": f"{r['__Tong__'].mean():,.0f} đ",
        "Chỉ tiêu 2": "Tỷ lệ BHYT TB", "Giá trị 2": f"{(r['__BHYT__'].sum()/r['__Tong__'].sum()*100):.1f}%",
    }])
    return tong_quan, theo_khoa, theo_benh, theo_thang


def feature_find_replace_bulk(df, operations):
    """Tìm & thay thế hàng loạt: nhiều cột, nhiều cặp tìm/thay, hỗ trợ regex"""
    result = df.copy()
    log = []
    for op in operations:
        col = op.get("col")
        find = op.get("find", "")
        replace = op.get("replace", "")
        use_regex = op.get("regex", False)
        case_sensitive = op.get("case", False)
        if not col or col not in result.columns or not find:
            continue
        before = result[col].copy()
        try:
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                result[col] = result[col].astype(str).str.replace(find, replace, regex=True, flags=flags)
            else:
                if case_sensitive:
                    result[col] = result[col].astype(str).str.replace(find, replace, regex=False)
                else:
                    result[col] = result[col].astype(str).str.replace(
                        find, replace, regex=False, case=False
                    )
            changed = (before.astype(str) != result[col].astype(str)).sum()
            log.append({"Cột": col, "Tìm": find, "Thay bằng": replace, "Regex": use_regex, "Số dòng thay": changed})
        except Exception as e:
            log.append({"Cột": col, "Tìm": find, "Thay bằng": replace, "Regex": use_regex, "Số dòng thay": f"Lỗi: {e}"})
    return result, pd.DataFrame(log)


def feature_rank_top_n(df, group_col, value_col, n=10, agg_func="sum", ascending=False):
    """Xếp hạng và lấy Top-N theo nhóm, có tính tỷ trọng"""
    result = df.copy()
    result["__Val__"] = pd.to_numeric(result[value_col], errors="coerce").fillna(0)
    if group_col:
        grouped = result.groupby(group_col)["__Val__"].agg(agg_func).round(0).reset_index()
        grouped.columns = [group_col, f"{agg_func.upper()}_{value_col}"]
        val_col_name = f"{agg_func.upper()}_{value_col}"
    else:
        grouped = result[[value_col]].copy()
        grouped["__Val__"] = pd.to_numeric(grouped[value_col], errors="coerce").fillna(0)
        grouped = grouped.rename(columns={"__Val__": f"{agg_func.upper()}_{value_col}"})
        val_col_name = f"{agg_func.upper()}_{value_col}"
    grouped = grouped.sort_values(val_col_name, ascending=ascending).reset_index(drop=True)
    grouped.insert(0, "Hạng", range(1, len(grouped) + 1))
    total = grouped[val_col_name].sum()
    grouped["Tỷ_trọng (%)"] = (grouped[val_col_name] / total * 100).round(2) if total > 0 else 0
    grouped["Tỷ_trọng_cộng_dồn (%)"] = grouped["Tỷ_trọng (%)"].cumsum().round(2)
    top_n = grouped.head(n).copy()
    return grouped, top_n


def feature_conditional_flag(df, rules):
    """Gắn nhãn / phân loại dòng theo nhiều điều kiện kết hợp (IF-THEN logic)"""
    result = df.copy()
    result["__Nhãn__"] = "Không xác định"
    for rule in rules:
        conditions = rule.get("conditions", [])
        label = rule.get("label", "?")
        mask = pd.Series([True] * len(result), index=result.index)
        for cond in conditions:
            col = cond.get("col")
            op = cond.get("op")
            val = str(cond.get("val", ""))
            if col not in result.columns:
                continue
            s = result[col].astype(str).str.strip()
            if op == "=":
                mask &= s.str.upper() == val.upper()
            elif op == "≠":
                mask &= s.str.upper() != val.upper()
            elif op == "chứa":
                mask &= s.str.contains(val, case=False, na=False)
            elif op == ">":
                try:
                    mask &= pd.to_numeric(result[col], errors="coerce") > float(val)
                except Exception:
                    pass
            elif op == "<":
                try:
                    mask &= pd.to_numeric(result[col], errors="coerce") < float(val)
                except Exception:
                    pass
            elif op == "trống":
                mask &= (s == "") | (s.str.upper() == "NAN")
            elif op == "không trống":
                mask &= (s != "") & (s.str.upper() != "NAN")
        result.loc[mask, "__Nhãn__"] = label
    summary = result["__Nhãn__"].value_counts().reset_index()
    summary.columns = ["Nhãn", "Số dòng"]
    return result, summary


# ====================== MAIN TAB RENDER ======================

def render_excel_compare_tab():
    st.markdown("### 📊 Công cụ xử lý Excel nâng cao")

    # ---- UPLOAD ----
    st.markdown("#### 📁 Tải lên file Excel")
    col_u1, col_u2 = st.columns(2)
    with col_u1:
        file1 = st.file_uploader("📂 File 1 (File chính)", type=["xlsx", "xls"], key="ec_file1")
    with col_u2:
        file2 = st.file_uploader("📂 File 2 (File so sánh / tra cứu)", type=["xlsx", "xls"], key="ec_file2")
    col_u3, col_u4, col_u5 = st.columns(3)
    with col_u3:
        file3 = st.file_uploader("📂 File 3 (tùy chọn)", type=["xlsx"], key="ec_file3")
    with col_u4:
        file4 = st.file_uploader("📂 File 4 (tùy chọn)", type=["xlsx"], key="ec_file4")
    with col_u5:
        file5 = st.file_uploader("📂 File 5 (tùy chọn)", type=["xlsx"], key="ec_file5")

    dfs = {}
    for fname, fobj in {"File 1": file1, "File 2": file2, "File 3": file3, "File 4": file4, "File 5": file5}.items():
        if fobj:
            fobj.seek(0)
            sheets = get_sheet_names(fobj)
            fobj.seek(0)
            chosen = st.selectbox(f"Sheet của {fname}:", sheets, key=f"sheet_{fname}") if len(sheets) > 1 else sheets[0]
            fobj.seek(0)
            df, err = read_excel_file(fobj, sheet_name=chosen)
            if err:
                st.error(f"❌ Lỗi đọc {fname}: {err}")
            else:
                dfs[fname] = df
                st.success(f"✅ {fname}: **{fobj.name}** — {len(df):,} dòng × {len(df.columns)} cột")

    if not dfs:
        st.info("👆 Vui lòng tải lên ít nhất 1 file Excel để bắt đầu.")
        return

    st.divider()

    # ---- CHỌN NHÓM & TÍNH NĂNG ----
    st.markdown("#### ⚙️ Chọn tác vụ")

    nhom = st.selectbox("📂 Nhóm tác vụ:", [
        "🔵 Xử lý dữ liệu cơ bản",
        "💰 Tài chính & BHYT",
        "💊 Thuốc & Vật tư y tế",
        "👥 Nhân sự & Lương",
        "🔍 Kiểm soát & Phát hiện lỗi",
        "🔄 Chuyển đổi & Tái cấu trúc dữ liệu",
        "🔬 Phân tích nâng cao",
    ], key="ec_nhom")

    features_map = {
        "🔵 Xử lý dữ liệu cơ bản": [
            "1. So sánh 2 file — Tìm điểm khác biệt",
            "2. Tìm & lọc dữ liệu TRÙNG",
            "3. VLOOKUP — Ghép dữ liệu từ file tra cứu",
            "4. Lọc dữ liệu nâng cao (nhiều điều kiện)",
            "5. Kiểm tra chéo 2 cột (Cross-check)",
            "6. Thống kê nhóm (Group By)",
            "7. Kiểm tra giá trị trống",
            "8. Gộp nhiều file thành 1",
            "9. Chuẩn hóa dữ liệu văn bản",
        ],
        "💰 Tài chính & BHYT": [
            "10. Đối chiếu BHYT — Tự trả",
            "11. Kiểm tra vượt trần quỹ BHYT theo khoa",
            "12. Báo cáo tổng hợp theo tháng / quý",
            "13. Pivot nhanh — Tổng hợp đa chiều",
        ],
        "💊 Thuốc & Vật tư y tế": [
            "14. Kiểm tra hạn sử dụng thuốc / vật tư",
            "15. Đối chiếu danh mục BHYT được chi trả",
            "16. Thống kê xuất nhập TỒN KHO",
        ],
        "👥 Nhân sự & Lương": [
            "17. Tính lương theo hệ số & ngày công",
            "18. Tổng hợp ngày công theo khoa / tháng",
            "19. Bảng khấu trừ BHXH / BHYT / BHTN",
        ],
        "🔍 Kiểm soát & Phát hiện lỗi": [
            "20. Kiểm tra định dạng mã thẻ BHYT / mã BN",
            "21. Phát hiện số âm & giá trị bất thường",
            "22. Kiểm tra ngày tháng sai định dạng",
        ],
        "🔄 Chuyển đổi & Tái cấu trúc dữ liệu": [
            "23. Tách cột (họ tên / ngày tháng / ký tự phân cách)",
            "24. Chuyển đổi định dạng ngày tháng",
            "25. Gộp nhiều cột thành 1",
        ],
        "🔬 Phân tích nâng cao": [
            "26. Phân tích xu hướng & tăng trưởng theo thời gian",
            "27. Báo cáo BHYT tổng hợp đa chiều (tự động)",
            "28. Tìm & Thay thế hàng loạt (hỗ trợ Regex)",
            "29. Xếp hạng & Top-N theo nhóm (Pareto 80/20)",
            "30. Phân loại & gắn nhãn dữ liệu (IF-THEN đa điều kiện)",
        ],
    }

    feature = st.selectbox("🔧 Tác vụ:", features_map[nhom], key="ec_feature")
    st.divider()
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # ============================================================
    # NHÓM 1: CƠ BẢN
    # ============================================================

    if feature.startswith("1."):
        st.markdown("##### 🔍 So sánh 2 file — Tìm điểm khác biệt")
        if len(dfs) < 2:
            st.warning("⚠️ Cần tải lên ít nhất 2 file.")
            return
        fkeys = list(dfs.keys())
        c1, c2 = st.columns(2)
        src1 = c1.selectbox("File gốc:", fkeys, key="cmp_src1")
        src2 = c2.selectbox("File so sánh:", [k for k in fkeys if k != src1], key="cmp_src2")
        df1, df2 = dfs[src1], dfs[src2]
        c3, c4 = st.columns(2)
        key1 = c3.selectbox("Cột khóa (File 1):", df1.columns.tolist(), key="cmp_k1")
        key2 = c4.selectbox("Cột khóa (File 2):", df2.columns.tolist(), key="cmp_k2")
        st.caption("Chọn thêm cột cần so sánh nội dung (không bắt buộc):")
        c5, c6 = st.columns(2)
        cc1 = c5.multiselect("Cột so sánh (File 1):", [c for c in df1.columns if c != key1], key="cmp_c1")
        cc2 = c6.multiselect("Cột so sánh (File 2):", [c for c in df2.columns if c != key2], key="cmp_c2")
        if len(cc1) != len(cc2):
            st.warning("⚠️ Số cột so sánh ở 2 file phải bằng nhau.")
        if st.button("🚀 BẮT ĐẦU SO SÁNH", type="primary", use_container_width=True, key="btn_cmp"):
            with st.spinner("Đang so sánh..."):
                res = feature_compare_two_files(df1, df2, key1, key2, cc1, cc2)
            stat = res["stats"]
            cols_s = st.columns(6)
            for i, (k, v) in enumerate(stat.items()):
                cols_s[i].metric(k, f"{v:,}")
            with st.expander(f"📋 Chỉ có trong {src1} ({len(res['only_in_file1'])})", expanded=True):
                if not res["only_in_file1"].empty:
                    st.dataframe(res["only_in_file1"], use_container_width=True)
                    dl_btn("📥 Tải xuống", to_excel_bytes(res["only_in_file1"]), f"chi_trong_file1_{ts}.xlsx", "dl_o1")
                else:
                    st.success("Không có")
            with st.expander(f"📋 Chỉ có trong {src2} ({len(res['only_in_file2'])})", expanded=True):
                if not res["only_in_file2"].empty:
                    st.dataframe(res["only_in_file2"], use_container_width=True)
                    dl_btn("📥 Tải xuống", to_excel_bytes(res["only_in_file2"]), f"chi_trong_file2_{ts}.xlsx", "dl_o2")
                else:
                    st.success("Không có")
            if not res["differences"].empty:
                with st.expander(f"⚠️ Khác biệt nội dung ({len(res['differences'])})", expanded=True):
                    st.dataframe(res["differences"], use_container_width=True)
                    dl_btn("📥 Tải xuống", to_excel_bytes(res["differences"]), f"khac_biet_{ts}.xlsx", "dl_diff")
            all_s = {
                f"Chỉ trong {src1}": res["only_in_file1"] if not res["only_in_file1"].empty else pd.DataFrame({"": ["Không có"]}),
                f"Chỉ trong {src2}": res["only_in_file2"] if not res["only_in_file2"].empty else pd.DataFrame({"": ["Không có"]}),
                "Khác biệt nội dung": res["differences"] if not res["differences"].empty else pd.DataFrame({"": ["Không có"]}),
                "Thống kê": pd.DataFrame(list(stat.items()), columns=["Chỉ tiêu", "Giá trị"]),
            }
            dl_btn("📥 Báo cáo đầy đủ (Excel)", multi_sheet_excel(all_s), f"BAO_CAO_SS_{ts}.xlsx", "dl_full")

    elif feature.startswith("2."):
        st.markdown("##### 🔁 Tìm & lọc dữ liệu TRÙNG")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="dup_src")
        df = dfs[src]
        dup_cols = st.multiselect("Cột xét trùng:", df.columns.tolist(), default=[df.columns[0]], key="dup_cols")
        keep_opt = st.radio("Giữ lại dòng nào?", ["first — giữ dòng đầu", "last — giữ dòng cuối", "none — xóa tất cả trùng"], horizontal=True, key="dup_keep")
        keep_val = "first" if keep_opt.startswith("first") else ("last" if keep_opt.startswith("last") else False)
        if st.button("🚀 TÌM DỮ LIỆU TRÙNG", type="primary", use_container_width=True, key="btn_dup"):
            if not dup_cols:
                st.warning("Chọn ít nhất 1 cột.")
                return
            with st.spinner("Đang xử lý..."):
                duplicates, cleaned, removed = feature_find_duplicates(df, dup_cols, keep=keep_val)
            m1, m2, m3 = st.columns(3)
            m1.metric("🔢 Tổng dòng gốc", f"{len(df):,}")
            m2.metric("🔴 Dòng trùng", f"{len(duplicates):,}")
            m3.metric("✅ Sau khi lọc", f"{len(cleaned):,}")
            with st.expander(f"🔴 Danh sách trùng ({len(duplicates)} dòng)", expanded=True):
                if not duplicates.empty:
                    st.dataframe(duplicates, use_container_width=True)
                    dl_btn("📥 Tải danh sách trùng", to_excel_bytes(duplicates), f"trung_{ts}.xlsx", "dl_dup")
                else:
                    st.success("✅ Không có dữ liệu trùng!")
            with st.expander(f"✅ File đã loại trùng ({len(cleaned)} dòng)"):
                st.dataframe(cleaned, use_container_width=True)
                dl_btn("📥 Tải file đã lọc trùng", to_excel_bytes(cleaned), f"da_loc_trung_{ts}.xlsx", "dl_cl")

    elif feature.startswith("3."):
        st.markdown("##### 🔗 VLOOKUP — Ghép dữ liệu từ file tra cứu")
        if len(dfs) < 2:
            st.warning("⚠️ Cần tải lên ít nhất 2 file.")
            return
        fkeys = list(dfs.keys())
        c1, c2 = st.columns(2)
        main_src = c1.selectbox("File chính:", fkeys, key="vl_main")
        lk_src = c2.selectbox("File tra cứu:", [k for k in fkeys if k != main_src], key="vl_lk")
        df_main, df_lk = dfs[main_src], dfs[lk_src]
        c3, c4 = st.columns(2)
        mk = c3.selectbox("Cột khóa (File chính):", df_main.columns.tolist(), key="vl_mk")
        lk = c4.selectbox("Cột khóa (File tra cứu):", df_lk.columns.tolist(), key="vl_lkk")
        lk_cols = st.multiselect("Cột cần lấy từ file tra cứu:", [c for c in df_lk.columns if c != lk], key="vl_cols")
        if st.button("🚀 THỰC HIỆN VLOOKUP", type="primary", use_container_width=True, key="btn_vl"):
            if not lk_cols:
                st.warning("Chọn ít nhất 1 cột.")
                return
            with st.spinner("Đang ghép..."):
                result, not_found = feature_vlookup(df_main, df_lk, mk, lk, lk_cols)
            st.metric("✅ Ghép được", f"{len(df_main)-len(not_found):,} / {len(df_main):,}")
            if not_found:
                st.warning(f"⚠️ {len(not_found)} dòng không khớp")
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải kết quả VLOOKUP", to_excel_bytes(result), f"vlookup_{ts}.xlsx", "dl_vl")

    elif feature.startswith("4."):
        st.markdown("##### 🔍 Lọc dữ liệu nâng cao")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="flt_src")
        df = dfs[src]
        num_conds = st.number_input("Số điều kiện:", min_value=1, max_value=10, value=1, key="flt_num")
        ops = ["=", "≠", "chứa", "không chứa", "bắt đầu bằng", "trống", "không trống", ">", "<"]
        conditions = []
        for i in range(int(num_conds)):
            cc = st.columns([3, 2, 3])
            col_s = cc[0].selectbox(f"Cột #{i+1}:", df.columns.tolist(), key=f"fc_{i}")
            op_s = cc[1].selectbox("Toán tử:", ops, key=f"fo_{i}")
            val_s = cc[2].text_input("Giá trị:", key=f"fv_{i}") if op_s not in ["trống", "không trống"] else ""
            conditions.append({"col": col_s, "operator": op_s, "value": val_s})
        if st.button("🚀 LỌC DỮ LIỆU", type="primary", use_container_width=True, key="btn_flt"):
            with st.spinner("Đang lọc..."):
                filtered = feature_filter_advanced(df, conditions)
            st.metric("📊 Kết quả", f"{len(filtered):,} / {len(df):,} dòng")
            st.dataframe(filtered, use_container_width=True)
            if not filtered.empty:
                dl_btn("📥 Tải kết quả lọc", to_excel_bytes(filtered), f"loc_{ts}.xlsx", "dl_flt")

    elif feature.startswith("5."):
        st.markdown("##### ✅ Kiểm tra chéo 2 cột")
        if len(dfs) < 2:
            st.warning("⚠️ Cần 2 file.")
            return
        fkeys = list(dfs.keys())
        c1, c2 = st.columns(2)
        s1 = c1.selectbox("File 1:", fkeys, key="cc_s1")
        s2 = c2.selectbox("File 2:", [k for k in fkeys if k != s1], key="cc_s2")
        df1, df2 = dfs[s1], dfs[s2]
        cc1, cc2 = st.columns(2)
        col1 = cc1.selectbox("Cột kiểm tra (File 1):", df1.columns.tolist(), key="cc_c1")
        col2 = cc2.selectbox("Cột đối chiếu (File 2):", df2.columns.tolist(), key="cc_c2")
        if st.button("🚀 KIỂM TRA CHÉO", type="primary", use_container_width=True, key="btn_cc"):
            with st.spinner("Đang kiểm tra..."):
                missing_rows, m2, m1 = feature_cross_check(df1, df2, col1, col2)
            st.columns(2)[0].metric(f"Trong {s1} nhưng không có trong {s2}", f"{m2:,}")
            st.columns(2)[1].metric(f"Trong {s2} nhưng không có trong {s1}", f"{m1:,}")
            with st.expander("Xem dữ liệu không khớp", expanded=True):
                if not missing_rows.empty:
                    st.dataframe(missing_rows, use_container_width=True)
                    dl_btn("📥 Tải xuống", to_excel_bytes(missing_rows), f"cross_check_{ts}.xlsx", "dl_cc")
                else:
                    st.success("✅ Tất cả khớp!")

    elif feature.startswith("6."):
        st.markdown("##### 📈 Thống kê nhóm (Group By)")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="grp_src")
        df = dfs[src]
        c1, c2 = st.columns(2)
        grp_cols = c1.multiselect("Nhóm theo cột:", df.columns.tolist(), key="grp_col")
        agg_cols = c2.multiselect("Cột thống kê (số liệu):", [c for c in df.columns if c not in grp_cols], key="grp_agg")
        if st.button("🚀 THỐNG KÊ", type="primary", use_container_width=True, key="btn_grp"):
            if not grp_cols or not agg_cols:
                st.warning("Chọn ít nhất 1 cột nhóm và 1 cột thống kê.")
                return
            with st.spinner("Đang tính toán..."):
                result = feature_summary_stats(df, grp_cols, agg_cols)
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải bảng thống kê", to_excel_bytes(result), f"thong_ke_nhom_{ts}.xlsx", "dl_grp")

    elif feature.startswith("7."):
        st.markdown("##### 🔎 Kiểm tra giá trị trống")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="mv_src")
        df = dfs[src]
        if st.button("🚀 KIỂM TRA", type="primary", use_container_width=True, key="btn_mv"):
            with st.spinner("Đang kiểm tra..."):
                report = feature_check_missing_values(df)
            if report.empty:
                st.success("✅ Không có giá trị trống!")
            else:
                st.warning(f"⚠️ {len(report)} cột có giá trị trống")
                st.dataframe(report, use_container_width=True)
                prob_cols = report["Cột"].tolist()
                mask = df[prob_cols].apply(lambda s: (s.astype(str).str.strip() == "") | (s.astype(str).str.upper() == "NAN")).any(axis=1)
                df_empty = df[mask].copy()
                st.markdown(f"**Dòng có giá trị trống ({len(df_empty)}):**")
                st.dataframe(df_empty, use_container_width=True)
                dl_btn("📥 Tải báo cáo trống", multi_sheet_excel({"Báo cáo cột trống": report, "Dòng có trống": df_empty}), f"bao_cao_trong_{ts}.xlsx", "dl_mv")

    elif feature.startswith("8."):
        st.markdown("##### 📎 Gộp nhiều file thành 1")
        sel = st.multiselect("Chọn file cần gộp:", list(dfs.keys()), default=list(dfs.keys()), key="cat_files")
        add_src = st.checkbox("Thêm cột '__Nguồn__'", value=True, key="cat_src")
        if st.button("🚀 GỘP FILE", type="primary", use_container_width=True, key="btn_cat"):
            if len(sel) < 2:
                st.warning("Chọn ít nhất 2 file.")
                return
            with st.spinner("Đang gộp..."):
                result = feature_concat_files([dfs[k] for k in sel], add_source=add_src, source_names=sel)
            st.columns(2)[0].metric("Tổng dòng sau gộp", f"{len(result):,}")
            st.columns(2)[1].metric("Số cột", f"{len(result.columns)}")
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải file đã gộp", to_excel_bytes(result), f"gop_file_{ts}.xlsx", "dl_cat")

    elif feature.startswith("9."):
        st.markdown("##### 🧹 Chuẩn hóa dữ liệu văn bản")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="norm_src")
        df = dfs[src]
        norm_cols = st.multiselect("Cột cần chuẩn hóa:", df.columns.tolist(), key="norm_cols")
        c1, c2, c3, c4 = st.columns(4)
        do_upper = c1.checkbox("Viết HOA", key="n_upper")
        do_lower = c2.checkbox("Viết thường", key="n_lower")
        do_spaces = c3.checkbox("Xóa khoảng trắng thừa", value=True, key="n_sp")
        do_special = c4.checkbox("Xóa ký tự đặc biệt", key="n_spec")
        if st.button("🚀 CHUẨN HÓA", type="primary", use_container_width=True, key="btn_norm"):
            if not norm_cols:
                st.warning("Chọn ít nhất 1 cột.")
                return
            with st.spinner("Đang chuẩn hóa..."):
                result = feature_normalize_data(df, norm_cols, upper=do_upper, lower=do_lower, remove_spaces=do_spaces, remove_special=do_special)
            st.success(f"✅ Đã chuẩn hóa {len(norm_cols)} cột / {len(result):,} dòng")
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải file đã chuẩn hóa", to_excel_bytes(result), f"chuan_hoa_{ts}.xlsx", "dl_norm")

    # ============================================================
    # NHÓM 2: TÀI CHÍNH & BHYT
    # ============================================================

    elif feature.startswith("10."):
        st.markdown("##### 💰 Đối chiếu BHYT — Tự trả")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="bh_src")
        df = dfs[src]
        c1, c2, c3, c4 = st.columns(4)
        col_ma = c1.selectbox("Cột mã thẻ / mã BN:", df.columns.tolist(), key="bh_ma")
        col_bhyt = c2.selectbox("Cột tiền BHYT:", df.columns.tolist(), key="bh_bhyt")
        col_tutra = c3.selectbox("Cột tiền tự trả:", df.columns.tolist(), key="bh_tt")
        col_tong = c4.selectbox("Cột tổng (để kiểm tra, tùy chọn):", ["— Không —"] + df.columns.tolist(), key="bh_tong")
        col_tong_real = None if col_tong == "— Không —" else col_tong
        if st.button("🚀 ĐỐI CHIẾU", type="primary", use_container_width=True, key="btn_bh"):
            with st.spinner("Đang đối chiếu..."):
                result, issues, summary = feature_doi_chieu_bhyt(df, col_ma, col_bhyt, col_tutra, col_tong_real)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Tổng hồ sơ", f"{summary['Tổng bệnh nhân']:,}")
            m2.metric("Tiền BHYT", f"{summary['Tổng tiền BHYT']:,.0f}đ")
            m3.metric("Tiền tự trả", f"{summary['Tổng tiền tự trả']:,.0f}đ")
            m4.metric("Tổng cộng", f"{summary['Tổng cộng']:,.0f}đ")
            if col_tong_real and isinstance(issues, pd.DataFrame) and not issues.empty:
                st.error(f"⚠️ {len(issues)} dòng có chênh lệch tổng tiền!")
                st.dataframe(issues, use_container_width=True)
                dl_btn("📥 Tải dòng chênh lệch", to_excel_bytes(issues), f"chenh_lech_{ts}.xlsx", "dl_bh_err")
            else:
                if col_tong_real:
                    st.success("✅ Tất cả tổng tiền khớp!")
            st.dataframe(result[[ col_ma, col_bhyt, col_tutra, "__Tong_tinh__" ]].rename(columns={"__Tong_tinh__": "Tổng tính"}), use_container_width=True)
            dl_btn("📥 Tải kết quả đối chiếu", to_excel_bytes(result), f"doi_chieu_bhyt_{ts}.xlsx", "dl_bh")

    elif feature.startswith("11."):
        st.markdown("##### ⚠️ Kiểm tra vượt trần quỹ BHYT theo khoa")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="vt_src")
        df = dfs[src]
        c1, c2 = st.columns(2)
        col_khoa = c1.selectbox("Cột khoa / loại:", df.columns.tolist(), key="vt_khoa")
        col_tien = c2.selectbox("Cột tiền:", df.columns.tolist(), key="vt_tien")
        st.markdown("**Nhập ngưỡng trần cho từng khoa (tùy chọn):**")
        if df[col_khoa].nunique() < 30:
            khoa_list = df[col_khoa].dropna().unique().tolist()
            nguong_dict = {}
            cols_ng = st.columns(min(3, len(khoa_list)))
            for i, k in enumerate(khoa_list):
                val = cols_ng[i % 3].number_input(f"Trần: {k}", min_value=0, value=0, step=1000000, key=f"ng_{i}")
                if val > 0:
                    nguong_dict[k] = val
        else:
            nguong_dict = {}
            st.info("Quá nhiều khoa — bỏ qua ngưỡng trần.")
        if st.button("🚀 KIỂM TRA VƯỢT TRẦN", type="primary", use_container_width=True, key="btn_vt"):
            with st.spinner("Đang kiểm tra..."):
                result = feature_kiem_tra_vuot_tran(df, col_khoa, col_tien, nguong_dict if nguong_dict else None)
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải báo cáo vượt trần", to_excel_bytes(result), f"vuot_tran_{ts}.xlsx", "dl_vt")

    elif feature.startswith("12."):
        st.markdown("##### 📅 Báo cáo tổng hợp theo tháng / quý")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="bc_src")
        df = dfs[src]
        c1, c2, c3, c4 = st.columns(4)
        col_ngay = c1.selectbox("Cột ngày:", df.columns.tolist(), key="bc_ngay")
        col_bhyt = c2.selectbox("Cột tiền BHYT:", df.columns.tolist(), key="bc_bhyt")
        col_tt = c3.selectbox("Cột tiền tự trả:", df.columns.tolist(), key="bc_tt")
        col_khoa = c4.selectbox("Nhóm theo khoa (tùy chọn):", ["— Không —"] + df.columns.tolist(), key="bc_khoa")
        col_khoa_real = None if col_khoa == "— Không —" else col_khoa
        if st.button("🚀 TẠO BÁO CÁO", type="primary", use_container_width=True, key="btn_bc"):
            with st.spinner("Đang tổng hợp..."):
                bao_cao = feature_bao_cao_tong_hop(df, col_ngay, col_bhyt, col_tt, col_khoa_real)
            st.dataframe(bao_cao, use_container_width=True)
            dl_btn("📥 Tải báo cáo tổng hợp", to_excel_bytes(bao_cao), f"bao_cao_thang_{ts}.xlsx", "dl_bc")

    elif feature.startswith("13."):
        st.markdown("##### 🔄 Pivot nhanh — Tổng hợp đa chiều")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="pv_src")
        df = dfs[src]
        c1, c2, c3, c4 = st.columns(4)
        idx = c1.selectbox("Dòng (index):", df.columns.tolist(), key="pv_idx")
        col = c2.selectbox("Cột (columns):", df.columns.tolist(), key="pv_col")
        val = c3.selectbox("Giá trị:", df.columns.tolist(), key="pv_val")
        agg = c4.selectbox("Hàm tổng hợp:", ["sum", "count", "mean", "max", "min"], key="pv_agg")
        if st.button("🚀 TẠO PIVOT", type="primary", use_container_width=True, key="btn_pv"):
            with st.spinner("Đang tạo pivot..."):
                result, err = feature_pivot_nhanh(df, idx, col, val, agg)
            if err:
                st.error(f"Lỗi: {err}")
            else:
                st.dataframe(result, use_container_width=True)
                dl_btn("📥 Tải bảng Pivot", to_excel_bytes(result), f"pivot_{ts}.xlsx", "dl_pv")

    # ============================================================
    # NHÓM 3: THUỐC & VẬT TƯ
    # ============================================================

    elif feature.startswith("14."):
        st.markdown("##### 💊 Kiểm tra hạn sử dụng thuốc / vật tư")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="hsd_src")
        df = dfs[src]
        c1, c2, c3 = st.columns(3)
        col_ten = c1.selectbox("Cột tên thuốc / vật tư:", df.columns.tolist(), key="hsd_ten")
        col_han = c2.selectbox("Cột hạn sử dụng:", df.columns.tolist(), key="hsd_han")
        ngay_cb = c3.number_input("Cảnh báo trước (ngày):", min_value=7, max_value=365, value=90, key="hsd_ngay")
        if st.button("🚀 KIỂM TRA HẠN", type="primary", use_container_width=True, key="btn_hsd"):
            with st.spinner("Đang kiểm tra..."):
                da_het, sap_het, con_han = feature_kiem_tra_han_su_dung(df, col_ten, col_han, int(ngay_cb))
            m1, m2, m3 = st.columns(3)
            m1.metric("❌ Đã hết hạn", f"{len(da_het):,}")
            m2.metric(f"⚠️ Sắp hết hạn (≤{int(ngay_cb)} ngày)", f"{len(sap_het):,}")
            m3.metric("✅ Còn hạn", f"{len(con_han):,}")
            with st.expander(f"❌ Đã hết hạn ({len(da_het)} mục)", expanded=True):
                if not da_het.empty:
                    st.dataframe(da_het[[col_ten, col_han, "__ConLai__", "__Trạng_thái__"]], use_container_width=True)
                    dl_btn("📥 Tải danh sách hết hạn", to_excel_bytes(da_het), f"het_han_{ts}.xlsx", "dl_hh")
                else:
                    st.success("Không có mục hết hạn")
            with st.expander(f"⚠️ Sắp hết hạn ({len(sap_het)} mục)", expanded=True):
                if not sap_het.empty:
                    st.dataframe(sap_het[[col_ten, col_han, "__ConLai__", "__Trạng_thái__"]], use_container_width=True)
                    dl_btn("📥 Tải danh sách sắp hết", to_excel_bytes(sap_het), f"sap_het_han_{ts}.xlsx", "dl_shh")
                else:
                    st.success(f"Không có mục nào hết hạn trong {int(ngay_cb)} ngày tới")
            all_sheets = {
                "Đã hết hạn": da_het if not da_het.empty else pd.DataFrame({"": ["Không có"]}),
                "Sắp hết hạn": sap_het if not sap_het.empty else pd.DataFrame({"": ["Không có"]}),
                "Còn hạn": con_han if not con_han.empty else pd.DataFrame({"": ["Không có"]}),
            }
            dl_btn("📥 Báo cáo đầy đủ hạn sử dụng", multi_sheet_excel(all_sheets), f"han_sd_{ts}.xlsx", "dl_hsd_full")

    elif feature.startswith("15."):
        st.markdown("##### 📋 Đối chiếu danh mục BHYT được chi trả")
        if len(dfs) < 2:
            st.warning("⚠️ Cần 2 file: file đơn hàng/kê đơn và file danh mục BHYT.")
            return
        fkeys = list(dfs.keys())
        c1, c2 = st.columns(2)
        don_src = c1.selectbox("File kê đơn / sử dụng:", fkeys, key="dm_don")
        dm_src = c2.selectbox("File danh mục BHYT:", [k for k in fkeys if k != don_src], key="dm_dm")
        df_don, df_dm = dfs[don_src], dfs[dm_src]
        c3, c4 = st.columns(2)
        col_ma_don = c3.selectbox("Cột mã (đơn hàng):", df_don.columns.tolist(), key="dm_ma_don")
        col_ma_dm = c4.selectbox("Cột mã (danh mục BHYT):", df_dm.columns.tolist(), key="dm_ma_dm")
        if st.button("🚀 ĐỐI CHIẾU DANH MỤC", type="primary", use_container_width=True, key="btn_dm"):
            with st.spinner("Đang đối chiếu..."):
                in_dm, not_in_dm = feature_doi_chieu_danh_muc_bhyt(df_don, df_dm, col_ma_don, col_ma_dm)
            m1, m2 = st.columns(2)
            m1.metric("✅ Trong danh mục BHYT", f"{len(in_dm):,}")
            m2.metric("❌ Ngoài danh mục BHYT", f"{len(not_in_dm):,}")
            with st.expander(f"❌ Không có trong danh mục BHYT ({len(not_in_dm)} dòng)", expanded=True):
                if not not_in_dm.empty:
                    st.dataframe(not_in_dm, use_container_width=True)
                    dl_btn("📥 Tải danh sách ngoài DM", to_excel_bytes(not_in_dm), f"ngoai_dm_bhyt_{ts}.xlsx", "dl_ndm")
                else:
                    st.success("✅ Tất cả đều nằm trong danh mục BHYT!")
            sheets = {"Trong danh mục BHYT": in_dm, "Ngoài danh mục BHYT": not_in_dm}
            dl_btn("📥 Báo cáo đầy đủ", multi_sheet_excel(sheets), f"doi_chieu_dm_{ts}.xlsx", "dl_dm_full")

    elif feature.startswith("16."):
        st.markdown("##### 📦 Thống kê xuất nhập TỒN KHO")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="tk_src")
        df = dfs[src]
        c1, c2, c3, c4, c5 = st.columns(5)
        col_ten = c1.selectbox("Cột tên hàng:", df.columns.tolist(), key="tk_ten")
        col_nhap = c2.selectbox("Cột số nhập:", df.columns.tolist(), key="tk_nhap")
        col_xuat = c3.selectbox("Cột số xuất:", df.columns.tolist(), key="tk_xuat")
        col_ton_dau = c4.selectbox("Cột tồn đầu kỳ (tùy chọn):", ["— Không —"] + df.columns.tolist(), key="tk_ton")
        nguong = c5.number_input("Ngưỡng cảnh báo tồn tối thiểu:", min_value=0, value=10, key="tk_nguong")
        col_ton_real = None if col_ton_dau == "— Không —" else col_ton_dau
        if st.button("🚀 TÍNH TỒN KHO", type="primary", use_container_width=True, key="btn_tk"):
            with st.spinner("Đang tính..."):
                result, canh_bao = feature_ton_kho(df, col_ten, col_nhap, col_xuat, col_ton_real, int(nguong))
            m1, m2 = st.columns(2)
            m1.metric("Tổng mặt hàng", f"{len(result):,}")
            m2.metric(f"⚠️ Dưới mức tối thiểu (≤{int(nguong)})", f"{len(canh_bao):,}")
            st.dataframe(result[[col_ten, col_nhap, col_xuat, "__TonCuoi__", "__CanhBao__"]].rename(columns={"__TonCuoi__": "Tồn cuối kỳ", "__CanhBao__": "Trạng thái"}), use_container_width=True)
            if not canh_bao.empty:
                with st.expander(f"⚠️ Hàng cần nhập thêm ({len(canh_bao)} mặt hàng)", expanded=True):
                    st.dataframe(canh_bao[[col_ten, "__TonCuoi__", "__CanhBao__"]].rename(columns={"__TonCuoi__": "Tồn cuối kỳ", "__CanhBao__": "Trạng thái"}), use_container_width=True)
            dl_btn("📥 Tải báo cáo tồn kho", to_excel_bytes(result), f"ton_kho_{ts}.xlsx", "dl_tk")

    # ============================================================
    # NHÓM 4: NHÂN SỰ & LƯƠNG
    # ============================================================

    elif feature.startswith("17."):
        st.markdown("##### 💵 Tính lương theo hệ số & ngày công")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="lg_src")
        df = dfs[src]
        c1, c2, c3 = st.columns(3)
        col_hs = c1.selectbox("Cột hệ số lương:", df.columns.tolist(), key="lg_hs")
        col_cb = c2.selectbox("Cột lương cơ bản (đ):", df.columns.tolist(), key="lg_cb")
        col_pc = c3.selectbox("Cột phụ cấp (tùy chọn):", ["— Không —"] + df.columns.tolist(), key="lg_pc")
        c4, c5 = st.columns(2)
        col_nc = c4.selectbox("Cột ngày công thực tế (tùy chọn):", ["— Không —"] + df.columns.tolist(), key="lg_nc")
        col_nch = c5.selectbox("Cột ngày công chuẩn (tùy chọn):", ["— Không —"] + df.columns.tolist(), key="lg_nch")
        col_pc_r = None if col_pc == "— Không —" else col_pc
        col_nc_r = None if col_nc == "— Không —" else col_nc
        col_nch_r = None if col_nch == "— Không —" else col_nch
        if st.button("🚀 TÍNH LƯƠNG", type="primary", use_container_width=True, key="btn_lg"):
            with st.spinner("Đang tính lương..."):
                result = feature_tinh_luong(df, col_hs, col_cb, col_pc_r, col_nc_r, col_nch_r)
            tong_thuoc_linh = result["__Luong_Thuc_Lanh__"].astype(float).sum()
            st.metric("💰 Tổng lương thực lãnh", f"{tong_thuoc_linh:,.0f} đ")
            show_cols = [c for c in df.columns] + ["__LuongCoBan__", "__PhuCap__", "__Luong_Truoc_KT__", "__BHXH_KT__", "__BHYT_KT__", "__BHTN_KT__", "__Luong_Thuc_Lanh__"]
            show_cols = [c for c in show_cols if c in result.columns]
            display = result[show_cols].rename(columns={
                "__LuongCoBan__": "Lương cơ bản",
                "__PhuCap__": "Phụ cấp",
                "__Luong_Truoc_KT__": "Lương trước KT",
                "__BHXH_KT__": "KT BHXH (8%)",
                "__BHYT_KT__": "KT BHYT (1.5%)",
                "__BHTN_KT__": "KT BHTN (1%)",
                "__Luong_Thuc_Lanh__": "Lương thực lãnh"
            })
            st.dataframe(display, use_container_width=True)
            dl_btn("📥 Tải bảng lương", to_excel_bytes(display), f"bang_luong_{ts}.xlsx", "dl_lg")

    elif feature.startswith("18."):
        st.markdown("##### 📆 Tổng hợp ngày công theo khoa / tháng")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="nc_src")
        df = dfs[src]
        c1, c2, c3, c4 = st.columns(4)
        col_ten = c1.selectbox("Cột tên nhân viên:", df.columns.tolist(), key="nc_ten")
        col_khoa = c2.selectbox("Cột khoa / phòng:", df.columns.tolist(), key="nc_khoa")
        col_thang = c3.selectbox("Cột tháng / kỳ:", df.columns.tolist(), key="nc_thang")
        col_nc = c4.selectbox("Cột số ngày công:", df.columns.tolist(), key="nc_nc")
        if st.button("🚀 TỔNG HỢP NGÀY CÔNG", type="primary", use_container_width=True, key="btn_nc"):
            with st.spinner("Đang tổng hợp..."):
                result = feature_tong_hop_ngay_cong(df, col_ten, col_khoa, col_thang, col_nc)
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải bảng ngày công", to_excel_bytes(result), f"ngay_cong_{ts}.xlsx", "dl_nc")

    elif feature.startswith("19."):
        st.markdown("##### 🛡️ Bảng khấu trừ BHXH / BHYT / BHTN")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="bx_src")
        df = dfs[src]
        c1, c2, c3 = st.columns(3)
        col_ten = c1.selectbox("Cột tên nhân viên:", df.columns.tolist(), key="bx_ten")
        col_cb = c2.selectbox("Cột lương cơ bản (đ):", df.columns.tolist(), key="bx_cb")
        col_hs = c3.selectbox("Cột hệ số lương:", df.columns.tolist(), key="bx_hs")
        if st.button("🚀 TÍNH KHẤU TRỪ BẢO HIỂM", type="primary", use_container_width=True, key="btn_bx"):
            with st.spinner("Đang tính..."):
                result = feature_bao_hiem_xa_hoi(df, col_ten, col_cb, col_hs)
            tong_nld = result["NLĐ_Tổng khấu trừ"].astype(float).sum()
            tong_dv = result["ĐVSDLĐ_Tổng đóng"].astype(float).sum()
            m1, m2 = st.columns(2)
            m1.metric("Tổng NLĐ phải đóng", f"{tong_nld:,.0f} đ")
            m2.metric("Tổng đơn vị phải đóng", f"{tong_dv:,.0f} đ")
            st.dataframe(result.drop(columns=["__LuongCB__", "__HeSo__", "__Luong_dong_BH__"], errors="ignore"), use_container_width=True)
            dl_btn("📥 Tải bảng BHXH/BHYT/BHTN", to_excel_bytes(result), f"bao_hiem_{ts}.xlsx", "dl_bx")

    # ============================================================
    # NHÓM 5: KIỂM SOÁT & PHÁT HIỆN LỖI
    # ============================================================

    elif feature.startswith("20."):
        st.markdown("##### 🔐 Kiểm tra định dạng mã thẻ BHYT / mã BN")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="val_src")
        df = dfs[src]
        c1, c2 = st.columns(2)
        col_ma = c1.selectbox("Cột cần kiểm tra:", df.columns.tolist(), key="val_col")
        loai = c2.selectbox("Loại định dạng:", {
            "the_bhyt": "Thẻ BHYT (VD: DN4012345678901)",
            "ma_benh_nhan": "Mã bệnh nhân (6–12 chữ số)",
            "cmnd_cccd": "CMND/CCCD (9 hoặc 12 số)",
            "ma_icd10": "Mã ICD-10 (VD: J06.9)",
            "so_dien_thoai": "Số điện thoại VN",
        }, key="val_loai", format_func=lambda x: {
            "the_bhyt": "Thẻ BHYT (VD: DN4012345678901)",
            "ma_benh_nhan": "Mã bệnh nhân (6–12 chữ số)",
            "cmnd_cccd": "CMND/CCCD (9 hoặc 12 số)",
            "ma_icd10": "Mã ICD-10 (VD: J06.9)",
            "so_dien_thoai": "Số điện thoại VN",
        }[x])
        if st.button("🚀 KIỂM TRA ĐỊNH DẠNG", type="primary", use_container_width=True, key="btn_val"):
            with st.spinner("Đang kiểm tra..."):
                hop_le, khong_hop_le = feature_validate_ma(df, col_ma, loai)
            m1, m2 = st.columns(2)
            m1.metric("✅ Hợp lệ", f"{len(hop_le):,}")
            m2.metric("❌ Sai định dạng", f"{len(khong_hop_le):,}")
            if not khong_hop_le.empty:
                with st.expander(f"❌ Danh sách sai định dạng ({len(khong_hop_le)})", expanded=True):
                    st.dataframe(khong_hop_le, use_container_width=True)
                    dl_btn("📥 Tải danh sách sai", to_excel_bytes(khong_hop_le), f"sai_dinh_dang_{ts}.xlsx", "dl_val")
            else:
                st.success("✅ Tất cả dữ liệu đúng định dạng!")
            sheets = {"Hợp lệ": hop_le, "Sai định dạng": khong_hop_le if not khong_hop_le.empty else pd.DataFrame({"": ["Không có"]})}
            dl_btn("📥 Báo cáo đầy đủ", multi_sheet_excel(sheets), f"kiem_tra_ma_{ts}.xlsx", "dl_val_full")

    elif feature.startswith("21."):
        st.markdown("##### 🔢 Phát hiện số âm & giá trị bất thường")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="an_src")
        df = dfs[src]
        c1, c2 = st.columns(2)
        col_tien = c1.selectbox("Cột số tiền / số lượng:", df.columns.tolist(), key="an_col")
        cho_phep_am = c2.checkbox("Cho phép số âm (bình thường)", value=False, key="an_am")
        if st.button("🚀 PHÁT HIỆN BẤT THƯỜNG", type="primary", use_container_width=True, key="btn_an"):
            with st.spinner("Đang phân tích..."):
                issues = feature_detect_anomaly_numbers(df, col_tien, cho_phep_am)
            if issues.empty:
                st.success("✅ Không phát hiện giá trị bất thường!")
            else:
                st.error(f"⚠️ Phát hiện {len(issues)} dòng có vấn đề")
                st.dataframe(issues, use_container_width=True)
                dl_btn("📥 Tải danh sách bất thường", to_excel_bytes(issues, highlight_color="FFD7D7"), f"bat_thuong_{ts}.xlsx", "dl_an")

    elif feature.startswith("22."):
        st.markdown("##### 📅 Kiểm tra ngày tháng sai định dạng")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="dt_src")
        df = dfs[src]
        col_ngay = st.selectbox("Cột ngày cần kiểm tra:", df.columns.tolist(), key="dt_col")
        if st.button("🚀 KIỂM TRA NGÀY THÁNG", type="primary", use_container_width=True, key="btn_dt"):
            with st.spinner("Đang kiểm tra..."):
                sai, tuong_lai, qua_cu, hop_le = feature_validate_ngay(df, col_ngay)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("✅ Hợp lệ", f"{len(hop_le):,}")
            m2.metric("❌ Sai định dạng", f"{len(sai):,}")
            m3.metric("⚠️ Ngày tương lai", f"{len(tuong_lai):,}")
            m4.metric("❌ Ngày quá cũ", f"{len(qua_cu):,}")
            all_err = pd.concat([sai, tuong_lai, qua_cu], ignore_index=True) if any(not x.empty for x in [sai, tuong_lai, qua_cu]) else pd.DataFrame()
            if not all_err.empty:
                with st.expander("Xem dữ liệu lỗi", expanded=True):
                    st.dataframe(all_err, use_container_width=True)
                    dl_btn("📥 Tải dữ liệu lỗi ngày", to_excel_bytes(all_err), f"loi_ngay_{ts}.xlsx", "dl_dt")
            else:
                st.success("✅ Tất cả ngày tháng đều hợp lệ!")

    # ============================================================
    # NHÓM 6: CHUYỂN ĐỔI & TÁI CẤU TRÚC
    # ============================================================

    elif feature.startswith("23."):
        st.markdown("##### ✂️ Tách cột")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="tc_src")
        df = dfs[src]
        col_ng = st.selectbox("Cột nguồn cần tách:", df.columns.tolist(), key="tc_col")
        loai = st.radio("Loại tách:", ["ho_ten — Họ và Tên", "ngay_thang — Ngày / Tháng / Năm", "tu_chon — Ký tự phân cách tùy chọn"], horizontal=True, key="tc_loai")
        sep_val = ""
        if loai.startswith("tu_chon"):
            sep_val = st.text_input("Nhập ký tự phân cách (VD: '/', ',', '-'):", value="/", key="tc_sep")
        loai_code = loai.split(" ")[0]
        if st.button("🚀 TÁCH CỘT", type="primary", use_container_width=True, key="btn_tc"):
            with st.spinner("Đang tách..."):
                result = feature_tach_cot(df, col_ng, loai_code, sep_val if sep_val else None)
            new_cols = [c for c in result.columns if c not in df.columns]
            st.success(f"✅ Đã tạo {len(new_cols)} cột mới: {', '.join(new_cols)}")
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải file đã tách cột", to_excel_bytes(result), f"tach_cot_{ts}.xlsx", "dl_tc")

    elif feature.startswith("24."):
        st.markdown("##### 🔁 Chuyển đổi định dạng ngày tháng")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="cd_src")
        df = dfs[src]
        col_ng = st.selectbox("Cột ngày cần chuyển:", df.columns.tolist(), key="cd_col")
        fmt_options = {
            "%d/%m/%Y": "dd/mm/yyyy (VD: 25/12/2024)",
            "%Y-%m-%d": "yyyy-mm-dd (VD: 2024-12-25)",
            "%d-%m-%Y": "dd-mm-yyyy (VD: 25-12-2024)",
            "%Y/%m/%d": "yyyy/mm/dd (VD: 2024/12/25)",
            "%d/%m/%y": "dd/mm/yy (VD: 25/12/24)",
        }
        c1, c2 = st.columns(2)
        fmt_in = c1.selectbox("Định dạng đầu vào:", list(fmt_options.keys()), format_func=lambda x: fmt_options[x], key="cd_in")
        fmt_out = c2.selectbox("Định dạng đầu ra:", list(fmt_options.keys()), format_func=lambda x: fmt_options[x], key="cd_out", index=1)
        if st.button("🚀 CHUYỂN ĐỔI NGÀY", type="primary", use_container_width=True, key="btn_cd"):
            with st.spinner("Đang chuyển đổi..."):
                result, loi = feature_chuyen_doi_ngay(df, col_ng, fmt_in, fmt_out)
            thanh_cong = len(df) - len(loi)
            m1, m2 = st.columns(2)
            m1.metric("✅ Chuyển thành công", f"{thanh_cong:,}")
            m2.metric("❌ Lỗi / không parse được", f"{len(loi):,}")
            st.dataframe(result[[col_ng, col_ng + "_mới"]], use_container_width=True)
            dl_btn("📥 Tải file đã chuyển đổi", to_excel_bytes(result), f"chuyen_ngay_{ts}.xlsx", "dl_cd")
            if not loi.empty:
                with st.expander(f"Xem {len(loi)} dòng lỗi"):
                    st.dataframe(loi, use_container_width=True)

    elif feature.startswith("25."):
        st.markdown("##### 🔗 Gộp nhiều cột thành 1")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="gc_src")
        df = dfs[src]
        cols_gop = st.multiselect("Chọn các cột cần gộp (theo thứ tự):", df.columns.tolist(), key="gc_cols")
        c1, c2 = st.columns(2)
        ten_moi = c1.text_input("Tên cột mới:", value="Cột_gộp", key="gc_ten")
        ngan_cach = c2.text_input("Ký tự ngăn cách:", value=" ", key="gc_sep")
        if st.button("🚀 GỘP CỘT", type="primary", use_container_width=True, key="btn_gc"):
            if not cols_gop:
                st.warning("Chọn ít nhất 2 cột để gộp.")
                return
            with st.spinner("Đang gộp..."):
                result = feature_gop_cot(df, cols_gop, ten_moi, ngan_cach)
            st.success(f"✅ Đã tạo cột '{ten_moi}'")
            st.dataframe(result[cols_gop + [ten_moi]], use_container_width=True)
            dl_btn("📥 Tải file đã gộp cột", to_excel_bytes(result), f"gop_cot_{ts}.xlsx", "dl_gc")

    # ============================================================
    # NHÓM 7: PHÂN TÍCH NÂNG CAO
    # ============================================================

    elif feature.startswith("26."):
        st.markdown("##### 📈 Phân tích xu hướng & tăng trưởng theo thời gian")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="tr_src")
        df = dfs[src]
        c1, c2, c3, c4 = st.columns(4)
        col_ngay = c1.selectbox("Cột ngày:", df.columns.tolist(), key="tr_ngay")
        col_val = c2.selectbox("Cột giá trị (số):", df.columns.tolist(), key="tr_val")
        col_nhom = c3.selectbox("Nhóm theo (tùy chọn):", ["— Không —"] + df.columns.tolist(), key="tr_nhom")
        ky = c4.selectbox("Kỳ phân tích:", {"M": "Tháng", "Q": "Quý", "Y": "Năm"},
                          format_func=lambda x: {"M": "📅 Tháng", "Q": "📅 Quý", "Y": "📅 Năm"}[x], key="tr_ky")
        col_nhom_real = None if col_nhom == "— Không —" else col_nhom
        if st.button("🚀 PHÂN TÍCH XU HƯỚNG", type="primary", use_container_width=True, key="btn_tr"):
            with st.spinner("Đang phân tích..."):
                trend, ky_col = feature_trend_analysis(df, col_ngay, col_val, col_nhom_real, ky)
            st.dataframe(trend, use_container_width=True)
            # Chart
            if col_nhom_real is None:
                chart_data = trend.set_index(ky_col)[["Tổng"]].copy()
                st.subheader("📊 Biểu đồ xu hướng tổng chi phí")
                st.bar_chart(chart_data)
                growth = trend[["Tăng_trưởng (%)"]].copy()
                st.subheader("📈 Tốc độ tăng trưởng (%)")
                st.line_chart(growth)
            else:
                pivot_chart = trend.pivot_table(index=ky_col, columns=col_nhom_real, values="Tổng", fill_value=0)
                st.subheader(f"📊 Biểu đồ theo {col_nhom_real}")
                st.bar_chart(pivot_chart)
            dl_btn("📥 Tải bảng xu hướng", to_excel_bytes(trend), f"xu_huong_{ts}.xlsx", "dl_tr")

    elif feature.startswith("27."):
        st.markdown("##### 📋 Báo cáo BHYT tổng hợp đa chiều (tự động)")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="ar_src")
        df = dfs[src]
        st.caption("Chọn các cột tương ứng trong file dữ liệu:")
        c1, c2, c3 = st.columns(3)
        col_ngay = c1.selectbox("Cột ngày KCB:", df.columns.tolist(), key="ar_ngay")
        col_khoa = c2.selectbox("Cột khoa/phòng:", df.columns.tolist(), key="ar_khoa")
        col_benh = c3.selectbox("Cột mã bệnh / nhóm bệnh:", df.columns.tolist(), key="ar_benh")
        c4, c5 = st.columns(2)
        col_bhyt = c4.selectbox("Cột tiền BHYT:", df.columns.tolist(), key="ar_bhyt")
        col_tt = c5.selectbox("Cột tiền tự trả:", df.columns.tolist(), key="ar_tt")
        if st.button("🚀 TẠO BÁO CÁO ĐA CHIỀU", type="primary", use_container_width=True, key="btn_ar"):
            with st.spinner("Đang tổng hợp báo cáo..."):
                tq, theo_khoa, theo_benh, theo_thang = feature_auto_report_bhyt(
                    df, col_ngay, col_khoa, col_benh, col_bhyt, col_tt)
            st.subheader("📌 Tổng quan")
            st.dataframe(tq, use_container_width=True, hide_index=True)
            st.subheader("🏥 Theo khoa / phòng")
            st.dataframe(theo_khoa, use_container_width=True, hide_index=True)
            chart_khoa = theo_khoa.set_index(col_khoa)[["Tổng_chi_phí"]].head(15)
            st.bar_chart(chart_khoa)
            st.subheader("🗓️ Theo tháng")
            st.dataframe(theo_thang, use_container_width=True, hide_index=True)
            chart_thang = theo_thang.set_index("Tháng")[["Tiền_BHYT", "Tiền_tự_trả"]]
            st.bar_chart(chart_thang)
            st.subheader("🦠 Top 50 nhóm bệnh chi phí cao nhất")
            st.dataframe(theo_benh, use_container_width=True, hide_index=True)
            all_sheets = {
                "Tổng quan": tq,
                "Theo khoa": theo_khoa,
                "Theo nhóm bệnh": theo_benh,
                "Theo tháng": theo_thang,
            }
            dl_btn("📥 Tải báo cáo đa chiều (Excel)", multi_sheet_excel(all_sheets), f"BC_BHYT_DA_CHIEU_{ts}.xlsx", "dl_ar")

    elif feature.startswith("28."):
        st.markdown("##### 🔁 Tìm & Thay thế hàng loạt (hỗ trợ Regex)")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="fr_src")
        df = dfs[src]
        num_ops = st.number_input("Số phép thay thế:", min_value=1, max_value=20, value=1, key="fr_num")
        operations = []
        for i in range(int(num_ops)):
            st.markdown(f"**Phép #{i+1}**")
            cc = st.columns([3, 3, 3, 1, 1])
            col_s = cc[0].selectbox(f"Cột #{i+1}:", df.columns.tolist(), key=f"fr_col_{i}")
            find_s = cc[1].text_input(f"Tìm:", key=f"fr_find_{i}")
            repl_s = cc[2].text_input(f"Thay bằng:", key=f"fr_repl_{i}")
            is_regex = cc[3].checkbox("Regex", key=f"fr_rx_{i}")
            is_case = cc[4].checkbox("A≠a", key=f"fr_cs_{i}")
            operations.append({"col": col_s, "find": find_s, "replace": repl_s, "regex": is_regex, "case": is_case})
        if st.button("🚀 THỰC HIỆN THAY THẾ", type="primary", use_container_width=True, key="btn_fr"):
            with st.spinner("Đang xử lý..."):
                result, log = feature_find_replace_bulk(df, operations)
            if not log.empty:
                total_changed = log["Số dòng thay"].apply(lambda x: x if isinstance(x, int) else 0).sum()
                st.success(f"✅ Hoàn thành — tổng {total_changed:,} ô được thay thế")
                st.dataframe(log, use_container_width=True, hide_index=True)
            st.dataframe(result, use_container_width=True)
            dl_btn("📥 Tải file đã thay thế", to_excel_bytes(result), f"find_replace_{ts}.xlsx", "dl_fr")

    elif feature.startswith("29."):
        st.markdown("##### 🏆 Xếp hạng & Top-N theo nhóm (Pareto 80/20)")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="rk_src")
        df = dfs[src]
        c1, c2, c3, c4, c5 = st.columns(5)
        group_col = c1.selectbox("Nhóm theo cột:", ["— Không nhóm —"] + df.columns.tolist(), key="rk_group")
        val_col = c2.selectbox("Cột giá trị:", df.columns.tolist(), key="rk_val")
        agg_fn = c3.selectbox("Hàm:", ["sum", "count", "mean", "max"], key="rk_agg")
        top_n = c4.number_input("Top N:", min_value=5, max_value=500, value=20, key="rk_n")
        order = c5.radio("Sắp xếp:", ["Cao → Thấp", "Thấp → Cao"], key="rk_order")
        group_col_real = None if group_col == "— Không nhóm —" else group_col
        asc = order == "Thấp → Cao"
        if st.button("🚀 XẾP HẠNG", type="primary", use_container_width=True, key="btn_rk"):
            with st.spinner("Đang xếp hạng..."):
                full, top = feature_rank_top_n(df, group_col_real, val_col, int(top_n), agg_fn, ascending=asc)
            val_col_name = f"{agg_fn.upper()}_{val_col}"
            m1, m2, m3 = st.columns(3)
            m1.metric("Tổng nhóm", f"{len(full):,}")
            top_pct = top["Tỷ_trọng (%)"].sum()
            m2.metric(f"Top {int(top_n)} chiếm", f"{top_pct:.1f}% tổng giá trị")
            pareto_n = (full["Tỷ_trọng_cộng_dồn (%)"] <= 80).sum()
            m3.metric("Nhóm đạt 80% giá trị (Pareto)", f"{pareto_n}")
            st.subheader(f"🏆 Top {int(top_n)}")
            st.dataframe(top, use_container_width=True, hide_index=True)
            if group_col_real:
                chart_top = top.set_index(group_col_real)[[val_col_name]].head(20)
                st.bar_chart(chart_top)
            st.subheader("📋 Bảng xếp hạng đầy đủ")
            st.dataframe(full, use_container_width=True, hide_index=True)
            sheets = {f"Top {int(top_n)}": top, "Xếp hạng đầy đủ": full}
            dl_btn("📥 Tải bảng xếp hạng", multi_sheet_excel(sheets), f"xep_hang_{ts}.xlsx", "dl_rk")

    elif feature.startswith("30."):
        st.markdown("##### 🏷️ Phân loại & gắn nhãn dữ liệu (IF-THEN đa điều kiện)")
        src = st.selectbox("Chọn file:", list(dfs.keys()), key="fl_src")
        df = dfs[src]
        st.caption("Định nghĩa các quy tắc phân loại (ưu tiên từ trên xuống dưới):")
        num_rules = st.number_input("Số quy tắc:", min_value=1, max_value=10, value=2, key="fl_num")
        rules = []
        ops_list = ["=", "≠", "chứa", ">", "<", "trống", "không trống"]
        for r_i in range(int(num_rules)):
            with st.expander(f"📌 Quy tắc #{r_i+1}", expanded=(r_i == 0)):
                label = st.text_input(f"Nhãn kết quả #{r_i+1}:", value=f"Loại {r_i+1}", key=f"fl_lbl_{r_i}")
                num_conds = st.number_input(f"Số điều kiện trong quy tắc #{r_i+1}:", min_value=1, max_value=5, value=1, key=f"fl_nc_{r_i}")
                conditions = []
                for c_i in range(int(num_conds)):
                    cc = st.columns([3, 2, 3])
                    c_col = cc[0].selectbox(f"Cột:", df.columns.tolist(), key=f"fl_col_{r_i}_{c_i}")
                    c_op = cc[1].selectbox(f"Điều kiện:", ops_list, key=f"fl_op_{r_i}_{c_i}")
                    c_val = cc[2].text_input(f"Giá trị:", key=f"fl_val_{r_i}_{c_i}") if c_op not in ["trống", "không trống"] else ""
                    conditions.append({"col": c_col, "op": c_op, "val": c_val})
                rules.append({"label": label, "conditions": conditions})
        if st.button("🚀 GẮN NHÃN", type="primary", use_container_width=True, key="btn_fl"):
            with st.spinner("Đang phân loại..."):
                result, summary = feature_conditional_flag(df, rules)
            st.subheader("📊 Kết quả phân loại")
            st.dataframe(summary, use_container_width=True, hide_index=True)
            chart_data = summary.set_index("Nhãn")
            st.bar_chart(chart_data)
            st.subheader("📋 Dữ liệu đã gắn nhãn")
            label_col_name = "__Nhãn__"
            filter_label = st.selectbox("Lọc theo nhãn:", ["— Tất cả —"] + result[label_col_name].unique().tolist(), key="fl_filter")
            display_df = result if filter_label == "— Tất cả —" else result[result[label_col_name] == filter_label]
            st.dataframe(display_df, use_container_width=True)
            result_renamed = result.rename(columns={"__Nhãn__": "Phân_loại"})
            sheets = {"Dữ liệu gắn nhãn": result_renamed, "Thống kê phân loại": summary}
            dl_btn("📥 Tải kết quả phân loại", multi_sheet_excel(sheets), f"phan_loai_{ts}.xlsx", "dl_fl")
